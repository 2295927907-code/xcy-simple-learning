from __future__ import annotations

import argparse
import csv
import io
import json
import math
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog, ttk
except Exception:  # pragma: no cover - makes core tests usable without Tk.
    tk = None
    filedialog = None
    messagebox = None
    simpledialog = None
    ttk = None


APP_DIR = Path(__file__).resolve().parent
DEFAULT_REFERENCE_GLOB = "*AM1.5*.xlsx"
DEFAULT_REFERENCE_KIND = "AM1.5G"
DEFAULT_REFERENCE_CHOICES = ("AM1.5G", "AM1.5D")
DEFAULT_SR_GLOBS = ("1027SR.xlsx", "1027SR*.xlsx", "1025SR.xlsx", "1025SR*.xlsx", "*光谱响应*.xlsx")
INPUT_FILE_FORMAT = "spectral_spd_input"
INPUT_FILE_VERSION = 1
EXCEL_TABLE_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
TEXT_TABLE_SUFFIXES = {".csv", ".txt", ".tsv", ".asd"}
TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "gbk", "gb18030")
DEFAULT_REFERENCE_TEMPERATURE_C = 25.0
DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C = 0.04
CURRENT_CORRECTION_START_NM = 300.0
CURRENT_CORRECTION_END_NM = 1200.0
IEC_A5_MIN_VALID_CV_COUNT = 5
IEC_A5_CV_DEVIATION_LIMIT_PERCENT = 1.5
IEC_A5_ISC_RANGE_LIMIT_PERCENT = 1.5
IEC_A5_CV_STD_LIMIT_PERCENT = 1.0

IEC_BANDS = [
    (300.0, 470.0),
    (470.0, 561.0),
    (561.0, 657.0),
    (657.0, 772.0),
    (772.0, 919.0),
    (919.0, 1200.0),
]

CLASS_LIMITS = [
    ("A+", 0.875, 1.125),
    ("A", 0.750, 1.250),
    ("B", 0.600, 1.400),
    ("C", 0.400, 2.000),
]

UNIT_FACTORS = {
    "W/(m^2 nm)": 1.0,
    "W/(m^2 um)": 1.0 / 1000.0,
    "W/(cm^2 nm)": 10000.0,
    "mW/(cm^2 nm)": 10.0,
    "uW/(cm^2 nm)": 0.01,
    "relative / a.u.": 1.0,
}
REFERENCE_SERIES_COLOR = "#2563eb"
TEST_SERIES_COLORS = (
    "#e11d48",
    "#0f766e",
    "#7c3aed",
    "#f59e0b",
    "#0891b2",
    "#be123c",
    "#4d7c0f",
    "#9333ea",
    "#dc2626",
    "#0284c7",
    "#16a34a",
    "#ca8a04",
    "#c026d3",
    "#ea580c",
    "#0d9488",
    "#475569",
    "#db2777",
    "#65a30d",
    "#1d4ed8",
    "#7e22ce",
)
TEST_SERIES_DASH_PATTERNS = (
    (),
    (9, 5),
    (2, 4),
    (9, 4, 2, 4),
)
TEST_SERIES_LINE_NAMES = ("实线", "虚线", "点线", "点划线")
VISIBLE_CHECKED_TEXT = "☑"
VISIBLE_UNCHECKED_TEXT = "☐"
FADED_SERIES_COLOR = "#cbd5e1"
FADED_LEGEND_TEXT_COLOR = "#94a3b8"
MIN_PLOT_AREA_WIDTH = 360
UI_FONT_FAMILY = "Microsoft YaHei UI"
UI_PAGE_BG = "#f4f6f8"
UI_PANEL_BG = "#ffffff"
UI_BORDER = "#d8dee8"
UI_BORDER_STRONG = "#b9c3d0"
UI_TEXT = "#161b22"
UI_TEXT_MUTED = "#5f6f83"
UI_TEXT_SUBTLE = "#6b7280"
UI_PRIMARY = "#0f62fe"
UI_PRIMARY_HOVER = "#0353e9"
UI_PRIMARY_PRESSED = "#002d9c"
UI_SECONDARY_BG = "#eef2f7"
UI_SECONDARY_HOVER = "#e2e8f0"
UI_GHOST_HOVER = "#f1f5f9"
UI_DANGER = "#da1e28"
UI_DANGER_BG = "#fff1f1"
UI_DANGER_HOVER = "#ffd7d9"
UI_SUCCESS = "#198038"
UI_WARNING = "#b7791f"
UI_GRID = "#edf1f7"
UI_GRID_MAJOR = "#dfe5ee"
UI_AXIS = "#8b99ab"
UI_FOCUS = "#0f62fe"


@dataclass
class Spectrum:
    wavelength_nm: np.ndarray
    irradiance_w_m2_nm: np.ndarray


@dataclass
class SpectralResponse:
    wavelength_nm: np.ndarray
    response_a_w: np.ndarray


@dataclass
class TestSpectrumDataset:
    label: str
    name: str
    path: Path | None
    sheet_name: str
    df: pd.DataFrame
    wavelength_column: str
    irradiance_column: str
    unit_name: str


@dataclass
class IVCurve:
    source_name: str
    sheet_name: str
    voltage_column: str
    current_column: str
    header_row: int | None
    first_data_row: int | None
    voltage_v: np.ndarray
    current_ma: np.ndarray
    isc_ma: float
    label: str = ""


@dataclass
class CurrentInput:
    label: str
    isc_ma: float
    spectrum_label: str
    temperature_c: float
    source: str
    iv_curve: IVCurve | None = None


@dataclass
class CalculationResult:
    reference_name: str
    test_name: str
    start_nm: float
    end_nm: float
    step_nm: float
    normalize: bool
    scale_factor: float
    grid_nm: np.ndarray
    reference: np.ndarray
    test_raw: np.ndarray
    test_scaled: np.ndarray
    absolute_error: np.ndarray
    reference_total: float
    test_raw_total: float
    test_scaled_total: float
    spd_percent: float
    spc_percent: float
    overall_class: str
    band_rows: list[dict[str, object]]


@dataclass
class CurrentCorrectionResult:
    sr_name: str
    sr_sheet_name: str
    sr_wavelength_column: str
    sr_column: str
    measured_isc: float
    test_temperature_c: float
    temperature_coefficient_percent_per_c: float
    reference_temperature_c: float
    grid_nm: np.ndarray
    sr: np.ndarray
    reference: np.ndarray
    test: np.ndarray
    reference_sr_weighted: np.ndarray
    test_sr_weighted: np.ndarray
    reference_irradiance_integral: float
    test_irradiance_integral: float
    reference_sr_integral: float
    test_sr_integral: float
    mg: float
    mt: float
    mmf: float
    corrected_cv: float


@dataclass
class CurrentCorrectionEvaluation:
    label: str
    spectrum_label: str
    spectrum_name: str
    source: str
    correction: CurrentCorrectionResult
    deviation_percent: float
    valid: bool
    reasons: list[str]


@dataclass(frozen=True)
class PlotLegendLayout:
    font_size: int
    row_step: float
    rows_per_column: int
    column_count: int
    column_width: float
    panel_width: float
    panel_gap: float
    line_length: float


@dataclass(frozen=True)
class PlotSeriesStyle:
    color: str
    dash: tuple[int, ...] = ()


@dataclass(frozen=True)
class PlotSeriesData:
    y_values: np.ndarray
    color: str
    label: str
    dash: tuple[int, ...] = ()
    spectrum_label: str = ""
    is_reference: bool = False


def find_default_reference_file() -> Path | None:
    matches = sorted(APP_DIR.glob(DEFAULT_REFERENCE_GLOB))
    return matches[0] if matches else None


def find_default_sr_file() -> Path | None:
    for pattern in DEFAULT_SR_GLOBS:
        matches = sorted(APP_DIR.glob(pattern))
        if matches:
            return matches[0]
    return None


def list_excel_sheets(path: Path) -> list[str]:
    return pd.ExcelFile(path).sheet_names


def read_table(path: Path, sheet_name: str | int | None = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in EXCEL_TABLE_SUFFIXES:
        return pd.read_excel(path, sheet_name=sheet_name or 0)
    if suffix in TEXT_TABLE_SUFFIXES:
        if suffix == ".tsv":
            return _read_delimited(path, sep="\t")
        return _read_delimited(path, sep=None)
    raise ValueError(f"暂不支持该文件类型：{path.suffix}")


def _read_delimited(path: Path, sep: str | None) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in TEXT_ENCODINGS:
        try:
            frame = pd.read_csv(path, sep=sep, engine="python", encoding=encoding)
            frame = _clean_text_frame(frame)
            if _is_usable_text_frame(frame):
                return frame
            last_error = ValueError("文本表格少于两列可用数字数据。")
        except Exception as exc:
            last_error = exc

    text = _read_text_with_fallback(path)
    delimiter_names = ["tab"] if sep == "\t" else _text_delimiter_candidates(path)
    best_frame: pd.DataFrame | None = None
    best_score: tuple[int, int, int, int] | None = None
    parse_errors: list[str] = []
    for delimiter_name in delimiter_names:
        try:
            frame, score = _dataframe_from_text(text, delimiter_name)
        except Exception as exc:
            parse_errors.append(str(exc))
            continue
        if best_score is None or score > best_score:
            best_frame = frame
            best_score = score

    if best_frame is not None:
        return best_frame

    detail = "; ".join(parse_errors) or str(last_error)
    raise ValueError(f"无法读取文本表格：{detail}")


def _is_usable_text_frame(frame: pd.DataFrame) -> bool:
    if frame.shape[1] < 2:
        return False
    numeric_columns = 0
    for column in frame.columns:
        values = frame[column].dropna()
        if values.map(_looks_numeric).sum() > 0:
            numeric_columns += 1
    return numeric_columns >= 2


def _clean_text_frame(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.replace(r"^\s*$", np.nan, regex=True).dropna(axis=1, how="all")


def _read_text_with_fallback(path: Path) -> str:
    last_error: Exception | None = None
    for encoding in TEXT_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except Exception as exc:
            last_error = exc
    raise ValueError(f"无法读取文本表格：{last_error}")


def _text_delimiter_candidates(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return ["comma", "tab", "semicolon", "whitespace"]
    if suffix == ".tsv":
        return ["tab", "whitespace"]
    return ["tab", "comma", "semicolon", "whitespace"]


def _dataframe_from_text(text: str, delimiter_name: str) -> tuple[pd.DataFrame, tuple[int, int, int, int]]:
    rows = _split_text_rows(text, delimiter_name)
    if not rows:
        raise ValueError("没有找到可读取的文本行。")

    numeric_flags = [_row_has_at_least_two_numbers(row) for row in rows]
    start, end = _longest_true_block(numeric_flags)
    if start is None or end is None:
        raise ValueError("没有找到至少两列数字数据。")

    header_index = start - 1 if start > 0 and _is_header_row(rows[start - 1]) else None
    data_rows = rows[start:end]
    max_columns = max(len(row) for row in data_rows)
    if header_index is not None:
        max_columns = max(max_columns, len(rows[header_index]))
        columns = _unique_text_columns(rows[header_index], max_columns)
    else:
        columns = ["波长"] + [f"数据{i}" for i in range(1, max_columns)]

    normalized_rows = [_pad_row(row, max_columns) for row in data_rows]
    frame = pd.DataFrame(normalized_rows, columns=columns)
    frame = frame.dropna(axis=1, how="all")
    frame = frame.loc[:, [col for col in frame.columns if str(col).strip()]]
    if frame.shape[1] < 2:
        raise ValueError("文本表格少于两列，无法作为光谱数据读取。")

    exact_width_rows = sum(1 for row in data_rows if len(row) == max_columns)
    numeric_columns = sum(1 for column in frame.columns if as_numeric(frame[column]).notna().sum() > 0)
    score = (len(data_rows), exact_width_rows, numeric_columns, -frame.shape[1])
    return frame, score


def _split_text_rows(text: str, delimiter_name: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip().lstrip("\ufeff")
        if not line:
            continue
        if delimiter_name == "whitespace":
            fields = re.split(r"\s+", line)
        else:
            delimiter = {"comma": ",", "tab": "\t", "semicolon": ";"}[delimiter_name]
            fields = next(csv.reader([line], delimiter=delimiter))
        cleaned = [field.strip().lstrip("\ufeff") for field in fields]
        cleaned = [field for field in cleaned if field != ""]
        if len(cleaned) >= 2:
            rows.append(cleaned)
    return rows


def _longest_true_block(flags: list[bool]) -> tuple[int | None, int | None]:
    best_start: int | None = None
    best_end: int | None = None
    current_start: int | None = None
    for index, flag in enumerate(flags + [False]):
        if flag and current_start is None:
            current_start = index
        elif not flag and current_start is not None:
            if best_start is None or index - current_start > best_end - best_start:
                best_start = current_start
                best_end = index
            current_start = None
    return best_start, best_end


def _is_header_row(row: list[str]) -> bool:
    return len(row) >= 2 and not _row_has_at_least_two_numbers(row)


def _row_has_at_least_two_numbers(row: list[str]) -> bool:
    return sum(1 for value in row if _looks_numeric(value)) >= 2


def _looks_numeric(value: object) -> bool:
    text = unicodedata.normalize("NFKC", str(value)).strip()
    text = text.replace("−", "-").replace("，", "").replace(",", "")
    if not text:
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _unique_text_columns(raw_columns: list[str], count: int) -> list[str]:
    columns = [(column.strip() if column.strip() else f"列{index + 1}") for index, column in enumerate(raw_columns)]
    while len(columns) < count:
        columns.append(f"列{len(columns) + 1}")

    seen: dict[str, int] = {}
    unique_columns: list[str] = []
    for column in columns[:count]:
        base = column
        seen[base] = seen.get(base, 0) + 1
        unique_columns.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return unique_columns


def _pad_row(row: list[str], width: int) -> list[str | None]:
    return row[:width] + [None] * max(width - len(row), 0)


def normalized_columns(df: pd.DataFrame) -> list[str]:
    return [str(col).strip() for col in df.columns]


def guess_wavelength_column(columns: Iterable[str]) -> str | None:
    lowered = [(col, col.lower()) for col in columns]
    preferred_tokens = ("wvlgth", "wavelength", "wave length", "lambda", "波长")
    for col, low in lowered:
        if any(token in low for token in preferred_tokens):
            return col
    for col, low in lowered:
        if "nm" in low:
            return col
    return next(iter(columns), None)


def guess_irradiance_column(columns: Iterable[str], reference_kind: str | None = None) -> str | None:
    cols = list(columns)
    if reference_kind:
        for col in cols:
            if col.strip().lower() == reference_kind.strip().lower():
                return col
        fallback_tokens = {
            "AM1.5G": ("global", "global_tilt", "global tilt", "tilt"),
            "AM1.5D": ("direct", "circumsolar"),
        }.get(reference_kind, ())
        for col in cols:
            low = col.lower()
            if any(token in low for token in fallback_tokens):
                return col

    score_tokens = (
        "am1.5g",
        "am1.5d",
        "irradiance",
        "irrad",
        "intensity",
        "power",
        "spectrum",
        "spectral",
        "辐照",
        "光谱",
        "强度",
        "功率",
    )
    wavelength_col = guess_wavelength_column(cols)
    candidates = []
    for col in cols:
        if col == wavelength_col:
            continue
        low = col.lower()
        score = sum(1 for token in score_tokens if token in low)
        if score:
            candidates.append((score, col))
    if candidates:
        return sorted(candidates, reverse=True)[0][1]

    for col in cols:
        if col != wavelength_col:
            return col
    return None


def reference_irradiance_choices(df: pd.DataFrame | None) -> list[str]:
    if df is None:
        return list(DEFAULT_REFERENCE_CHOICES)

    columns = normalized_columns(df)
    wavelength_col = guess_wavelength_column(columns)
    am_choices: list[str] = []
    semantic_choices: list[str] = []
    named_numeric_choices: list[str] = []
    numeric_choices: list[str] = []
    seen: set[str] = set()

    for index, column in enumerate(columns):
        label = str(column).strip()
        if not label or label == wavelength_col or label in seen:
            continue
        if index >= df.shape[1] or as_numeric(df.iloc[:, index]).notna().sum() < 2:
            continue

        seen.add(label)
        numeric_choices.append(label)
        if not _is_placeholder_reference_column(label):
            named_numeric_choices.append(label)

        tier = _reference_column_tier(label)
        if tier >= 3:
            am_choices.append(label)
        elif tier > 0 and not _is_placeholder_reference_column(label):
            semantic_choices.append(label)

    if am_choices:
        return am_choices
    if semantic_choices:
        return semantic_choices
    if named_numeric_choices:
        return named_numeric_choices
    return numeric_choices


def _is_placeholder_reference_column(label: str) -> bool:
    low = label.strip().lower()
    return low.startswith("unnamed") or low.startswith("列")


def _reference_column_tier(label: str) -> int:
    low = label.strip().lower()
    compact = re.sub(r"[\s_\-]+", "", low)
    if re.fullmatch(r"am\d+(?:\.\d+)?[a-z]?", compact) or re.search(r"\bam\s*\d", low):
        return 3
    if any(token in low for token in ("irradiance", "irrad", "radiance", "spectrum", "spectral", "辐照", "光谱")):
        return 2
    if any(token in low for token in ("global", "direct", "diffuse", "circumsolar", "etr")):
        return 1
    return 0


def guess_response_column(columns: Iterable[str]) -> str | None:
    cols = list(columns)
    wavelength_col = guess_wavelength_column(cols)
    score_tokens = (
        "sr",
        "a/w",
        "responsivity",
        "response",
        "spectral response",
        "响应",
        "光谱响应",
        "量子效率",
    )
    candidates = []
    for col in cols:
        if col == wavelength_col:
            continue
        low = col.lower()
        score = sum(1 for token in score_tokens if token in low)
        if score:
            candidates.append((score, col))
    if candidates:
        return sorted(candidates, reverse=True)[0][1]

    for col in cols:
        if col != wavelength_col:
            return col
    return None


def guess_voltage_column(columns: Iterable[str]) -> str | None:
    columns_list = guess_voltage_columns(columns)
    return columns_list[0] if columns_list else None


def guess_current_column(columns: Iterable[str]) -> str | None:
    columns_list = guess_current_columns(columns)
    return columns_list[0] if columns_list else None


def voltage_column_score(column: str) -> int:
    low = column.lower()
    skip_tokens = ("time", "时间", "index", "索引", "current", "电流", "power", "功率", "resistance", "电阻")
    if any(token in low for token in skip_tokens):
        return 0
    score_tokens = ("voltage", "volt", "bias", "v)", "(v", "电压")
    score = sum(1 for token in score_tokens if token in low)
    text = column.strip().lower()
    if text in {"v", "u"} or re.fullmatch(r"v\s*\d+", text):
        score += 2
    return score


def current_column_score(column: str) -> int:
    low = column.lower().replace("μ", "u").replace("µ", "u")
    skip_tokens = ("time", "时间", "index", "索引", "voltage", "电压", "power", "功率", "resistance", "电阻")
    if any(token in low for token in skip_tokens):
        return 0
    score_tokens = ("current", "curr", "isc", "i_sc", "a)", "(a", "ma", "ua", "na", "电流", "短路电流")
    score = sum(1 for token in score_tokens if token in low)
    text = column.strip().lower()
    if text in {"i", "a"} or re.fullmatch(r"i\s*\d+", text) or re.fullmatch(r"i[_-]?\d+", text):
        score += 2
    return score


def guess_voltage_columns(columns: Iterable[str]) -> list[str]:
    candidates = [(voltage_column_score(col), index, col) for index, col in enumerate(columns)]
    candidates = [item for item in candidates if item[0] > 0]
    return [col for _score, _index, col in sorted(candidates, key=lambda item: (-item[0], item[1]))]


def guess_current_columns(columns: Iterable[str]) -> list[str]:
    cols = list(columns)
    candidates = [(current_column_score(col), index, col) for index, col in enumerate(cols)]
    candidates = [item for item in candidates if item[0] > 0]
    return [col for _score, _index, col in sorted(candidates, key=lambda item: (-item[0], item[1]))]


def voltage_column_for_current(columns: Iterable[str], current_column: str) -> str | None:
    cols = list(columns)
    voltage_columns = guess_voltage_columns(cols)
    if not voltage_columns:
        return None
    current_index = cols.index(current_column) if current_column in cols else len(cols)
    voltage_before = [col for col in voltage_columns if cols.index(col) < current_index]
    if voltage_before:
        return max(voltage_before, key=lambda col: cols.index(col))
    return voltage_columns[0]


def current_to_ma_factor(column_name: str) -> float:
    low = column_name.lower().replace("μ", "u").replace("µ", "u")
    if "ma" in low or "毫安" in low:
        return 1.0
    if "ua" in low or "微安" in low:
        return 0.001
    if "na" in low or "纳安" in low:
        return 0.000001
    return 1000.0


def voltage_to_v_factor(column_name: str) -> float:
    low = column_name.lower().replace("μ", "u").replace("µ", "u")
    if "mv" in low or "毫伏" in low:
        return 0.001
    return 1.0


def parse_current_values_ma(text: str) -> list[float]:
    normalized = (
        text.strip()
        .replace("，", ",")
        .replace("；", ",")
        .replace(";", ",")
        .replace("、", ",")
        .replace("?", ",")
        .replace("\n", ",")
        .replace("\t", ",")
    )
    if not normalized:
        return []
    values = []
    for token in re.split(r"[,\s]+", normalized):
        token = token.strip()
        if not token:
            continue
        token = re.sub(r"(?i)\s*m?a$", "", token).strip()
        if not token:
            continue
        values.append(float(token))
    return values


def as_numeric(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")
    text = (
        series.astype(str)
        .str.strip()
        .str.replace("，", "", regex=False)
        .str.replace(",", "", regex=False)
    )
    return pd.to_numeric(text, errors="coerce")


def spectrum_from_dataframe(
    df: pd.DataFrame,
    wavelength_col: str,
    irradiance_col: str,
    unit_factor: float = 1.0,
) -> Spectrum:
    if wavelength_col not in df.columns:
        raise ValueError(f"找不到波长列：{wavelength_col}")
    if irradiance_col not in df.columns:
        raise ValueError(f"找不到辐照度列：{irradiance_col}")

    wavelengths = as_numeric(df[wavelength_col])
    irradiance = as_numeric(df[irradiance_col]) * float(unit_factor)
    data = pd.DataFrame({"wavelength": wavelengths, "irradiance": irradiance})
    data = data.replace([np.inf, -np.inf], np.nan).dropna()
    data = data[data["wavelength"] >= 0]
    if data.empty:
        raise ValueError("所选两列没有可用的数字数据。")

    data = data.groupby("wavelength", as_index=False)["irradiance"].mean()
    data = data.sort_values("wavelength")
    return Spectrum(
        wavelength_nm=data["wavelength"].to_numpy(dtype=float),
        irradiance_w_m2_nm=data["irradiance"].to_numpy(dtype=float),
    )


def response_from_dataframe(
    df: pd.DataFrame,
    wavelength_col: str,
    response_col: str,
) -> SpectralResponse:
    if wavelength_col not in df.columns:
        raise ValueError(f"找不到 SR 波长列：{wavelength_col}")
    if response_col not in df.columns:
        raise ValueError(f"找不到 SR 列：{response_col}")

    wavelengths = as_numeric(df[wavelength_col])
    response = as_numeric(df[response_col])
    data = pd.DataFrame({"wavelength": wavelengths, "response": response})
    data = data.replace([np.inf, -np.inf], np.nan).dropna()
    data = data[data["wavelength"] >= 0]
    if data.empty:
        raise ValueError("所选 SR 两列没有可用的数字数据。")

    data = data.groupby("wavelength", as_index=False)["response"].mean()
    data = data.sort_values("wavelength")
    return SpectralResponse(
        wavelength_nm=data["wavelength"].to_numpy(dtype=float),
        response_a_w=data["response"].to_numpy(dtype=float),
    )


def read_iv_curve(path: Path, sheet_name: str | int | None = None) -> IVCurve:
    curves = read_iv_curves(path, sheet_name)
    if not curves:
        raise ValueError("没有找到可用的 IV 数据。")
    return curves[0]


def read_iv_curves(path: Path, sheet_name: str | int | None = None) -> list[IVCurve]:
    suffix = path.suffix.lower()
    curves: list[IVCurve] = []
    if suffix in EXCEL_TABLE_SUFFIXES:
        sheet_names: list[str | int]
        if sheet_name is not None:
            sheet_names = [sheet_name]
        else:
            sheet_names = list_excel_sheets(path)
        for sheet in sheet_names:
            raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
            rows = raw.where(pd.notna(raw), "").to_numpy().tolist()
            header_index = find_iv_header_row(rows)
            if header_index is None:
                continue
            df, source_rows, header_row = iv_dataframe_from_rows(rows, header_index)
            curves.extend(iv_curves_from_dataframe(
                df,
                source_name=path.name,
                sheet_name=str(sheet),
                header_row=header_row,
                source_rows=source_rows,
            ))
        if curves:
            return curves
        raise ValueError("没有找到同时包含电压列和电流列的 IV 数据表头。")

    if suffix in TEXT_TABLE_SUFFIXES:
        rows, header_index = read_delimited_iv_rows(path)
        df, source_rows, header_row = iv_dataframe_from_rows(rows, header_index)
        return iv_curves_from_dataframe(
            df,
            source_name=path.name,
            sheet_name="",
            header_row=header_row,
            source_rows=source_rows,
        )

    raise ValueError(f"暂不支持该 IV 文件类型：{path.suffix}")


def read_delimited_iv_rows(path: Path) -> tuple[list[list[str]], int]:
    last_error: Exception | None = None
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except Exception as exc:
            last_error = exc
    if text is None:
        raise ValueError(f"无法读取 IV 文本表格：{last_error}")

    delimiter_names = ["tab"] if path.suffix.lower() == ".tsv" else _text_delimiter_candidates(path)
    for delimiter_name in delimiter_names:
        rows = _split_text_rows(text, delimiter_name)
        header_index = find_iv_header_row(rows)
        if header_index is not None:
            return rows, header_index

    for delimiter_name in delimiter_names:
        rows = _split_text_rows(text, delimiter_name)
        synthetic = synthetic_iv_rows_from_numeric_block(rows)
        if synthetic is not None:
            return synthetic, 0
    raise ValueError("没有找到同时包含电压列和电流列的 IV 数据表头。")


def synthetic_iv_rows_from_numeric_block(rows: list[list[str]]) -> list[list[str]] | None:
    numeric_flags = [_row_has_at_least_two_numbers(row) for row in rows]
    start, end = _longest_true_block(numeric_flags)
    if start is None or end is None:
        return None

    data_rows = rows[start:end]
    max_columns = max(len(row) for row in data_rows)
    if max_columns < 2:
        return None
    headers = ["电压 (V)"] + [f"电流{i} (A)" for i in range(1, max_columns)]
    return [headers] + [_pad_row(row, max_columns) for row in data_rows]


def clean_header_cell(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lstrip("\ufeff")


def find_iv_header_row(rows: list[list[object]]) -> int | None:
    for index, row in enumerate(rows):
        columns = [clean_header_cell(value) for value in row]
        if len([column for column in columns if column]) < 2:
            continue
        if guess_voltage_column(columns) and guess_current_column(columns):
            return index
    return None


def unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    output = []
    for index, header in enumerate(headers):
        name = header or f"列{index + 1}"
        count = counts.get(name, 0)
        counts[name] = count + 1
        output.append(name if count == 0 else f"{name}_{count + 1}")
    return output


def iv_dataframe_from_rows(
    rows: list[list[object]],
    header_index: int,
) -> tuple[pd.DataFrame, list[int], int]:
    raw_headers = [clean_header_cell(value) for value in rows[header_index]]
    width = len(raw_headers)
    while width > 0 and not raw_headers[width - 1]:
        width -= 1
    headers = unique_headers(raw_headers[:width])
    records = []
    source_rows = []
    for source_row, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = [clean_header_cell(value) for value in row[:width]]
        if len(values) < width:
            values.extend([""] * (width - len(values)))
        if not any(values):
            continue
        records.append(values)
        source_rows.append(source_row)
    return pd.DataFrame(records, columns=headers), source_rows, header_index + 1


def iv_curve_from_dataframe(
    df: pd.DataFrame,
    *,
    source_name: str,
    sheet_name: str,
    header_row: int | None,
    source_rows: list[int] | None,
) -> IVCurve:
    curves = iv_curves_from_dataframe(
        df,
        source_name=source_name,
        sheet_name=sheet_name,
        header_row=header_row,
        source_rows=source_rows,
    )
    if not curves:
        raise ValueError("IV 表格中没有可用的电压/电流数据。")
    return curves[0]


def iv_curves_from_dataframe(
    df: pd.DataFrame,
    *,
    source_name: str,
    sheet_name: str,
    header_row: int | None,
    source_rows: list[int] | None,
) -> list[IVCurve]:
    df.columns = normalized_columns(df)
    columns = normalized_columns(df)
    current_cols = guess_current_columns(columns)
    if not guess_voltage_columns(columns):
        raise ValueError("IV 表格中无法识别电压列。")
    if not current_cols:
        raise ValueError("IV 表格中无法识别电流列。")

    curves: list[IVCurve] = []
    for current_col in current_cols:
        voltage_col = voltage_column_for_current(columns, current_col)
        if voltage_col is None:
            continue

        voltage = as_numeric(df[voltage_col]) * voltage_to_v_factor(voltage_col)
        current_ma = as_numeric(df[current_col]) * current_to_ma_factor(current_col)
        valid_mask = voltage.notna() & current_ma.notna()
        valid_positions = np.flatnonzero(valid_mask.to_numpy())
        if valid_positions.size == 0:
            continue

        first_position = int(valid_positions[0])
        first_data_row = (
            source_rows[first_position]
            if source_rows is not None and first_position < len(source_rows)
            else None
        )
        data = pd.DataFrame(
            {
                "voltage": voltage.loc[valid_mask].to_numpy(dtype=float),
                "current_ma": current_ma.loc[valid_mask].to_numpy(dtype=float),
            }
        )
        if len(data) < 2:
            continue

        curves.append(
            IVCurve(
                source_name=source_name,
                sheet_name=sheet_name,
                voltage_column=voltage_col,
                current_column=current_col,
                header_row=header_row,
                first_data_row=first_data_row,
                voltage_v=data["voltage"].to_numpy(dtype=float),
                current_ma=data["current_ma"].to_numpy(dtype=float),
                isc_ma=abs(float(current_ma.iloc[first_position])),
            )
        )

    if not curves:
        raise ValueError("IV 表格的电压/电流列没有可用的数字数据。")
    return curves


def make_grid(start_nm: float, end_nm: float, step_nm: float) -> np.ndarray:
    if step_nm <= 0:
        raise ValueError("网格步长必须大于 0。")
    if end_nm <= start_nm:
        raise ValueError("终止波长必须大于起始波长。")
    count = int(math.floor((end_nm - start_nm) / step_nm + 0.5))
    grid = start_nm + np.arange(count + 1, dtype=float) * step_nm
    if grid[-1] < end_nm:
        grid = np.append(grid, end_nm)
    elif grid[-1] > end_nm:
        grid[-1] = end_nm
    return grid


def interpolate_spectrum(spectrum: Spectrum, grid_nm: np.ndarray, *, strict: bool) -> np.ndarray:
    wavelengths = spectrum.wavelength_nm
    irradiance = spectrum.irradiance_w_m2_nm
    if wavelengths.size < 2:
        raise ValueError("光谱至少需要两个波长点才能插值和积分。")

    start = float(grid_nm[0])
    end = float(grid_nm[-1])
    min_w = float(np.min(wavelengths))
    max_w = float(np.max(wavelengths))
    if strict and (min_w > start or max_w < end):
        raise ValueError(
            f"参考光谱覆盖范围不足：当前为 {min_w:g}-{max_w:g} nm，需要 {start:g}-{end:g} nm。"
        )

    values = np.interp(grid_nm, wavelengths, irradiance)
    values[grid_nm < min_w] = 0.0
    values[grid_nm > max_w] = 0.0
    return values


def integrate_xy(x: np.ndarray, y: np.ndarray) -> float:
    if x.size < 2:
        return 0.0
    return float(np.trapezoid(y, x))


def integrate_range(x: np.ndarray, y: np.ndarray, start_nm: float, end_nm: float) -> float:
    inner = x[(x > start_nm) & (x < end_nm)]
    local_x = np.concatenate(([start_nm], inner, [end_nm]))
    local_y = np.interp(local_x, x, y)
    return integrate_xy(local_x, local_y)


def classify_ratio(ratio: float) -> str:
    for name, lower, upper in CLASS_LIMITS:
        if lower <= ratio <= upper:
            return name
    return "未达 C"


def classify_overall(ratios: Iterable[float]) -> str:
    ratio_list = list(ratios)
    if not ratio_list or any(not np.isfinite(ratio) for ratio in ratio_list):
        return "未达 C"
    for name, lower, upper in CLASS_LIMITS:
        if all(lower <= ratio <= upper for ratio in ratio_list):
            return name
    return "未达 C"


def calculate_spd(
    reference_spectrum: Spectrum,
    test_spectrum: Spectrum,
    *,
    reference_name: str,
    test_name: str,
    start_nm: float = 300.0,
    end_nm: float = 1200.0,
    step_nm: float = 1.0,
    normalize: bool = False,
) -> CalculationResult:
    grid = make_grid(start_nm, end_nm, step_nm)
    reference = interpolate_spectrum(reference_spectrum, grid, strict=True)
    test_raw = interpolate_spectrum(test_spectrum, grid, strict=False)

    reference_total = integrate_xy(grid, reference)
    test_raw_total = integrate_xy(grid, test_raw)
    if reference_total <= 0:
        raise ValueError("参考光谱在计算波段内的积分辐照度为 0。")
    if test_raw_total <= 0:
        raise ValueError("测试光谱在计算波段内的积分辐照度为 0。")

    scale_factor = reference_total / test_raw_total if normalize else 1.0
    test_scaled = test_raw * scale_factor
    test_scaled_total = integrate_xy(grid, test_scaled)
    absolute_error = np.abs(test_scaled - reference)

    spd_percent = integrate_xy(grid, absolute_error) / reference_total * 100.0
    spc_mask = test_scaled >= (0.1 * reference)
    spc_percent = integrate_xy(grid, np.where(spc_mask, reference, 0.0)) / reference_total * 100.0

    band_rows: list[dict[str, object]] = []
    ratios = []
    for start, end in IEC_BANDS:
        clipped_start = max(float(start_nm), start)
        clipped_end = min(float(end_nm), end)
        if clipped_end <= clipped_start:
            continue
        ref_band = integrate_range(grid, reference, clipped_start, clipped_end)
        test_band = integrate_range(grid, test_scaled, clipped_start, clipped_end)
        ref_percent = ref_band / reference_total * 100.0 if reference_total else np.nan
        test_percent = test_band / test_scaled_total * 100.0 if test_scaled_total else np.nan
        ratio = test_percent / ref_percent if ref_percent > 0 else np.nan
        ratios.append(ratio)
        band_rows.append(
            {
                "range": f"{clipped_start:g}-{clipped_end:g}",
                "ref_integral": ref_band,
                "test_integral": test_band,
                "ref_percent": ref_percent,
                "test_percent": test_percent,
                "ratio": ratio,
                "class": classify_ratio(ratio),
            }
        )

    return CalculationResult(
        reference_name=reference_name,
        test_name=test_name,
        start_nm=start_nm,
        end_nm=end_nm,
        step_nm=step_nm,
        normalize=normalize,
        scale_factor=scale_factor,
        grid_nm=grid,
        reference=reference,
        test_raw=test_raw,
        test_scaled=test_scaled,
        absolute_error=absolute_error,
        reference_total=reference_total,
        test_raw_total=test_raw_total,
        test_scaled_total=test_scaled_total,
        spd_percent=spd_percent,
        spc_percent=spc_percent,
        overall_class=classify_overall(ratios),
        band_rows=band_rows,
    )


def calculate_current_correction(
    reference_spectrum: Spectrum,
    test_spectrum: Spectrum,
    spectral_response: SpectralResponse,
    *,
    sr_name: str,
    sr_sheet_name: str,
    sr_wavelength_column: str,
    sr_column: str,
    measured_isc: float,
    test_temperature_c: float,
    temperature_coefficient_percent_per_c: float = DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C,
    reference_temperature_c: float = DEFAULT_REFERENCE_TEMPERATURE_C,
    start_nm: float | None = None,
    end_nm: float | None = None,
) -> CurrentCorrectionResult:
    data = pd.DataFrame(
        {
            "wavelength": spectral_response.wavelength_nm,
            "sr": spectral_response.response_a_w,
        }
    )
    data = data.replace([np.inf, -np.inf], np.nan).dropna()
    if start_nm is not None:
        data = data[data["wavelength"] >= float(start_nm)]
    if end_nm is not None:
        data = data[data["wavelength"] <= float(end_nm)]
    data = data.sort_values("wavelength")

    if len(data) < 2:
        raise ValueError("SR 波段至少需要两个有效波长点。")

    grid = data["wavelength"].to_numpy(dtype=float)
    sr = data["sr"].to_numpy(dtype=float)
    reference = interpolate_spectrum(reference_spectrum, grid, strict=False)
    test = interpolate_spectrum(test_spectrum, grid, strict=False)
    reference_sr_weighted = reference * sr
    test_sr_weighted = test * sr

    reference_irradiance_integral = integrate_xy(grid, reference)
    test_irradiance_integral = integrate_xy(grid, test)
    reference_sr_integral = integrate_xy(grid, reference_sr_weighted)
    test_sr_integral = integrate_xy(grid, test_sr_weighted)

    if reference_irradiance_integral <= 0:
        raise ValueError("SR 波段内的标准光谱辐照度积分为 0。")
    if test_irradiance_integral <= 0:
        raise ValueError("SR 波段内的实测光谱辐照度积分为 0。")
    if reference_sr_integral <= 0:
        raise ValueError("SR 加权后的标准光谱积分为 0。")
    if test_sr_integral <= 0:
        raise ValueError("SR 加权后的实测光谱积分为 0。")
    if measured_isc <= 0:
        raise ValueError("实测短路电流 ISC 必须大于 0。")

    mg = reference_irradiance_integral / test_irradiance_integral
    temperature_alpha = temperature_coefficient_percent_per_c / 100.0
    temperature_denominator = 1.0 - temperature_alpha * (reference_temperature_c - test_temperature_c)
    if temperature_denominator <= 0:
        raise ValueError("温度修正分母小于等于 0，请检查测试温度或温度系数。")
    mt = 1.0 / temperature_denominator
    mmf = (test_sr_integral / test_irradiance_integral) / (
        reference_sr_integral / reference_irradiance_integral
    )
    if mmf <= 0 or not np.isfinite(mmf):
        raise ValueError("光谱失配因子 MMF 无法计算，请检查 SR 和光谱数据。")

    corrected_cv = measured_isc * (mg * mt) / mmf
    return CurrentCorrectionResult(
        sr_name=sr_name,
        sr_sheet_name=sr_sheet_name,
        sr_wavelength_column=sr_wavelength_column,
        sr_column=sr_column,
        measured_isc=measured_isc,
        test_temperature_c=test_temperature_c,
        temperature_coefficient_percent_per_c=temperature_coefficient_percent_per_c,
        reference_temperature_c=reference_temperature_c,
        grid_nm=grid,
        sr=sr,
        reference=reference,
        test=test,
        reference_sr_weighted=reference_sr_weighted,
        test_sr_weighted=test_sr_weighted,
        reference_irradiance_integral=reference_irradiance_integral,
        test_irradiance_integral=test_irradiance_integral,
        reference_sr_integral=reference_sr_integral,
        test_sr_integral=test_sr_integral,
        mg=mg,
        mt=mt,
        mmf=mmf,
        corrected_cv=corrected_cv,
    )


def evaluate_current_corrections(
    labeled_corrections: list[tuple],
) -> list[CurrentCorrectionEvaluation]:
    if not labeled_corrections:
        return []

    def item_label(index: int) -> str:
        return str(labeled_corrections[index][0])

    def item_correction(index: int) -> CurrentCorrectionResult:
        return labeled_corrections[index][1]

    def item_spectrum_label(index: int) -> str:
        return str(labeled_corrections[index][2]) if len(labeled_corrections[index]) > 2 else ""

    def item_spectrum_name(index: int) -> str:
        return str(labeled_corrections[index][3]) if len(labeled_corrections[index]) > 3 else ""

    def item_source(index: int) -> str:
        return str(labeled_corrections[index][4]) if len(labeled_corrections[index]) > 4 else ""

    active = [True] * len(labeled_corrections)
    reasons: list[list[str]] = [[] for _item in labeled_corrections]

    while True:
        active_indices = [idx for idx, is_active in enumerate(active) if is_active]
        if len(active_indices) <= 1:
            break

        active_cvs = np.array(
            [item_correction(idx).corrected_cv for idx in active_indices],
            dtype=float,
        )
        mean_cv = float(np.mean(active_cvs))
        if mean_cv <= 0 or not np.isfinite(mean_cv):
            break

        rejected = False
        for idx in active_indices:
            cv = item_correction(idx).corrected_cv
            deviation = abs(cv - mean_cv) / mean_cv * 100.0
            if deviation > IEC_A5_CV_DEVIATION_LIMIT_PERCENT:
                reasons[idx].append(
                    f"CV偏离平均值 {deviation:.2f}% > {IEC_A5_CV_DEVIATION_LIMIT_PERCENT:g}%"
                )
                active[idx] = False
                rejected = True
        if rejected:
            continue

        active_indices = [idx for idx, is_active in enumerate(active) if is_active]
        if len(active_indices) <= 1:
            break

        active_currents = np.array(
            [item_correction(idx).measured_isc for idx in active_indices],
            dtype=float,
        )
        mean_isc = float(np.mean(active_currents))
        if mean_isc > 0 and np.isfinite(mean_isc):
            isc_range_percent = (float(np.max(active_currents)) - float(np.min(active_currents))) / mean_isc * 100.0
            if isc_range_percent > IEC_A5_ISC_RANGE_LIMIT_PERCENT:
                median_isc = float(np.median(active_currents))
                target = max(
                    active_indices,
                    key=lambda idx: abs(item_correction(idx).measured_isc - median_isc),
                )
                reasons[target].append(
                    f"ISC范围 {isc_range_percent:.2f}% > {IEC_A5_ISC_RANGE_LIMIT_PERCENT:g}%"
                )
                active[target] = False
                continue

        active_indices = [idx for idx, is_active in enumerate(active) if is_active]
        if len(active_indices) <= 1:
            break

        active_cvs = np.array(
            [item_correction(idx).corrected_cv for idx in active_indices],
            dtype=float,
        )
        mean_cv = float(np.mean(active_cvs))
        cv_std_percent = (
            float(np.std(active_cvs, ddof=1)) / mean_cv * 100.0
            if len(active_cvs) > 1 and mean_cv > 0
            else 0.0
        )
        if cv_std_percent > IEC_A5_CV_STD_LIMIT_PERCENT:
            target = max(
                active_indices,
                key=lambda idx: abs(item_correction(idx).corrected_cv - mean_cv),
            )
            reasons[target].append(
                f"CV标准偏差 {cv_std_percent:.2f}% > {IEC_A5_CV_STD_LIMIT_PERCENT:g}%"
            )
            active[target] = False
            continue

        break

    active_indices = [idx for idx, is_active in enumerate(active) if is_active]
    valid_count = len(active_indices)
    mean_for_deviation = (
        float(np.mean([item_correction(idx).corrected_cv for idx in active_indices]))
        if active_indices
        else float(np.mean([item_correction(idx).corrected_cv for idx in range(len(labeled_corrections))]))
    )

    evaluations: list[CurrentCorrectionEvaluation] = []
    for idx in range(len(labeled_corrections)):
        label = item_label(idx)
        correction = item_correction(idx)
        deviation = (
            abs(correction.corrected_cv - mean_for_deviation) / mean_for_deviation * 100.0
            if mean_for_deviation > 0 and np.isfinite(mean_for_deviation)
            else 0.0
        )
        row_reasons = list(reasons[idx])
        row_valid = active[idx]
        if valid_count < IEC_A5_MIN_VALID_CV_COUNT:
            row_valid = False
            row_reasons.append(f"有效CV少于{IEC_A5_MIN_VALID_CV_COUNT}组")
        evaluations.append(
            CurrentCorrectionEvaluation(
                label=label,
                spectrum_label=item_spectrum_label(idx),
                spectrum_name=item_spectrum_name(idx),
                source=item_source(idx),
                correction=correction,
                deviation_percent=deviation,
                valid=row_valid,
                reasons=row_reasons,
            )
        )
    return evaluations


def fmt_number(value: float, digits: int = 4) -> str:
    if not np.isfinite(value):
        return "-"
    if value == 0:
        return "0"
    if abs(value) >= 10000 or abs(value) < 0.001:
        return f"{value:.{digits}e}"
    return f"{value:.{digits}g}"


def reference_legend_label(reference_name: str | None) -> str:
    label = str(reference_name or "").strip()
    if not label:
        return "参考"
    match = re.search(r"AM\s*1\.5\s*[GD]", label, re.IGNORECASE)
    if match:
        return match.group(0).upper().replace(" ", "")
    return label


def spectrum_color_for_index(index: int) -> str:
    return TEST_SERIES_COLORS[index % len(TEST_SERIES_COLORS)]


def spectrum_style_for_index(index: int) -> PlotSeriesStyle:
    index = max(0, int(index))
    color = spectrum_color_for_index(index)
    dash_group = index // len(TEST_SERIES_COLORS)
    dash = TEST_SERIES_DASH_PATTERNS[dash_group % len(TEST_SERIES_DASH_PATTERNS)]
    return PlotSeriesStyle(color=color, dash=dash)


def spectrum_line_name_for_index(index: int) -> str:
    index = max(0, int(index))
    dash_group = index // len(TEST_SERIES_COLORS)
    return TEST_SERIES_LINE_NAMES[dash_group % len(TEST_SERIES_LINE_NAMES)]


def plot_series_data(
    y_values: np.ndarray,
    color: str,
    label: str,
    *,
    dash: tuple[int, ...] = (),
    spectrum_label: str = "",
    is_reference: bool = False,
) -> PlotSeriesData:
    return PlotSeriesData(
        y_values=np.asarray(y_values, dtype=float),
        color=color,
        label=label,
        dash=tuple(int(part) for part in dash if int(part) > 0),
        spectrum_label=spectrum_label,
        is_reference=is_reference,
    )


def _coerce_plot_series_data(item: object) -> PlotSeriesData:
    if isinstance(item, PlotSeriesData):
        return plot_series_data(
            item.y_values,
            item.color,
            item.label,
            dash=item.dash,
            spectrum_label=item.spectrum_label,
            is_reference=item.is_reference,
        )
    values, color, label, *rest = item  # type: ignore[misc]
    dash = tuple(rest[0]) if rest else ()
    return plot_series_data(np.asarray(values, dtype=float), str(color), str(label), dash=dash)


def _scaled_dash(dash: tuple[int, ...], scale: int) -> tuple[int, ...]:
    return tuple(max(1, int(round(part * scale))) for part in dash)


def _draw_dashed_polyline(draw, points: list[tuple[int, int]], *, fill: str, width: int, dash: tuple[int, ...]) -> None:
    if len(points) < 2:
        return
    if not dash:
        draw.line(points, fill=fill, width=width, joint="curve")
        return
    pattern = list(dash if len(dash) % 2 == 0 else dash + dash)
    pattern_index = 0
    draw_on = True
    remaining = float(pattern[pattern_index])

    for start, end in zip(points, points[1:]):
        x0, y0 = start
        x1, y1 = end
        dx = x1 - x0
        dy = y1 - y0
        length = math.hypot(dx, dy)
        if length <= 0:
            continue
        pos = 0.0
        while pos < length:
            step = min(remaining, length - pos)
            start_ratio = pos / length
            end_ratio = (pos + step) / length
            segment_start = (int(round(x0 + dx * start_ratio)), int(round(y0 + dy * start_ratio)))
            segment_end = (int(round(x0 + dx * end_ratio)), int(round(y0 + dy * end_ratio)))
            if draw_on:
                draw.line([segment_start, segment_end], fill=fill, width=width)
            pos += step
            remaining -= step
            if remaining <= 1e-9:
                pattern_index = (pattern_index + 1) % len(pattern)
                remaining = float(pattern[pattern_index])
                draw_on = pattern_index % 2 == 0


def _legend_font_size(series_count: int) -> int:
    if series_count <= 8:
        return 9
    return max(5, 9 - math.ceil((series_count - 8) / 8))


def _plot_legend_layout(
    series_count: int,
    width: int,
    height: int,
    pad_left: int,
    pad_top: int,
    pad_bottom: int,
) -> PlotLegendLayout:
    count = max(1, int(series_count))
    font_size = _legend_font_size(count)
    base_row_step = max(12, font_size + 10)
    plot_h = max(1, height - pad_top - pad_bottom)
    rows_that_fit = max(1, int((plot_h - 8) // base_row_step))
    requested_columns = max(1, math.ceil(count / rows_that_fit))
    column_width = max(136, min(176, 110 + font_size * 7))
    desired_panel_width = requested_columns * column_width + 22
    min_plot_width = MIN_PLOT_AREA_WIDTH if width >= 820 else 260
    max_panel_width = max(120, width - pad_left - min_plot_width - 18)
    panel_width = min(desired_panel_width, max_panel_width)
    column_count = max(1, min(requested_columns, int(max(panel_width - 22, 1) // column_width) or 1))
    if column_count < requested_columns:
        column_width = max(112, (panel_width - 22) / column_count)
    rows_per_column = max(1, math.ceil(count / column_count))
    row_step = min(base_row_step, max(font_size + 4, (plot_h - 8) / rows_per_column))
    return PlotLegendLayout(
        font_size=font_size,
        row_step=row_step,
        rows_per_column=rows_per_column,
        column_count=column_count,
        column_width=column_width,
        panel_width=panel_width,
        panel_gap=18,
        line_length=max(18, font_size * 3),
    )


def _plot_image_dimensions_for_series(
    series_count: int,
    *,
    base_width: int = 1100,
    base_height: int = 400,
) -> tuple[int, int]:
    extra_width = max(0, math.ceil(max(series_count - 10, 0) / 10) * 120)
    return base_width + min(extra_width, 520), base_height


def _fit_canvas_text(canvas: tk.Canvas, text: str, font: tuple[str, int] | tuple[str, int, str], max_width: float) -> str:
    if max_width <= 0:
        return "..."
    try:
        measure = int(canvas.tk.call("font", "measure", font, text))
        if measure <= max_width:
            return text
        ellipsis = "..."
        clipped = text
        while clipped and int(canvas.tk.call("font", "measure", font, clipped + ellipsis)) > max_width:
            clipped = clipped[:-1]
        return clipped + ellipsis if clipped else ellipsis
    except Exception:
        approx_width = max(1, int(max_width / 7))
        return text if len(text) <= approx_width else text[: max(0, approx_width - 3)] + "..."


def _chart_font(size: int, *, bold: bool = False):
    from PIL import ImageFont

    font_dir = Path(os.environ.get("WINDIR", "C:\\Windows")) / "Fonts"
    candidates = [
        font_dir / ("msyhbd.ttc" if bold else "msyh.ttc"),
        font_dir / ("simhei.ttf" if bold else "msyh.ttc"),
        font_dir / ("arialbd.ttf" if bold else "arial.ttf"),
    ]
    for font_path in candidates:
        if font_path.exists():
            try:
                return ImageFont.truetype(str(font_path), size=size)
            except Exception:
                continue
    return ImageFont.load_default()


def _text_size(draw, text: str, font) -> tuple[int, int]:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def _fit_text(draw, text: str, font, max_width: int) -> str:
    if max_width <= 0 or _text_size(draw, text, font)[0] <= max_width:
        return text
    ellipsis = "..."
    clipped = text
    while clipped and _text_size(draw, clipped + ellipsis, font)[0] > max_width:
        clipped = clipped[:-1]
    return clipped + ellipsis if clipped else ellipsis


def _draw_centered_text(draw, xy: tuple[int, int], text: str, font, fill: str) -> None:
    width, height = _text_size(draw, text, font)
    draw.text((xy[0] - width / 2, xy[1] - height / 2), text, fill=fill, font=font)


def _draw_right_centered_text(draw, xy: tuple[int, int], text: str, font, fill: str) -> None:
    width, height = _text_size(draw, text, font)
    draw.text((xy[0] - width, xy[1] - height / 2), text, fill=fill, font=font)


def _render_plot_image(
    x_values: np.ndarray,
    series: list[object],
    x_label: str,
    y_label: str,
    subtitle: str,
    *,
    width: int = 1100,
    height: int = 400,
    scale: int = 2,
    isc_marker: tuple[float, str] | None = None,
) -> io.BytesIO | None:
    from PIL import Image, ImageDraw

    x = np.asarray(x_values, dtype=float)
    plot_series = [_coerce_plot_series_data(item) for item in series]
    all_y = np.concatenate([item.y_values for item in plot_series if item.y_values.size])
    finite_x = x[np.isfinite(x)]
    finite_y = all_y[np.isfinite(all_y)]
    if finite_x.size < 2 or finite_y.size == 0:
        return None

    x_min, x_max = float(np.min(finite_x)), float(np.max(finite_x))
    if x_max <= x_min:
        x_min -= 0.5
        x_max += 0.5

    raw_y_min, raw_y_max = float(np.min(finite_y)), float(np.max(finite_y))
    if raw_y_min >= 0:
        y_min, y_max = 0.0, raw_y_max
    elif raw_y_max <= 0:
        y_min, y_max = raw_y_min, 0.0
    else:
        y_min, y_max = raw_y_min, raw_y_max
    if y_max <= y_min:
        span = abs(y_max) if y_max else 1.0
        y_min -= span * 0.5
        y_max += span * 0.5
    padding = (y_max - y_min) * 0.08
    if raw_y_min < 0:
        y_min -= padding
    if raw_y_max > 0:
        y_max += padding

    def sx(value: float) -> int:
        return int(round(value * scale))

    pad_left, pad_top, pad_bottom = 84, 36, 56
    legend_layout = _plot_legend_layout(len(plot_series), width, height, pad_left, pad_top, pad_bottom)
    pad_right = int(math.ceil(legend_layout.panel_width + legend_layout.panel_gap))
    plot_w = max(80, width - pad_left - pad_right)
    plot_h = height - pad_top - pad_bottom
    plot_right = pad_left + plot_w

    def px(value: float) -> int:
        return sx(pad_left + (value - x_min) / (x_max - x_min) * plot_w)

    def py(value: float) -> int:
        return sx(pad_top + plot_h - ((value - y_min) / (y_max - y_min)) * plot_h)

    image = Image.new("RGB", (width * scale, height * scale), "#ffffff")
    draw = ImageDraw.Draw(image)

    font_small = _chart_font(8 * scale)
    font_label = _chart_font(9 * scale)
    font_legend = _chart_font(legend_layout.font_size * scale)
    font_marker = _chart_font(9 * scale, bold=True)

    for i in range(6):
        y_value = y_min + (y_max - y_min) * i / 5
        yy = py(y_value)
        draw.line([(sx(pad_left), yy), (sx(plot_right), yy)], fill="#e5eaf1", width=max(1, sx(1)))
        _draw_right_centered_text(
            draw,
            (sx(pad_left - 8), yy),
            fmt_number(y_value, 3),
            font_small,
            "#607089",
        )

    for x_value in np.linspace(x_min, x_max, 7):
        xx = px(float(x_value))
        draw.line([(xx, sx(pad_top)), (xx, sx(height - pad_bottom))], fill="#eef2f7", width=max(1, sx(1)))
        tick_text = f"{x_value:.0f}" if x_label == "Wavelength (nm)" else fmt_number(float(x_value), 4)
        _draw_centered_text(draw, (xx, sx(height - pad_bottom + 18)), tick_text, font_small, "#607089")

    axis_width = max(1, sx(1))
    draw.line([(sx(pad_left), sx(pad_top)), (sx(pad_left), sx(height - pad_bottom))], fill="#9aa8ba", width=axis_width)
    draw.line([(sx(pad_left), sx(height - pad_bottom)), (sx(plot_right), sx(height - pad_bottom))], fill="#9aa8ba", width=axis_width)
    if y_min < 0 < y_max:
        yy = py(0.0)
        draw.line([(sx(pad_left), yy), (sx(plot_right), yy)], fill="#cbd5e1", width=axis_width)

    if isc_marker is not None and np.isfinite(isc_marker[0]):
        marker_y = py(isc_marker[0])
        draw.line([(sx(pad_left - 6), marker_y), (sx(pad_left + 6), marker_y)], fill="#dc2626", width=max(2, sx(2)))
        draw.text((sx(pad_left + 10), marker_y - _text_size(draw, isc_marker[1], font_marker)[1] / 2), isc_marker[1], fill="#dc2626", font=font_marker)

    _draw_centered_text(draw, (sx(pad_left + plot_w / 2), sx(height - 12)), x_label, font_label, "#334155")

    y_label_width, y_label_height = _text_size(draw, y_label, font_label)
    label_image = Image.new("RGBA", (y_label_width + sx(8), y_label_height + sx(8)), (255, 255, 255, 0))
    label_draw = ImageDraw.Draw(label_image)
    label_draw.text((sx(4), sx(4)), y_label, fill="#334155", font=font_label)
    rotated_label = label_image.rotate(90, expand=True)
    image.paste(
        rotated_label,
        (sx(18) - rotated_label.width // 2, sx(height / 2) - rotated_label.height // 2),
        rotated_label,
    )

    for item in plot_series:
        points = []
        for x_value, y_value in zip(x, item.y_values):
            if not np.isfinite(x_value) or not np.isfinite(y_value):
                continue
            points.append((px(float(x_value)), py(float(y_value))))
        if len(points) >= 2:
            _draw_dashed_polyline(
                draw,
                points,
                fill=item.color,
                width=max(2, sx(2)),
                dash=_scaled_dash(item.dash, scale),
            )

    panel_left = pad_left + plot_w + legend_layout.panel_gap / 2
    draw.rectangle(
        [(sx(panel_left), sx(pad_top)), (sx(width - 8), sx(height - pad_bottom))],
        fill="#fbfdff",
        outline="#e2e8f0",
        width=max(1, sx(1)),
    )
    legend_x = pad_left + plot_w + legend_layout.panel_gap
    legend_y = pad_top + 8
    for idx, item in enumerate(plot_series):
        column = idx // legend_layout.rows_per_column
        row = idx % legend_layout.rows_per_column
        item_x = legend_x + column * legend_layout.column_width
        y0 = sx(legend_y + row * legend_layout.row_step)
        line_end = item_x + legend_layout.line_length
        _draw_dashed_polyline(
            draw,
            [(sx(item_x), y0), (sx(line_end), y0)],
            fill=item.color,
            width=max(2, sx(2)),
            dash=_scaled_dash(item.dash, scale),
        )
        text_x = line_end + 8
        text_width = max(16, legend_layout.column_width - legend_layout.line_length - 14)
        legend_text = _fit_text(draw, item.label, font_legend, sx(text_width))
        draw.text((sx(text_x), y0 - _text_size(draw, legend_text, font_legend)[1] / 2), legend_text, fill="#172033", font=font_legend)

    subtitle_text = _fit_text(draw, subtitle, font_small, sx(plot_w))
    draw.text((sx(pad_left), sx(12)), subtitle_text, fill="#475569", font=font_small)

    buffer = io.BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    buffer.seek(0)
    return buffer


class SpectralSPDApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("光谱 SPD 自动计算器")
        self.root.geometry("1180x820")
        self.root.minsize(1040, 720)

        self.reference_path: Path | None = None
        self.reference_df: pd.DataFrame | None = None
        self.reference_sheets: list[str] = []
        self.test_path: Path | None = None
        self.test_df: pd.DataFrame | None = None
        self.test_sheets: list[str] = []
        self.test_is_reference = False
        self.pending_test_path: Path | None = None
        self.pending_test_df: pd.DataFrame | None = None
        self.pending_test_sheets: list[str] = []
        self.editing_pending_test_source = False
        self.suppress_spectrum_select = False
        self.test_spectra: list[TestSpectrumDataset] = []
        self.active_test_label: str = ""
        self.visible_spectrum_labels: set[str] = set()
        self.highlighted_spectrum_labels: set[str] = set()
        self.legend_hitboxes: list[tuple[float, float, float, float, str | None]] = []
        self.visibility_popup: tk.Toplevel | None = None
        self.sr_path: Path | None = None
        self.sr_df: pd.DataFrame | None = None
        self.sr_sheets: list[str] = []
        self.iv_path: Path | None = None
        self.iv_curves: list[IVCurve] = []
        self.current_inputs: list[CurrentInput] = []
        self.last_iv_curve: IVCurve | None = None
        self.last_result: CalculationResult | None = None
        self.last_results_by_label: dict[str, CalculationResult] = {}
        self.last_correction: CurrentCorrectionResult | None = None
        self.last_correction_evaluations: list[CurrentCorrectionEvaluation] = []
        self.current_input_source_by_label: dict[str, str] = {}

        self._build_style()
        self._build_ui()
        self._load_default_sr()
        self._load_default_reference()

    def _build_style(self) -> None:
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        base_font = (UI_FONT_FAMILY, 9)
        bold_font = (UI_FONT_FAMILY, 9, "bold")
        style.configure(".", font=base_font, background=UI_PAGE_BG, foreground=UI_TEXT)
        style.configure("TFrame", background=UI_PAGE_BG)
        style.configure("Panel.TFrame", background=UI_PANEL_BG)
        style.configure("MetricCard.TFrame", background=UI_PANEL_BG, relief="solid", borderwidth=1)
        style.configure("TLabel", background=UI_PAGE_BG, foreground=UI_TEXT)
        style.configure("Panel.TLabel", background=UI_PANEL_BG, foreground=UI_TEXT)
        style.configure("Muted.TLabel", background=UI_PANEL_BG, foreground=UI_TEXT_MUTED)
        style.configure("Subtle.TLabel", background=UI_PANEL_BG, foreground=UI_TEXT_SUBTLE)
        style.configure("Title.TLabel", background=UI_PANEL_BG, foreground=UI_TEXT, font=(UI_FONT_FAMILY, 12, "bold"))
        style.configure("Metric.TLabel", background=UI_PANEL_BG, foreground=UI_TEXT, font=(UI_FONT_FAMILY, 16, "bold"))
        style.configure("Metric.Good.TLabel", background=UI_PANEL_BG, foreground=UI_SUCCESS, font=(UI_FONT_FAMILY, 16, "bold"))
        style.configure("Metric.Warn.TLabel", background=UI_PANEL_BG, foreground=UI_WARNING, font=(UI_FONT_FAMILY, 16, "bold"))
        style.configure("Metric.Bad.TLabel", background=UI_PANEL_BG, foreground=UI_DANGER, font=(UI_FONT_FAMILY, 16, "bold"))
        style.configure(
            "TButton",
            padding=(10, 5),
            background="#f4f6f8",
            foreground=UI_TEXT,
            bordercolor=UI_BORDER,
            focusthickness=1,
            focuscolor=UI_FOCUS,
        )
        style.map(
            "TButton",
            background=[("active", UI_GHOST_HOVER), ("pressed", UI_SECONDARY_HOVER)],
            bordercolor=[("focus", UI_FOCUS), ("active", UI_BORDER_STRONG)],
        )
        style.configure("Primary.TButton", font=bold_font, background=UI_PRIMARY, foreground="#ffffff", bordercolor=UI_PRIMARY)
        style.map(
            "Primary.TButton",
            background=[("pressed", UI_PRIMARY_PRESSED), ("active", UI_PRIMARY_HOVER), ("disabled", "#cbd5e1")],
            foreground=[("disabled", "#f8fafc"), ("!disabled", "#ffffff")],
            bordercolor=[("focus", UI_PRIMARY_PRESSED), ("active", UI_PRIMARY_HOVER)],
        )
        style.configure("Secondary.TButton", font=bold_font, background=UI_SECONDARY_BG, foreground=UI_TEXT, bordercolor=UI_BORDER_STRONG)
        style.map(
            "Secondary.TButton",
            background=[("pressed", UI_SECONDARY_HOVER), ("active", UI_SECONDARY_HOVER)],
            bordercolor=[("focus", UI_FOCUS), ("active", UI_BORDER_STRONG)],
        )
        style.configure("Ghost.TButton", background=UI_PANEL_BG, foreground=UI_TEXT_MUTED, bordercolor=UI_PANEL_BG)
        style.map(
            "Ghost.TButton",
            background=[("pressed", UI_GHOST_HOVER), ("active", UI_GHOST_HOVER)],
            foreground=[("active", UI_TEXT), ("!active", UI_TEXT_MUTED)],
            bordercolor=[("focus", UI_FOCUS), ("active", UI_BORDER)],
        )
        style.configure("Danger.TButton", font=bold_font, background=UI_DANGER_BG, foreground=UI_DANGER, bordercolor="#ffb3b8")
        style.map(
            "Danger.TButton",
            background=[("pressed", UI_DANGER_HOVER), ("active", UI_DANGER_HOVER)],
            foreground=[("disabled", UI_TEXT_SUBTLE), ("!disabled", UI_DANGER)],
            bordercolor=[("focus", UI_DANGER), ("active", UI_DANGER)],
        )
        style.configure("TEntry", fieldbackground="#fbfcfe", bordercolor=UI_BORDER, lightcolor=UI_BORDER, darkcolor=UI_BORDER)
        style.map("TEntry", bordercolor=[("focus", UI_FOCUS)])
        style.configure("TCombobox", fieldbackground="#fbfcfe", background="#fbfcfe", bordercolor=UI_BORDER, arrowcolor=UI_TEXT_MUTED)
        style.map(
            "TCombobox",
            fieldbackground=[("readonly", "#fbfcfe")],
            bordercolor=[("focus", UI_FOCUS), ("active", UI_BORDER_STRONG)],
            arrowcolor=[("active", UI_PRIMARY), ("!active", UI_TEXT_MUTED)],
        )
        style.configure("TLabelframe", background=UI_PANEL_BG, bordercolor=UI_BORDER, relief="solid")
        style.configure("TLabelframe.Label", background=UI_PANEL_BG, foreground=UI_TEXT, font=bold_font)
        style.configure("Section.TLabelframe", background=UI_PANEL_BG, bordercolor=UI_BORDER, relief="solid")
        style.configure("Section.TLabelframe.Label", background=UI_PANEL_BG, foreground=UI_TEXT, font=bold_font)
        style.configure("TNotebook", background=UI_PAGE_BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(14, 7), background="#e9eef5", foreground=UI_TEXT_MUTED, font=bold_font)
        style.map(
            "TNotebook.Tab",
            background=[("selected", UI_PANEL_BG), ("active", "#f8fafc")],
            foreground=[("selected", UI_TEXT), ("active", UI_TEXT)],
        )
        style.configure(
            "Treeview",
            rowheight=27,
            fieldbackground=UI_PANEL_BG,
            background=UI_PANEL_BG,
            foreground=UI_TEXT,
            bordercolor=UI_BORDER,
            lightcolor=UI_BORDER,
            darkcolor=UI_BORDER,
        )
        style.map(
            "Treeview",
            background=[("selected", "#d0e2ff")],
            foreground=[("selected", UI_TEXT)],
        )
        style.configure("Treeview.Heading", font=bold_font, background="#f1f5f9", foreground=UI_TEXT, bordercolor=UI_BORDER)
        style.map("Treeview.Heading", background=[("active", "#e2e8f0")])

    def _build_ui(self) -> None:
        self.root.configure(bg=UI_PAGE_BG)
        self.page_canvas = tk.Canvas(self.root, bg=UI_PAGE_BG, highlightthickness=0)
        page_scroll = ttk.Scrollbar(self.root, orient="vertical", command=self.page_canvas.yview)
        self.page_canvas.configure(yscrollcommand=page_scroll.set)
        self.page_canvas.pack(side="left", fill="both", expand=True)
        page_scroll.pack(side="right", fill="y")

        main = ttk.Frame(self.page_canvas, padding=12)
        self.page_inner = main
        self.page_window = self.page_canvas.create_window((0, 0), window=main, anchor="nw")
        self.page_canvas.bind("<Configure>", self._sync_page_scroll)
        main.bind("<Configure>", self._sync_page_scroll)
        self.root.bind_all("<MouseWheel>", self._on_page_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._on_page_mousewheel, add="+")
        self.root.bind_all("<Button-5>", self._on_page_mousewheel, add="+")

        left = ttk.Frame(main, style="Panel.TFrame", padding=12, width=340)
        left.pack(side="left", fill="y", padx=(0, 12))
        left.pack_propagate(False)

        right = ttk.Frame(main)
        right.pack(side="left", fill="both", expand=True)

        ttk.Label(left, text="光谱 SPD 计算", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            left,
            text="读取参考光谱和测试光谱表格，计算 300-1200 nm 光谱偏差。",
            style="Muted.TLabel",
            wraplength=295,
        ).pack(anchor="w", pady=(4, 14))

        actions = ttk.Frame(left, style="Panel.TFrame")
        actions.pack(fill="x", pady=(0, 12))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        ttk.Button(actions, text="开始计算", style="Primary.TButton", command=self.calculate).grid(
            row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8), ipady=5
        )
        ttk.Button(actions, text="导出结果", style="Secondary.TButton", command=self.export_result).grid(
            row=1, column=0, sticky="ew", padx=(0, 4), pady=(0, 6)
        )
        ttk.Button(actions, text="导出光谱数据", style="Secondary.TButton", command=self.export_spectrum_report).grid(
            row=1, column=1, sticky="ew", padx=(4, 0), pady=(0, 6)
        )
        ttk.Button(actions, text="读取输入文件", command=self.open_input_file).grid(row=2, column=0, sticky="ew", padx=(0, 4))
        ttk.Button(actions, text="保存输入文件", style="Ghost.TButton", command=self.save_input_file).grid(
            row=2, column=1, sticky="ew", padx=(4, 0)
        )
        ttk.Button(actions, text="恢复初始", style="Ghost.TButton", command=self.reset_to_initial_state).grid(
            row=3, column=0, columnspan=2, sticky="ew", pady=(6, 0)
        )

        self._build_reference_controls(left)
        self._build_test_controls(left)
        self._build_setting_controls(left)
        self._build_formula_panel(left)

        self.status_var = tk.StringVar(value="正在载入标准光谱...")
        ttk.Label(left, textvariable=self.status_var, style="Muted.TLabel", wraplength=295).pack(anchor="w", pady=(12, 0))

        self._build_result_area(right)

    def _sync_page_scroll(self, _event: tk.Event | None = None) -> None:
        canvas_width = max(self.page_canvas.winfo_width(), 1)
        canvas_height = max(self.page_canvas.winfo_height(), 1)
        content_height = max(self.page_inner.winfo_reqheight(), canvas_height)
        self.page_canvas.itemconfigure(self.page_window, width=canvas_width, height=content_height)
        self.page_canvas.configure(scrollregion=self.page_canvas.bbox(self.page_window))

    def _on_page_mousewheel(self, event: tk.Event) -> str | None:
        if self._is_child_scroll_target(getattr(event, "widget", None)):
            return None

        steps = self._wheel_steps(event)
        if steps == 0:
            return None

        first, last = self.page_canvas.yview()
        if first <= 0.0 and steps < 0:
            return None
        if last >= 1.0 and steps > 0:
            return None
        if last - first >= 0.999:
            return None

        self.page_canvas.yview_scroll(steps * 3, "units")
        return "break"

    def _wheel_steps(self, event: tk.Event) -> int:
        if getattr(event, "num", None) == 4:
            return -1
        if getattr(event, "num", None) == 5:
            return 1
        delta = getattr(event, "delta", 0)
        if delta == 0:
            return 0
        steps = -int(delta / 120)
        if steps == 0:
            steps = -1 if delta > 0 else 1
        return steps

    def _is_child_scroll_target(self, widget: object) -> bool:
        if widget is None:
            return False

        widget_path = str(widget)
        if "popdown" in widget_path:
            return True

        while isinstance(widget, tk.Widget):
            if isinstance(widget, (ttk.Treeview, ttk.Combobox, tk.Listbox)):
                return True
            widget = widget.master

        try:
            widget_class = self.root.tk.call("winfo", "class", widget_path)
        except tk.TclError:
            widget_class = ""
        if widget_class in {"TCombobox", "Treeview", "Listbox"}:
            return True

        return False

    def _build_reference_controls(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="参考光谱", style="Section.TLabelframe", padding=8)
        box.pack(fill="x", pady=(0, 8))

        self.ref_file_var = tk.StringVar(value="未载入")
        ttk.Label(box, textvariable=self.ref_file_var, style="Panel.TLabel", wraplength=285).pack(anchor="w")

        row = ttk.Frame(box, style="Panel.TFrame")
        row.pack(fill="x", pady=(8, 0))
        ttk.Label(row, text="标准", style="Panel.TLabel").pack(side="left")
        self.ref_kind_var = tk.StringVar(value=DEFAULT_REFERENCE_KIND)
        self.ref_kind_combo = ttk.Combobox(
            row,
            textvariable=self.ref_kind_var,
            values=list(DEFAULT_REFERENCE_CHOICES),
            state="readonly",
            width=22,
        )
        self.ref_kind_combo.pack(side="right")
        self.ref_kind_combo.bind("<<ComboboxSelected>>", lambda _event: self._update_reference_column())

        ttk.Button(box, text="选择参考表格", command=self.open_reference_file).pack(fill="x", pady=(8, 0))

    def _build_test_controls(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="测试光谱", style="Section.TLabelframe", padding=8)
        box.pack(fill="x", pady=(0, 8))

        self.test_file_var = tk.StringVar(value="未选择测试表格")
        ttk.Label(box, textvariable=self.test_file_var, style="Panel.TLabel", wraplength=285).pack(anchor="w")
        ttk.Button(box, text="选择测试表格", command=self.open_test_file).pack(fill="x", pady=(8, 8))
        ttk.Button(box, text="批量导入光谱", style="Secondary.TButton", command=self.open_test_files).pack(fill="x", pady=(0, 8))
        ttk.Button(box, text="添加当前列", style="Secondary.TButton", command=self.add_selected_test_spectrum).pack(fill="x", pady=(0, 8))
        ttk.Button(box, text="用当前参考光谱自测", style="Ghost.TButton", command=self.use_reference_as_test).pack(fill="x", pady=(0, 8))

        self.test_spectrum_var = tk.StringVar()
        self.test_name_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.wave_col_var = tk.StringVar()
        self.irrad_col_var = tk.StringVar()
        self.unit_var = tk.StringVar(value="W/(m^2 nm)")

        self.test_spectrum_combo = self._labeled_combo(box, "当前光谱", self.test_spectrum_var, [], self._change_active_test_spectrum)
        ttk.Label(box, text="光谱名称", style="Panel.TLabel").pack(anchor="w", pady=(4, 0))
        ttk.Entry(box, textvariable=self.test_name_var).pack(fill="x")
        self.sheet_combo = self._labeled_combo(box, "工作表", self.sheet_var, [], self._change_test_sheet)
        self.wave_combo = self._labeled_combo(box, "波长列", self.wave_col_var, [], self._update_active_test_config)
        self.irrad_combo = self._labeled_combo(box, "辐照度列", self.irrad_col_var, [], self._update_active_test_config)
        self.unit_combo = self._labeled_combo(box, "输入单位", self.unit_var, list(UNIT_FACTORS.keys()), self._update_active_test_config)

    def _build_setting_controls(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="计算设置", style="Section.TLabelframe", padding=8)
        box.pack(fill="x", pady=(0, 8))

        row = ttk.Frame(box, style="Panel.TFrame")
        row.pack(fill="x")
        self.start_var = tk.StringVar(value="300")
        self.end_var = tk.StringVar(value="1200")
        self.step_var = tk.StringVar(value="1")
        self._small_entry(row, "起始 nm", self.start_var)
        self._small_entry(row, "终止 nm", self.end_var)
        self._small_entry(row, "步长 nm", self.step_var)

    def _build_formula_panel(self, parent: ttk.Frame) -> None:
        box = ttk.LabelFrame(parent, text="公式", style="Section.TLabelframe", padding=8)
        box.pack(fill="x", pady=(0, 8))
        header = ttk.Frame(box, style="Panel.TFrame")
        header.pack(fill="x")
        ttk.Label(header, text="SPD / SPC / CV 判定", style="Subtle.TLabel").pack(side="left")
        self.formula_expanded = tk.BooleanVar(value=False)
        self.formula_toggle = ttk.Button(header, text="展开公式", style="Ghost.TButton", command=self._toggle_formula_panel)
        self.formula_toggle.pack(side="right")
        text = (
            "SPD% = ∫|E_test(λ)-E_ref(λ)|dλ / ∫E_ref(λ)dλ × 100%\n"
            "SPC% = 测试光谱 ≥ 10% 参考光谱的参考积分占比\n"
            "固定使用原始测试光谱，不做总辐照度归一化。\n"
            "CVi(mA) = Ii(mA) × (MG × MT) / MMF\n"
            "CV筛选(A5)：|CVi-CVmean|/CVmean ≤ 1.5%，ISC范围=(Imax-Imin)/Imean ≤ 1.5%，"
            "CV标准偏差/CVmean ≤ 1.0%；超限时剔除偏离最大的组并重复筛选，最终有效CV不少于5组。"
        )
        self.formula_content = ttk.Frame(box, style="Panel.TFrame")
        self.formula_label = ttk.Label(self.formula_content, text=text, style="Panel.TLabel", justify="left", wraplength=285)
        self.formula_label.pack(anchor="w", pady=(8, 0))
        self._sync_formula_panel()

    def _toggle_formula_panel(self) -> None:
        self.formula_expanded.set(not self.formula_expanded.get())
        self._sync_formula_panel()

    def _sync_formula_panel(self) -> None:
        if not hasattr(self, "formula_content"):
            return
        if self.formula_expanded.get():
            self.formula_toggle.configure(text="收起公式")
            self.formula_content.pack(fill="x")
        else:
            self.formula_toggle.configure(text="展开公式")
            self.formula_content.pack_forget()

    def _labeled_combo(
        self,
        parent: ttk.Frame,
        label: str,
        variable: tk.StringVar,
        values: list[str],
        callback,
    ) -> ttk.Combobox:
        row = ttk.Frame(parent, style="Panel.TFrame")
        row.pack(fill="x", pady=(4, 0))
        ttk.Label(row, text=label, style="Panel.TLabel").pack(side="left")
        combo = ttk.Combobox(row, textvariable=variable, values=values, state="readonly", width=22)
        combo.pack(side="right", fill="x", expand=True, padx=(10, 0))
        if callback:
            combo.bind("<<ComboboxSelected>>", lambda _event: callback())
        return combo

    def _small_entry(self, parent: ttk.Frame, label: str, variable: tk.StringVar) -> None:
        frame = ttk.Frame(parent, style="Panel.TFrame")
        frame.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Label(frame, text=label, style="Panel.TLabel").pack(anchor="w")
        ttk.Entry(frame, textvariable=variable, width=8).pack(fill="x")

    def _build_result_area(self, parent: ttk.Frame) -> None:
        metrics = ttk.Frame(parent)
        metrics.pack(fill="x")
        self.metric_vars = {
            "spd": tk.StringVar(value="-"),
            "spc": tk.StringVar(value="-"),
            "class": tk.StringVar(value="-"),
            "test_total": tk.StringVar(value="-"),
        }
        self._metric_card(metrics, "SPD 光谱偏差", self.metric_vars["spd"]).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._metric_card(metrics, "SPC 覆盖率", self.metric_vars["spc"]).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._metric_card(metrics, "分段等级", self.metric_vars["class"]).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._metric_card(metrics, "测试积分", self.metric_vars["test_total"]).pack(side="left", fill="x", expand=True)

        canvas_frame = ttk.Frame(parent, style="Panel.TFrame", padding=10)
        canvas_frame.pack(fill="both", expand=True, pady=(12, 12))
        header = ttk.Frame(canvas_frame, style="Panel.TFrame")
        header.pack(fill="x")
        self.plot_title_var = tk.StringVar(value="光谱曲线")
        ttk.Label(header, textvariable=self.plot_title_var, style="Title.TLabel").pack(side="left")
        self.plot_mode_var = tk.StringVar(value="参考 / 测试")
        self.plot_mode_combo = ttk.Combobox(
            header,
            textvariable=self.plot_mode_var,
            values=["参考 / 测试", "绝对偏差", "SR 曲线", "IV 曲线"],
            state="readonly",
            width=14,
        )
        self.plot_mode_combo.pack(side="right")
        self.plot_mode_combo.bind("<<ComboboxSelected>>", lambda _event: self.draw_plot())
        self.canvas = tk.Canvas(canvas_frame, bg=UI_PANEL_BG, highlightthickness=0, height=320)
        self.canvas.pack(fill="both", expand=True, pady=(8, 0))
        self.canvas.bind("<Configure>", lambda _event: self.draw_plot())
        self.canvas.bind("<Double-1>", self._on_plot_double_click)
        self.plot_status_var = tk.StringVar(value="")
        ttk.Label(canvas_frame, textvariable=self.plot_status_var, style="Muted.TLabel", wraplength=760).pack(anchor="w", pady=(6, 0))

        notebook = ttk.Notebook(parent)
        notebook.pack(fill="both", expand=True)

        band_tab = ttk.Frame(notebook, style="Panel.TFrame", padding=8)
        data_tab = ttk.Frame(notebook, style="Panel.TFrame", padding=8)
        spectrum_tab = ttk.Frame(notebook, style="Panel.TFrame", padding=8)
        correction_tab = ttk.Frame(notebook, style="Panel.TFrame", padding=8)
        notebook.add(band_tab, text="分段结果")
        notebook.add(data_tab, text="数据预览")
        notebook.add(spectrum_tab, text="光谱记录")
        notebook.add(correction_tab, text="ISC 修正")

        self.band_tree = ttk.Treeview(
            band_tab,
            columns=("range", "ref_pct", "test_pct", "ratio", "class"),
            show="headings",
            height=7,
        )
        for col, title, width in [
            ("range", "波段 nm", 120),
            ("ref_pct", "参考占比 %", 120),
            ("test_pct", "测试占比 %", 120),
            ("ratio", "匹配比", 90),
            ("class", "等级", 90),
        ]:
            self.band_tree.heading(col, text=title)
            self.band_tree.column(col, width=width, anchor="center")
        self._style_tree(self.band_tree)
        self.band_tree.pack(fill="both", expand=True)

        self.data_tree = ttk.Treeview(
            data_tab,
            columns=("wavelength", "reference", "test", "error"),
            show="headings",
            height=9,
        )
        for col, title, width in [
            ("wavelength", "波长 nm", 90),
            ("reference", "参考 W/(m^2 nm)", 150),
            ("test", "测试 W/(m^2 nm)", 150),
            ("error", "绝对偏差", 120),
        ]:
            self.data_tree.heading(col, text=title)
            self.data_tree.column(col, width=width, anchor="center")
        self._style_tree(self.data_tree)
        yscroll = ttk.Scrollbar(data_tab, orient="vertical", command=self.data_tree.yview)
        self.data_tree.configure(yscrollcommand=yscroll.set)
        self.data_tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        self._build_spectrum_records_tab(spectrum_tab)
        self._build_correction_tab(correction_tab)

    def _metric_card(self, parent: ttk.Frame, title: str, variable: tk.StringVar) -> ttk.Frame:
        frame = ttk.Frame(parent, style="MetricCard.TFrame")
        accent = tk.Frame(frame, width=4, bg=UI_BORDER_STRONG, highlightthickness=0)
        accent.pack(side="left", fill="y")
        body = ttk.Frame(frame, style="Panel.TFrame", padding=(10, 9))
        body.pack(side="left", fill="both", expand=True)
        ttk.Label(body, text=title, style="Muted.TLabel").pack(anchor="w")
        value_label = ttk.Label(body, textvariable=variable, style="Metric.TLabel")
        value_label.pack(anchor="w", pady=(4, 0))

        def refresh_metric_style(*_args: object) -> None:
            label_style, accent_color = self._metric_visual_state(title, variable.get())
            value_label.configure(style=label_style)
            accent.configure(bg=accent_color)

        variable.trace_add("write", refresh_metric_style)
        refresh_metric_style()
        return frame

    def _metric_visual_state(self, title: str, value: str) -> tuple[str, str]:
        text = value.strip().upper()
        if not text or text == "-":
            return "Metric.TLabel", UI_BORDER_STRONG
        if "分段等级" in title:
            if text.startswith("A"):
                return "Metric.Good.TLabel", UI_SUCCESS
            if text.startswith("B"):
                return "Metric.Warn.TLabel", UI_WARNING
            if text.startswith("C"):
                return "Metric.Bad.TLabel", UI_DANGER
        return "Metric.TLabel", UI_PRIMARY

    def _style_tree(self, tree: ttk.Treeview) -> None:
        tree.tag_configure("odd", background=UI_PANEL_BG)
        tree.tag_configure("even", background="#f8fafc")
        tree.tag_configure("good", background="#ecfdf3", foreground="#0e6027")
        tree.tag_configure("warn", background="#fff8e6", foreground="#7c4f0b")
        tree.tag_configure("bad", background="#fff1f1", foreground="#9f1239")

    def _build_spectrum_records_tab(self, parent: ttk.Frame) -> None:
        toolbar = ttk.Frame(parent, style="Panel.TFrame")
        toolbar.pack(fill="x", pady=(0, 6))
        ttk.Button(toolbar, text="删除选中光谱", style="Danger.TButton", command=self.delete_selected_spectrum).pack(side="left")
        ttk.Button(toolbar, text="批量显示", style="Ghost.TButton", command=self._show_spectrum_visibility_menu).pack(side="left", padx=(6, 0))
        ttk.Label(
            toolbar,
            text="空格切换显示，F2 改名，Enter 高亮；图例会同步使用这里的名称。",
            style="Muted.TLabel",
        ).pack(side="left", padx=(10, 0))

        table_frame = ttk.Frame(parent, style="Panel.TFrame")
        table_frame.pack(fill="both", expand=True)
        self.spectrum_tree = ttk.Treeview(
            table_frame,
            columns=("visible", "label", "color", "name", "file", "sheet", "wave", "irrad", "unit"),
            show="headings",
            height=8,
        )
        for col, title, width in [
            ("visible", "独立显示", 82),
            ("label", "编号", 60),
            ("color", "曲线样式", 120),
            ("name", "光谱名称", 180),
            ("file", "文件", 160),
            ("sheet", "工作表", 90),
            ("wave", "波长列", 100),
            ("irrad", "数据列", 140),
            ("unit", "单位", 110),
        ]:
            heading_options = {"text": title}
            if col == "visible":
                heading_options["command"] = self._show_spectrum_visibility_menu
            self.spectrum_tree.heading(col, **heading_options)
            self.spectrum_tree.column(col, width=width, anchor="center")
        self._style_tree(self.spectrum_tree)
        scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.spectrum_tree.yview)
        self.spectrum_tree.configure(yscrollcommand=scroll.set)
        self.spectrum_tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.spectrum_tree.bind("<<TreeviewSelect>>", lambda _event: self._select_spectrum_record())
        self.spectrum_tree.bind("<Button-1>", self._on_spectrum_record_click)
        self.spectrum_tree.bind("<Double-1>", self._on_spectrum_record_double_click)
        self.spectrum_tree.bind("<space>", self._on_spectrum_visibility_key)
        self.spectrum_tree.bind("<F2>", self._on_spectrum_rename_key)
        self.spectrum_tree.bind("<Return>", self._on_spectrum_highlight_key)

    def _spectrum_tree_column_name(self, column_id: str) -> str:
        if not column_id.startswith("#"):
            return ""
        try:
            index = int(column_id[1:]) - 1
        except ValueError:
            return ""
        columns = tuple(self.spectrum_tree["columns"])
        if 0 <= index < len(columns):
            return str(columns[index])
        return ""

    def _on_spectrum_record_click(self, event: tk.Event) -> str | None:
        if self.spectrum_tree.identify_region(event.x, event.y) != "cell":
            return None
        if self._spectrum_tree_column_name(self.spectrum_tree.identify_column(event.x)) != "visible":
            return None
        label = self.spectrum_tree.identify_row(event.y)
        if label:
            self._toggle_spectrum_visibility(str(label))
        return "break"

    def _on_spectrum_record_double_click(self, event: tk.Event) -> str | None:
        if self.spectrum_tree.identify_region(event.x, event.y) != "cell":
            return None
        if self._spectrum_tree_column_name(self.spectrum_tree.identify_column(event.x)) != "name":
            return "break"
        label = self.spectrum_tree.identify_row(event.y)
        if label:
            self.spectrum_tree.selection_set(label)
            self.rename_selected_spectrum()
        return "break"

    def _focused_spectrum_label(self) -> str:
        label = str(self.spectrum_tree.focus() or "")
        if label:
            return label
        selection = self.spectrum_tree.selection()
        return str(selection[0]) if selection else ""

    def _on_spectrum_visibility_key(self, _event: tk.Event) -> str:
        label = self._focused_spectrum_label()
        if label:
            self._toggle_spectrum_visibility(label)
        return "break"

    def _on_spectrum_rename_key(self, _event: tk.Event) -> str:
        label = self._focused_spectrum_label()
        if label:
            self.spectrum_tree.selection_set(label)
            self.rename_selected_spectrum()
        return "break"

    def _on_spectrum_highlight_key(self, _event: tk.Event) -> str:
        label = self._focused_spectrum_label()
        if not label or self._test_spectrum_by_label(label) is None:
            return "break"
        if label in self.highlighted_spectrum_labels:
            self.highlighted_spectrum_labels.remove(label)
        elif self._is_spectrum_visible(label):
            self.highlighted_spectrum_labels.add(label)
        self._prune_spectrum_display_state()
        self.draw_plot()
        return "break"

    def _test_spectrum_display(self, dataset: TestSpectrumDataset) -> str:
        return f"{dataset.label} {dataset.name}"

    def _label_from_display(self, value: str) -> str:
        return value.split(" ", 1)[0] if value else ""

    def _active_test_spectrum(self) -> TestSpectrumDataset | None:
        label = self.active_test_label or self._label_from_display(self.test_spectrum_var.get())
        return self._test_spectrum_by_label(label)

    def _test_spectrum_by_label(self, label: str) -> TestSpectrumDataset | None:
        for dataset in self.test_spectra:
            if dataset.label == label:
                return dataset
        return None

    def _spectrum_color_for_label(self, label: str) -> str:
        for index, dataset in enumerate(self.test_spectra):
            if dataset.label == label:
                return spectrum_color_for_index(index)
        return spectrum_color_for_index(0)

    def _spectrum_color_for_dataset(self, dataset: TestSpectrumDataset) -> str:
        return self._spectrum_color_for_label(dataset.label)

    def _spectrum_style_for_label(self, label: str) -> PlotSeriesStyle:
        for index, dataset in enumerate(self.test_spectra):
            if dataset.label == label:
                return spectrum_style_for_index(index)
        return spectrum_style_for_index(0)

    def _spectrum_style_for_dataset(self, dataset: TestSpectrumDataset) -> PlotSeriesStyle:
        return self._spectrum_style_for_label(dataset.label)

    def _spectrum_style_text_for_dataset(self, dataset: TestSpectrumDataset) -> str:
        for index, item in enumerate(self.test_spectra):
            if item.label == dataset.label:
                style = spectrum_style_for_index(index)
                return f"{style.color} {spectrum_line_name_for_index(index)}"
        style = spectrum_style_for_index(0)
        return f"{style.color} {spectrum_line_name_for_index(0)}"

    def _plot_series_for_dataset(self, dataset: TestSpectrumDataset, y_values: np.ndarray) -> PlotSeriesData:
        style = self._spectrum_style_for_dataset(dataset)
        return plot_series_data(
            y_values,
            style.color,
            dataset.name,
            dash=style.dash,
            spectrum_label=dataset.label,
        )

    def _plot_series_for_label(self, label: str, name: str, y_values: np.ndarray) -> PlotSeriesData:
        style = self._spectrum_style_for_label(label)
        return plot_series_data(
            y_values,
            style.color,
            name,
            dash=style.dash,
            spectrum_label=label,
        )

    def _plot_reference_series(self, y_values: np.ndarray, label: str) -> PlotSeriesData:
        return plot_series_data(
            y_values,
            REFERENCE_SERIES_COLOR,
            label,
            is_reference=True,
        )

    def _current_spectrum_labels(self) -> set[str]:
        return {dataset.label for dataset in self.test_spectra}

    def _prune_spectrum_display_state(self) -> None:
        labels = self._current_spectrum_labels()
        if not hasattr(self, "visible_spectrum_labels"):
            self.visible_spectrum_labels = set(labels)
        if not hasattr(self, "highlighted_spectrum_labels"):
            self.highlighted_spectrum_labels = set()
        self.visible_spectrum_labels.intersection_update(labels)
        self.highlighted_spectrum_labels.intersection_update(self.visible_spectrum_labels)

    def _mark_spectrum_visible(self, label: str) -> None:
        if not hasattr(self, "visible_spectrum_labels"):
            self.visible_spectrum_labels = set()
        if label:
            self.visible_spectrum_labels.add(label)

    def _is_spectrum_visible(self, label: str) -> bool:
        visible_labels = getattr(self, "visible_spectrum_labels", None)
        if visible_labels is None:
            return True
        return label in visible_labels

    def _toggle_spectrum_visibility(self, label: str) -> None:
        dataset = self._test_spectrum_by_label(label)
        if dataset is None:
            return
        if self._is_spectrum_visible(label):
            self.visible_spectrum_labels.discard(label)
            self.highlighted_spectrum_labels.discard(label)
        else:
            self.visible_spectrum_labels.add(label)
        self._refresh_spectrum_records(selected_label=label)
        self.draw_plot()

    def _set_all_spectrum_visibility(self, visible: bool) -> None:
        labels = self._current_spectrum_labels()
        self.visible_spectrum_labels = set(labels) if visible else set()
        self.highlighted_spectrum_labels.intersection_update(self.visible_spectrum_labels)
        self._refresh_spectrum_records(selected_label=self.active_test_label)
        self.draw_plot()

    def _hide_visibility_popup(self) -> None:
        popup = getattr(self, "visibility_popup", None)
        if popup is None:
            return
        try:
            if popup.winfo_exists():
                popup.destroy()
        except Exception:
            pass
        self.visibility_popup = None

    def _show_spectrum_visibility_menu(self) -> None:
        if not hasattr(self, "root"):
            return
        self._hide_visibility_popup()
        popup = tk.Toplevel(self.root)
        popup.withdraw()
        popup.overrideredirect(True)
        popup.configure(bg="#ffffff")
        frame = ttk.Frame(popup, padding=6, style="Panel.TFrame")
        frame.pack(fill="both", expand=True)

        def apply_choice(visible: bool) -> None:
            self._hide_visibility_popup()
            self._set_all_spectrum_visibility(visible)

        select_button = ttk.Button(frame, text="全选", command=lambda: apply_choice(True))
        select_button.pack(side="left", padx=(0, 6))
        clear_button = ttk.Button(frame, text="全不选", command=lambda: apply_choice(False))
        clear_button.pack(side="left")
        x = self.root.winfo_pointerx()
        y = self.root.winfo_pointery()
        popup.geometry(f"+{x}+{y}")
        popup.deiconify()
        popup.lift()
        select_button.focus_set()
        popup.bind("<Escape>", lambda _event: self._hide_visibility_popup())
        frame.bind("<Escape>", lambda _event: self._hide_visibility_popup())
        select_button.bind("<Escape>", lambda _event: self._hide_visibility_popup())
        clear_button.bind("<Escape>", lambda _event: self._hide_visibility_popup())
        popup.bind("<FocusOut>", lambda _event: self._hide_visibility_popup())
        self.visibility_popup = popup

    def _renumber_test_spectra(self) -> dict[str, str]:
        label_map: dict[str, str] = {}
        for index, dataset in enumerate(self.test_spectra, start=1):
            old_label = dataset.label
            new_label = f"S{index}"
            label_map[old_label] = new_label
            dataset.label = new_label
        if hasattr(self, "visible_spectrum_labels"):
            self.visible_spectrum_labels = {
                label_map.get(label, label)
                for label in self.visible_spectrum_labels
            }
        if hasattr(self, "highlighted_spectrum_labels"):
            self.highlighted_spectrum_labels = {
                label_map.get(label, label)
                for label in self.highlighted_spectrum_labels
            }
        self._prune_spectrum_display_state()
        if all(old_label == new_label for old_label, new_label in label_map.items()):
            return label_map

        for current in self.current_inputs:
            if current.spectrum_label in label_map:
                current.spectrum_label = label_map[current.spectrum_label]

        self.active_test_label = label_map.get(self.active_test_label, self.active_test_label)
        if self.last_results_by_label:
            self.last_results_by_label = {
                label_map.get(label, label): result
                for label, result in self.last_results_by_label.items()
                if label in label_map or self._test_spectrum_by_label(label) is not None
            }
        for evaluation in self.last_correction_evaluations:
            if evaluation.spectrum_label in label_map:
                evaluation.spectrum_label = label_map[evaluation.spectrum_label]
        return label_map

    def _active_calculation_result(self) -> CalculationResult | None:
        if self.last_results_by_label:
            label = self.active_test_label or self._label_from_display(self.test_spectrum_var.get())
            result = self.last_results_by_label.get(label)
            if result is not None:
                return result
        return self.last_result

    def _refresh_test_spectrum_controls(self, selected_label: str | None = None, *, sync_fields: bool = True) -> None:
        if not hasattr(self, "test_spectrum_combo"):
            return
        values = [self._test_spectrum_display(dataset) for dataset in self.test_spectra]
        self.test_spectrum_combo.configure(values=values)
        if selected_label:
            self.active_test_label = selected_label
        elif self.active_test_label not in {dataset.label for dataset in self.test_spectra}:
            self.active_test_label = self.test_spectra[0].label if self.test_spectra else ""
        active = self._active_test_spectrum()
        if active is None:
            self.test_spectrum_var.set("")
            if self.editing_pending_test_source and self.pending_test_path is not None:
                self.test_file_var.set(f"待添加：{self.pending_test_path.name}")
            else:
                self.test_file_var.set("未选择测试表格")
            self._refresh_current_spectrum_choices()
            self._refresh_spectrum_records(selected_label="")
            return
        self.test_spectrum_var.set(self._test_spectrum_display(active))
        if sync_fields:
            self.editing_pending_test_source = False
            self._sync_active_test_fields()
        self._refresh_current_spectrum_choices()
        record_selection = self.active_test_label if sync_fields or not self.editing_pending_test_source else ""
        self._refresh_spectrum_records(selected_label=record_selection)

    def _sync_active_test_fields(self) -> None:
        dataset = self._active_test_spectrum()
        if dataset is None:
            return
        self.test_path = dataset.path
        self.test_df = dataset.df
        self.sheet_var.set(dataset.sheet_name)
        self.wave_col_var.set(dataset.wavelength_column)
        self.irrad_col_var.set(dataset.irradiance_column)
        self.unit_var.set(dataset.unit_name)
        self.test_file_var.set(f"已载入 {len(self.test_spectra)} 组测试光谱")
        columns = normalized_columns(dataset.df)
        for combo in (self.wave_combo, self.irrad_combo):
            combo.configure(values=columns)
        if dataset.path is not None and dataset.path.suffix.lower() in EXCEL_TABLE_SUFFIXES:
            try:
                self.test_sheets = list_excel_sheets(dataset.path)
            except Exception:
                self.test_sheets = [dataset.sheet_name] if dataset.sheet_name else []
        else:
            self.test_sheets = []
        self.sheet_combo_values(self.test_sheets, dataset.sheet_name if self.test_sheets else "")

    def _update_active_test_config(self) -> None:
        if self.editing_pending_test_source:
            self.test_df = self.pending_test_df
            return
        dataset = self._active_test_spectrum()
        if dataset is None:
            return
        dataset.wavelength_column = self.wave_col_var.get()
        dataset.irradiance_column = self.irrad_col_var.get()
        dataset.unit_name = self.unit_var.get()
        self._clear_correction()
        self.last_result = None
        self.last_results_by_label = {}
        self.draw_plot()

    def _change_active_test_spectrum(self) -> None:
        label = self._label_from_display(self.test_spectrum_var.get())
        if not label:
            return
        self.editing_pending_test_source = False
        self.active_test_label = label
        self._sync_active_test_fields()
        result = self.last_results_by_label.get(label)
        if result is not None:
            self.last_result = result
            self._render_result(result)
        else:
            self.draw_plot()

    def _refresh_spectrum_records(self, selected_label: str | None = None) -> None:
        if not hasattr(self, "spectrum_tree"):
            return
        if selected_label is None:
            selected_label = self.active_test_label
        self.suppress_spectrum_select = True
        try:
            self._prune_spectrum_display_state()
            self.spectrum_tree.delete(*self.spectrum_tree.get_children())
            for row_index, dataset in enumerate(self.test_spectra):
                self.spectrum_tree.insert(
                    "",
                    "end",
                    iid=dataset.label,
                    values=(
                        VISIBLE_CHECKED_TEXT if self._is_spectrum_visible(dataset.label) else VISIBLE_UNCHECKED_TEXT,
                        dataset.label,
                        self._spectrum_style_text_for_dataset(dataset),
                        dataset.name,
                        dataset.path.name if dataset.path is not None else "",
                        dataset.sheet_name,
                        dataset.wavelength_column,
                        dataset.irradiance_column,
                        dataset.unit_name,
                    ),
                    tags=("even" if row_index % 2 else "odd",),
                )
            if selected_label and selected_label in self.spectrum_tree.get_children():
                self.spectrum_tree.selection_set(selected_label)
            else:
                current_selection = self.spectrum_tree.selection()
                if current_selection:
                    self.spectrum_tree.selection_remove(*current_selection)
        finally:
            self.suppress_spectrum_select = False

    def _select_spectrum_record(self) -> None:
        if self.suppress_spectrum_select:
            return
        selected = self.spectrum_tree.selection()
        if not selected:
            return
        label = str(selected[0])
        dataset = self._test_spectrum_by_label(label)
        if dataset is None:
            return
        self.editing_pending_test_source = False
        self.active_test_label = label
        self.test_spectrum_var.set(self._test_spectrum_display(dataset))
        self._sync_active_test_fields()
        result = self.last_results_by_label.get(label)
        if result is not None:
            self.last_result = result
            self._render_result(result)
        else:
            self.draw_plot()

    def rename_selected_spectrum(self) -> None:
        selected = self.spectrum_tree.selection() if hasattr(self, "spectrum_tree") else ()
        if not selected:
            return
        dataset = self._test_spectrum_by_label(str(selected[0]))
        if dataset is None:
            return
        new_name = simpledialog.askstring("修改光谱名称", "请输入新的光谱名称：", initialvalue=dataset.name)
        if new_name is None:
            return
        new_name = new_name.strip()
        if not new_name:
            return
        dataset.name = new_name
        result = self.last_results_by_label.get(dataset.label)
        if result is not None:
            result.test_name = new_name
        for evaluation in self.last_correction_evaluations:
            if evaluation.spectrum_label == dataset.label:
                evaluation.spectrum_name = new_name
        self._refresh_test_spectrum_controls(selected_label=dataset.label)
        self._refresh_current_inputs_view(self.last_correction_evaluations)
        self.draw_plot()

    def delete_selected_spectrum(self) -> None:
        selected = self.spectrum_tree.selection() if hasattr(self, "spectrum_tree") else ()
        if selected:
            label = str(selected[0])
        else:
            label = self.active_test_label or self._label_from_display(self.test_spectrum_var.get())
        if not label:
            return
        dataset = self._test_spectrum_by_label(label)
        if dataset is None:
            return
        if not messagebox.askyesno("删除光谱", f"确定删除光谱：{dataset.name}？"):
            return
        self.test_spectra = [item for item in self.test_spectra if item.label != label]
        self.visible_spectrum_labels.discard(label)
        self.highlighted_spectrum_labels.discard(label)
        self.last_results_by_label.pop(label, None)
        if self.last_result is not None and self.last_result.test_name == dataset.name:
            self.last_result = None
        fallback_label = self.test_spectra[0].label if self.test_spectra else ""
        for current in self.current_inputs:
            if current.spectrum_label == label:
                current.spectrum_label = fallback_label
        self._renumber_test_spectra()
        fallback_label = self.test_spectra[0].label if self.test_spectra else ""
        self.active_test_label = fallback_label
        self._refresh_test_spectrum_controls(selected_label=fallback_label)
        self._refresh_current_spectrum_choices()
        self._refresh_current_inputs_view()
        self._clear_correction()
        self.draw_plot()
        self.status_var.set(f"已删除光谱：{dataset.name}")

    def _refresh_plot_modes(self, select_mode: str | None = None) -> None:
        if not hasattr(self, "plot_mode_combo"):
            return
        modes = ["参考 / 测试", "绝对偏差", "SR 曲线", "IV 曲线"]
        self.plot_mode_combo.configure(values=modes)
        current = self.plot_mode_var.get()
        if select_mode in modes:
            self.plot_mode_var.set(select_mode)
        elif current not in modes:
            self.plot_mode_var.set(modes[0])

    def _iv_curve_for_mode(self, mode: str) -> IVCurve | None:
        if mode != "IV 曲线" or not self.iv_curves:
            return None
        current = self._current_input_by_label(self.current_group_var.get())
        if current is not None and current.iv_curve is not None:
            return current.iv_curve
        return None

    def _build_correction_tab(self, parent: ttk.Frame) -> None:
        controls = ttk.Frame(parent, style="Panel.TFrame", padding=10)
        controls.pack(fill="x")

        file_row = ttk.Frame(controls, style="Panel.TFrame")
        file_row.pack(fill="x")
        self.sr_file_var = tk.StringVar(value="未选择 SR 文件")
        ttk.Label(file_row, textvariable=self.sr_file_var, style="Panel.TLabel", wraplength=620).pack(side="left", fill="x", expand=True)
        ttk.Button(file_row, text="选择光谱响应文件", command=self.open_sr_file).pack(side="right", padx=(10, 0))

        input_row = ttk.Frame(controls, style="Panel.TFrame")
        input_row.pack(fill="x", pady=(8, 2))
        self.isc_var = tk.StringVar()
        self.temperature_var = tk.StringVar(value=f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}")
        self.temp_coeff_var = tk.StringVar(value=f"{DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C:g}")
        isc_frame = ttk.Frame(input_row, style="Panel.TFrame")
        isc_frame.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Label(isc_frame, text="I1,I2... (mA)", style="Panel.TLabel").pack(anchor="w")
        isc_line = ttk.Frame(isc_frame, style="Panel.TFrame")
        isc_line.pack(fill="x")
        ttk.Entry(isc_line, textvariable=self.isc_var, width=24).pack(side="left", fill="x", expand=True)
        ttk.Button(isc_line, text="添加 I", style="Secondary.TButton", command=self.add_manual_current).pack(side="left", padx=(6, 0))
        ttk.Button(isc_line, text="添加 IV", style="Secondary.TButton", command=self.open_iv_file).pack(side="left", padx=(6, 0))
        self._small_entry(input_row, "默认温度 ℃", self.temperature_var)
        self._small_entry(input_row, "温度系数 %/℃", self.temp_coeff_var)
        self.iv_file_var = tk.StringVar(value="未读取 IV 表格")
        ttk.Label(controls, textvariable=self.iv_file_var, style="Muted.TLabel", wraplength=760).pack(anchor="w", pady=(2, 0))

        assign_row = ttk.Frame(controls, style="Panel.TFrame")
        assign_row.pack(fill="x", pady=(8, 0))
        self.current_group_var = tk.StringVar()
        self.current_spectrum_var = tk.StringVar()
        self.current_temp_edit_var = tk.StringVar(value=f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}")
        self.current_group_combo = self._labeled_combo(assign_row, "电流组", self.current_group_var, [], self._select_current_input)
        self.current_spectrum_combo = self._labeled_combo(assign_row, "使用光谱", self.current_spectrum_var, [], None)
        temp_frame = ttk.Frame(assign_row, style="Panel.TFrame")
        temp_frame.pack(side="left", fill="x", expand=True, padx=(8, 0))
        ttk.Label(temp_frame, text="温度 ℃", style="Panel.TLabel").pack(anchor="w")
        ttk.Entry(temp_frame, textvariable=self.current_temp_edit_var, width=8).pack(fill="x")
        ttk.Button(assign_row, text="应用设置", style="Secondary.TButton", command=self.apply_current_settings).pack(side="left", padx=(8, 0), pady=(18, 0))
        ttk.Button(assign_row, text="删除组", style="Danger.TButton", command=self.delete_current_input).pack(side="left", padx=(6, 0), pady=(18, 0))

        self.sr_sheet_var = tk.StringVar()
        self.sr_wave_col_var = tk.StringVar()
        self.sr_col_var = tk.StringVar()
        self.sr_sheet_combo = self._labeled_combo(controls, "SR 工作表", self.sr_sheet_var, [], self._change_sr_sheet)
        self.sr_wave_combo = self._labeled_combo(controls, "SR 波长列", self.sr_wave_col_var, [], None)
        self.sr_col_combo = self._labeled_combo(controls, "SR 列", self.sr_col_var, [], None)
        ttk.Button(controls, text="计算 ISC 修正", style="Primary.TButton", command=self.calculate_correction).pack(fill="x", pady=(8, 0), ipady=4)

        metrics = ttk.Frame(parent)
        metrics.pack(fill="x", pady=(8, 8))
        self.correction_vars = {
            "mg": tk.StringVar(value="-"),
            "mt": tk.StringVar(value="-"),
            "mmf": tk.StringVar(value="-"),
            "cv": tk.StringVar(value="-"),
        }
        self._metric_card(metrics, "MG", self.correction_vars["mg"]).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._metric_card(metrics, "MT", self.correction_vars["mt"]).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._metric_card(metrics, "MMF", self.correction_vars["mmf"]).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._metric_card(metrics, "有效平均 CV (mA)", self.correction_vars["cv"]).pack(side="left", fill="x", expand=True)

        self.correction_detail_var = tk.StringVar(value="载入 SR 并读取 IV 或输入 ISC 后计算。")
        ttk.Label(parent, textvariable=self.correction_detail_var, style="Muted.TLabel", wraplength=760).pack(anchor="w", pady=(0, 6))

        cv_frame = ttk.Frame(parent, style="Panel.TFrame")
        cv_frame.pack(fill="x", pady=(0, 8))
        self.cv_tree = ttk.Treeview(
            cv_frame,
            columns=("label", "isc", "spectrum", "temperature", "cv", "deviation", "status", "source", "mg", "mt", "mmf"),
            show="headings",
            height=6,
        )
        for col, title, width in [
            ("label", "组别", 70),
            ("isc", "Ii (mA)", 100),
            ("spectrum", "光谱", 130),
            ("temperature", "温度 ℃", 80),
            ("cv", "CVi (mA)", 115),
            ("deviation", "偏离 %", 80),
            ("status", "A5 判定", 140),
            ("source", "来源", 160),
            ("mg", "MG", 90),
            ("mt", "MT", 90),
            ("mmf", "MMF", 90),
        ]:
            self.cv_tree.heading(col, text=title)
            self.cv_tree.column(col, width=width, minwidth=width, anchor="center", stretch=False)
        self._style_tree(self.cv_tree)
        self.cv_tree.tag_configure("valid", background="#ecfdf3", foreground="#0e6027")
        self.cv_tree.tag_configure("invalid", background="#fff1f1", foreground="#9f1239")
        cv_scroll = ttk.Scrollbar(cv_frame, orient="vertical", command=self.cv_tree.yview)
        cv_x_scroll = ttk.Scrollbar(cv_frame, orient="horizontal", command=self.cv_tree.xview)
        self.cv_tree.configure(yscrollcommand=cv_scroll.set, xscrollcommand=cv_x_scroll.set)
        cv_frame.columnconfigure(0, weight=1)
        cv_frame.rowconfigure(0, weight=1)
        self.cv_tree.grid(row=0, column=0, sticky="ew")
        cv_scroll.grid(row=0, column=1, sticky="ns")
        cv_x_scroll.grid(row=1, column=0, sticky="ew")
        self.cv_tree.bind("<<TreeviewSelect>>", lambda _event: self._select_current_input_from_tree())

        table_frame = ttk.Frame(parent, style="Panel.TFrame")
        table_frame.pack(fill="both", expand=True)
        self.correction_tree = ttk.Treeview(
            table_frame,
            columns=("wavelength", "sr", "reference", "test", "ref_sr", "test_sr"),
            show="headings",
            height=7,
        )
        for col, title, width in [
            ("wavelength", "波长 nm", 90),
            ("sr", "SR", 90),
            ("reference", "标准光谱", 120),
            ("test", "实测光谱", 120),
            ("ref_sr", "标准×SR", 120),
            ("test_sr", "实测×SR", 120),
        ]:
            self.correction_tree.heading(col, text=title)
            self.correction_tree.column(col, width=width, anchor="center")
        self._style_tree(self.correction_tree)
        correction_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.correction_tree.yview)
        self.correction_tree.configure(yscrollcommand=correction_scroll.set)
        self.correction_tree.pack(side="left", fill="both", expand=True)
        correction_scroll.pack(side="right", fill="y")

    def _load_default_sr(self) -> None:
        path = find_default_sr_file()
        if path is None:
            return
        try:
            self._load_sr_path(path)
        except Exception:
            self.sr_file_var.set("默认 SR 文件读取失败，请手动选择。")

    def _load_default_reference(self) -> None:
        path = find_default_reference_file()
        if path is None:
            self.status_var.set("未找到 AM1.5 标准光谱文件，请手动选择参考表格。")
            return
        try:
            self._load_reference_path(path)
        except Exception as exc:
            self.status_var.set(f"载入标准光谱失败：{exc}")

    def reset_to_initial_state(self) -> None:
        self.reference_path = None
        self.reference_df = None
        self.reference_sheets = []
        self.test_path = None
        self.test_df = None
        self.test_sheets = []
        self.test_is_reference = False
        self.pending_test_path = None
        self.pending_test_df = None
        self.pending_test_sheets = []
        self.editing_pending_test_source = False
        self.test_spectra = []
        self.active_test_label = ""
        self.visible_spectrum_labels.clear()
        self.highlighted_spectrum_labels.clear()
        self.legend_hitboxes = []
        self._hide_visibility_popup()
        self.sr_path = None
        self.sr_df = None
        self.sr_sheets = []
        self.iv_path = None
        self.iv_curves = []
        self.current_inputs = []
        self.last_iv_curve = None
        self.current_input_source_by_label = {}

        self.ref_kind_combo.configure(values=list(DEFAULT_REFERENCE_CHOICES))
        self.ref_kind_var.set(DEFAULT_REFERENCE_KIND)
        self.ref_file_var.set("未载入")
        self.test_file_var.set("未选择测试表格")
        self.test_spectrum_var.set("")
        self.test_name_var.set("")
        self.sheet_combo_values([], "")
        self.wave_col_var.set("")
        self.irrad_col_var.set("")
        self.unit_var.set("W/(m^2 nm)")
        for combo in (self.wave_combo, self.irrad_combo):
            combo.configure(values=[])

        self.start_var.set("300")
        self.end_var.set("1200")
        self.step_var.set("1")
        self.plot_mode_var.set("参考 / 测试")
        self.plot_title_var.set("光谱曲线")

        self.sr_file_var.set("未选择 SR 文件")
        self.sr_sheet_combo_values([], "")
        self.sr_wave_col_var.set("")
        self.sr_col_var.set("")
        for combo in (self.sr_wave_combo, self.sr_col_combo):
            combo.configure(values=[])

        self.isc_var.set("")
        self.temperature_var.set(f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}")
        self.temp_coeff_var.set(f"{DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C:g}")
        self.iv_file_var.set("未读取 IV 表格")
        self.current_group_var.set("")
        self.current_spectrum_var.set("")
        self.current_temp_edit_var.set(f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}")
        self.current_group_combo.configure(values=[])
        self.current_spectrum_combo.configure(values=[])

        self._clear_result_display()
        self._refresh_test_spectrum_controls(selected_label="")
        self._refresh_spectrum_records(selected_label="")
        self._refresh_current_spectrum_choices()
        self._refresh_current_inputs_view()
        self._clear_correction()
        self._refresh_plot_modes(select_mode="参考 / 测试")

        self.status_var.set("正在载入标准光谱...")
        self._load_default_sr()
        self._load_default_reference()
        if self.reference_df is not None:
            self.status_var.set("已恢复初始状态。")

    def _load_reference_path(self, path: Path) -> None:
        preferred_reference = self.ref_kind_var.get()
        self.reference_path = path
        if path.suffix.lower() in EXCEL_TABLE_SUFFIXES:
            self.reference_sheets = list_excel_sheets(path)
            sheet = self.reference_sheets[0]
            self.reference_df = read_table(path, sheet)
        else:
            self.reference_sheets = []
            self.reference_df = read_table(path)
        self.reference_df.columns = normalized_columns(self.reference_df)
        self.ref_file_var.set(path.name)
        self._refresh_reference_choices(preferred=preferred_reference)
        self._update_reference_column()
        self._clear_correction()
        self.status_var.set("已载入参考光谱。")

    def _refresh_reference_choices(self, preferred: str | None = None) -> None:
        if not hasattr(self, "ref_kind_combo"):
            return
        choices = reference_irradiance_choices(self.reference_df)
        self.ref_kind_combo.configure(values=choices)
        if not choices:
            self.ref_kind_var.set("")
            return

        preferred_text = str(preferred or "").strip()
        current_text = self.ref_kind_var.get().strip()
        if preferred_text in choices:
            selected = preferred_text
        elif current_text in choices:
            selected = current_text
        else:
            guessed = guess_irradiance_column(choices, preferred_text or current_text)
            selected = guessed if guessed in choices else choices[0]
        self.ref_kind_var.set(selected)

    def _update_reference_column(self) -> None:
        if self.reference_df is None:
            return
        columns = normalized_columns(self.reference_df)
        col = guess_irradiance_column(columns, self.ref_kind_var.get())
        if col is None:
            self.status_var.set("参考表格中没有找到可用的辐照度列。")
        else:
            if self.test_is_reference and self.test_df is not None:
                self.irrad_col_var.set(col)
                active = self._active_test_spectrum()
                if active is not None:
                    active.irradiance_column = col
            self.status_var.set(f"参考光谱列：{col}")

    def open_reference_file(self) -> None:
        path_str = filedialog.askopenfilename(
            title="选择参考光谱表格",
            initialdir=str(APP_DIR),
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.xls *.csv *.txt *.tsv *.asd"),
                ("所有文件", "*.*"),
            ],
        )
        if not path_str:
            return
        try:
            self._load_reference_path(Path(path_str))
        except Exception as exc:
            messagebox.showerror("读取失败", str(exc))

    def open_test_file(self) -> None:
        path_str = filedialog.askopenfilename(
            title="添加测试光谱表格",
            initialdir=str(APP_DIR),
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.xls *.csv *.txt *.tsv *.asd"),
                ("所有文件", "*.*"),
            ],
        )
        if not path_str:
            return
        try:
            self._load_test_source_path(Path(path_str))
        except Exception as exc:
            messagebox.showerror("读取失败", str(exc))

    def open_test_files(self) -> None:
        path_values = filedialog.askopenfilenames(
            title="批量导入测试光谱",
            initialdir=str(APP_DIR),
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.xls *.csv *.txt *.tsv *.asd"),
                ("所有文件", "*.*"),
            ],
        )
        paths = [Path(value) for value in path_values]
        if not paths:
            return
        self._load_test_batch_paths(paths)

    def open_sr_file(self) -> None:
        path_str = filedialog.askopenfilename(
            title="选择光谱响应 SR 表格",
            initialdir=str(APP_DIR),
            filetypes=[
                ("表格文件", "*.xlsx *.xlsm *.xls *.csv *.txt *.tsv *.asd"),
                ("所有文件", "*.*"),
            ],
        )
        if not path_str:
            return
        try:
            self._load_sr_path(Path(path_str))
        except Exception as exc:
            messagebox.showerror("读取失败", str(exc))

    def open_iv_file(self) -> None:
        path_values = filedialog.askopenfilenames(
            title="添加实测 IV 表格",
            initialdir=str(APP_DIR),
            filetypes=[
                ("IV 表格文件", "*.xlsx *.xlsm *.xls *.csv *.txt *.tsv *.asd"),
                ("所有文件", "*.*"),
            ],
        )
        paths = [Path(value) for value in path_values]
        if not paths:
            return
        try:
            failures = self._load_iv_paths(paths)
            if failures:
                preview = "\n".join(failures[:8])
                suffix = "\n..." if len(failures) > 8 else ""
                messagebox.showwarning("部分 IV 未读取", f"{preview}{suffix}")
        except Exception as exc:
            messagebox.showerror("读取 IV 失败", str(exc))

    def open_input_file(self) -> None:
        path_str = filedialog.askopenfilename(
            title="读取 SPD 输入文件",
            initialdir=str(APP_DIR),
            filetypes=[
                ("SPD 输入文件", "*.inp"),
                ("所有文件", "*.*"),
            ],
        )
        if not path_str:
            return
        path = Path(path_str)
        try:
            self._load_input_state_path(path)
            messagebox.showinfo("读取完成", f"输入文件已读取：\n{path}")
        except Exception as exc:
            messagebox.showerror("读取输入文件失败", str(exc))

    def save_input_file(self) -> None:
        path_str = filedialog.asksaveasfilename(
            title="保存 SPD 输入文件",
            initialdir=str(APP_DIR),
            initialfile="SPD计算输入.inp",
            defaultextension=".inp",
            filetypes=[
                ("SPD 输入文件", "*.inp"),
                ("所有文件", "*.*"),
            ],
        )
        if not path_str:
            return
        path = self._normalized_input_path(Path(path_str))
        try:
            self._save_input_state_path(path)
            self.status_var.set(f"已保存输入文件：{path.name}")
            messagebox.showinfo("保存完成", f"输入文件已保存到：\n{path}")
        except Exception as exc:
            messagebox.showerror("保存输入文件失败", str(exc))

    def _normalized_input_path(self, path: Path) -> Path:
        return path if path.suffix.lower() == ".inp" else path.with_suffix(".inp")

    def _save_input_state_path(self, path: Path) -> Path:
        path = self._normalized_input_path(path)
        payload = self._build_input_state()
        with path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2, allow_nan=False)
        return path

    def _load_input_state_path(self, path: Path) -> None:
        with path.open("r", encoding="utf-8-sig") as handle:
            payload = json.load(handle)
        if not isinstance(payload, dict) or payload.get("format") != INPUT_FILE_FORMAT:
            raise ValueError("该文件不是本程序保存的 SPD 输入文件。")
        self._apply_input_state(payload, source_path=path)

    def _build_input_state(self) -> dict[str, object]:
        return {
            "format": INPUT_FILE_FORMAT,
            "version": INPUT_FILE_VERSION,
            "settings": {
                "ref_kind": self._string_var_value("ref_kind_var", DEFAULT_REFERENCE_KIND),
                "start_nm": self._string_var_value("start_var", "300"),
                "end_nm": self._string_var_value("end_var", "1200"),
                "step_nm": self._string_var_value("step_var", "1"),
                "manual_current_text": self._string_var_value("isc_var", ""),
                "default_temperature_c": self._string_var_value("temperature_var", f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}"),
                "temperature_coefficient_percent_per_c": self._string_var_value("temp_coeff_var", f"{DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C:g}"),
                "plot_mode": self._string_var_value("plot_mode_var", "参考 / 测试"),
            },
            "reference": {
                "path": self._path_to_text(self.reference_path),
                "display_name": self._string_var_value("ref_file_var", ""),
                "sheets": list(self.reference_sheets),
                "data": self._dataframe_to_input_payload(self.reference_df),
            },
            "test_is_reference": self.test_is_reference,
            "pending_test": {
                "editing": bool(self.editing_pending_test_source and self.pending_test_df is not None),
                "path": self._path_to_text(self.pending_test_path),
                "sheets": list(self.pending_test_sheets),
                "sheet_name": self._string_var_value("sheet_var", "") if self.editing_pending_test_source else "",
                "wavelength_column": self._string_var_value("wave_col_var", "") if self.editing_pending_test_source else "",
                "irradiance_column": self._string_var_value("irrad_col_var", "") if self.editing_pending_test_source else "",
                "unit_name": self._string_var_value("unit_var", "W/(m^2 nm)") if self.editing_pending_test_source else "W/(m^2 nm)",
                "name": self._string_var_value("test_name_var", "") if self.editing_pending_test_source else "",
                "data": self._dataframe_to_input_payload(self.pending_test_df),
            },
            "test_spectra": [
                {
                    "label": dataset.label,
                    "name": dataset.name,
                    "path": self._path_to_text(dataset.path),
                    "sheet_name": dataset.sheet_name,
                    "wavelength_column": dataset.wavelength_column,
                    "irradiance_column": dataset.irradiance_column,
                    "unit_name": dataset.unit_name,
                    "data": self._dataframe_to_input_payload(dataset.df),
                }
                for dataset in self.test_spectra
            ],
            "sr": {
                "path": self._path_to_text(self.sr_path),
                "display_name": self._string_var_value("sr_file_var", ""),
                "sheets": list(self.sr_sheets),
                "sheet_name": self._string_var_value("sr_sheet_var", ""),
                "wavelength_column": self._string_var_value("sr_wave_col_var", ""),
                "response_column": self._string_var_value("sr_col_var", ""),
                "data": self._dataframe_to_input_payload(self.sr_df),
            },
            "iv_path": self._path_to_text(self.iv_path),
            "iv_curves": [self._iv_curve_to_input_payload(curve) for curve in self.iv_curves],
            "current_inputs": [
                {
                    "label": current.label,
                    "isc_ma": current.isc_ma,
                    "spectrum_label": current.spectrum_label,
                    "temperature_c": current.temperature_c,
                    "source": current.source,
                    "iv_curve_label": current.iv_curve.label if current.iv_curve is not None else "",
                }
                for current in self.current_inputs
            ],
            "selection": {
                "active_test_label": self.active_test_label,
                "current_group": self._string_var_value("current_group_var", ""),
                "current_spectrum": self._string_var_value("current_spectrum_var", ""),
                "current_temperature": self._string_var_value("current_temp_edit_var", f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}"),
            },
        }

    def _apply_input_state(self, payload: dict[str, object], *, source_path: Path) -> None:
        settings = self._payload_dict(payload.get("settings"))
        reference = self._payload_dict(payload.get("reference"))
        pending = self._payload_dict(payload.get("pending_test"))
        sr_section = self._payload_dict(payload.get("sr"))
        selection = self._payload_dict(payload.get("selection"))

        restored_ref_kind = str(settings.get("ref_kind") or DEFAULT_REFERENCE_KIND)
        self.ref_kind_var.set(restored_ref_kind)
        self.start_var.set(str(settings.get("start_nm") or "300"))
        self.end_var.set(str(settings.get("end_nm") or "1200"))
        self.step_var.set(str(settings.get("step_nm") or "1"))
        self.isc_var.set(str(settings.get("manual_current_text") or ""))
        self.temperature_var.set(str(settings.get("default_temperature_c") or f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}"))
        self.temp_coeff_var.set(str(settings.get("temperature_coefficient_percent_per_c") or f"{DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C:g}"))

        self.reference_path = self._path_from_text(reference.get("path"))
        self.reference_sheets = self._string_list(reference.get("sheets"))
        self.reference_df = self._dataframe_from_input_payload(reference.get("data"))
        if self.reference_df is not None:
            self.reference_df.columns = normalized_columns(self.reference_df)
        self.ref_file_var.set(
            self._restored_display_name(reference, self.reference_path, "输入文件内参考光谱", "未载入")
        )
        self._refresh_reference_choices(preferred=restored_ref_kind)

        self.test_is_reference = bool(payload.get("test_is_reference"))
        self.pending_test_path = self._path_from_text(pending.get("path"))
        self.pending_test_sheets = self._string_list(pending.get("sheets"))
        self.pending_test_df = self._dataframe_from_input_payload(pending.get("data"))
        if self.pending_test_df is not None:
            self.pending_test_df.columns = normalized_columns(self.pending_test_df)
        self.editing_pending_test_source = bool(pending.get("editing")) and self.pending_test_df is not None

        self.test_spectra = []
        for item in self._payload_list(payload.get("test_spectra")):
            dataset_payload = self._payload_dict(item)
            df = self._dataframe_from_input_payload(dataset_payload.get("data"))
            if df is None:
                continue
            df.columns = normalized_columns(df)
            label = str(dataset_payload.get("label") or self._next_spectrum_label())
            self.test_spectra.append(
                TestSpectrumDataset(
                    label=label,
                    name=str(dataset_payload.get("name") or label),
                    path=self._path_from_text(dataset_payload.get("path")),
                    sheet_name=str(dataset_payload.get("sheet_name") or ""),
                    df=df,
                    wavelength_column=str(dataset_payload.get("wavelength_column") or ""),
                    irradiance_column=str(dataset_payload.get("irradiance_column") or ""),
                    unit_name=str(dataset_payload.get("unit_name") or "W/(m^2 nm)"),
                )
            )
        self.visible_spectrum_labels = {dataset.label for dataset in self.test_spectra}
        self.highlighted_spectrum_labels.clear()

        self.sr_path = self._path_from_text(sr_section.get("path"))
        self.sr_sheets = self._string_list(sr_section.get("sheets"))
        self.sr_df = self._dataframe_from_input_payload(sr_section.get("data"))
        if self.sr_df is not None:
            self.sr_df.columns = normalized_columns(self.sr_df)
        self._restore_sr_controls(sr_section)

        self.iv_path = self._path_from_text(payload.get("iv_path"))
        self.iv_curves = []
        for item in self._payload_list(payload.get("iv_curves")):
            curve = self._iv_curve_from_input_payload(self._payload_dict(item))
            if curve is not None:
                self.iv_curves.append(curve)
        curves_by_label = {curve.label: curve for curve in self.iv_curves if curve.label}
        self.current_inputs = []
        for item in self._payload_list(payload.get("current_inputs")):
            current_payload = self._payload_dict(item)
            label = str(current_payload.get("label") or self._next_current_label())
            iv_curve_label = str(current_payload.get("iv_curve_label") or "")
            self.current_inputs.append(
                CurrentInput(
                    label=label,
                    isc_ma=self._float_value(current_payload.get("isc_ma"), 0.0),
                    spectrum_label=str(current_payload.get("spectrum_label") or ""),
                    temperature_c=self._float_value(current_payload.get("temperature_c"), DEFAULT_REFERENCE_TEMPERATURE_C),
                    source=str(current_payload.get("source") or "输入文件"),
                    iv_curve=curves_by_label.get(iv_curve_label),
                )
            )
        self.last_iv_curve = self.iv_curves[-1] if self.iv_curves else None
        self.current_input_source_by_label = {}

        active_label = str(selection.get("active_test_label") or "")
        if active_label not in {dataset.label for dataset in self.test_spectra}:
            active_label = self.test_spectra[0].label if self.test_spectra else ""
        self.active_test_label = active_label
        self._refresh_test_spectrum_controls(selected_label=active_label, sync_fields=not self.editing_pending_test_source)
        if self.editing_pending_test_source:
            self._restore_pending_test_controls(pending)
        self._refresh_current_spectrum_choices()
        self._refresh_current_inputs_view(select_label=str(selection.get("current_group") or ""))
        if not self.current_inputs:
            self.current_spectrum_var.set(str(selection.get("current_spectrum") or ""))
            self.current_temp_edit_var.set(str(selection.get("current_temperature") or f"{DEFAULT_REFERENCE_TEMPERATURE_C:g}"))
        self.iv_file_var.set(self._restored_iv_status_text())
        self._refresh_plot_modes(select_mode=str(settings.get("plot_mode") or ""))

        self._clear_result_display()
        self._clear_correction()
        status = f"已读取输入文件：{source_path.name}"
        try:
            if self.reference_df is not None and self.test_spectra:
                results, result = self._calculate_spectrum_results()
                self.last_results_by_label = results
                self.last_result = result
                self._render_result(result)
                self._refresh_correction_if_ready(show_errors=False)
                status += "，并已自动重新计算"
        except Exception as exc:
            status += f"，但自动计算失败：{exc}"
        self.status_var.set(status + "。")

    def _restore_pending_test_controls(self, pending: dict[str, object]) -> None:
        if self.pending_test_df is None:
            return
        self.test_path = self.pending_test_path
        self.test_df = self.pending_test_df
        sheet_name = str(pending.get("sheet_name") or (self.pending_test_sheets[0] if self.pending_test_sheets else ""))
        self.sheet_combo_values(self.pending_test_sheets, sheet_name if self.pending_test_sheets else "")
        columns = normalized_columns(self.pending_test_df)
        for combo in (self.wave_combo, self.irrad_combo):
            combo.configure(values=columns)
        self.wave_col_var.set(str(pending.get("wavelength_column") or guess_wavelength_column(columns) or (columns[0] if columns else "")))
        self.irrad_col_var.set(
            str(
                pending.get("irradiance_column")
                or guess_irradiance_column(columns, self.ref_kind_var.get())
                or (columns[1] if len(columns) > 1 else self.wave_col_var.get())
            )
        )
        self.unit_var.set(str(pending.get("unit_name") or "W/(m^2 nm)"))
        self.test_name_var.set(str(pending.get("name") or ""))
        name = self.pending_test_path.name if self.pending_test_path is not None else "输入文件内测试光谱"
        self.test_file_var.set(f"待添加：{name}")
        self._refresh_spectrum_records(selected_label="")

    def _restore_sr_controls(self, sr_section: dict[str, object]) -> None:
        if self.sr_df is None:
            self.sr_file_var.set("未选择 SR 文件")
            self.sr_sheet_combo_values([], "")
            self.sr_wave_col_var.set("")
            self.sr_col_var.set("")
            for combo in (self.sr_wave_combo, self.sr_col_combo):
                combo.configure(values=[])
            return
        self.sr_file_var.set(self._restored_display_name(sr_section, self.sr_path, "输入文件内 SR", "未选择 SR 文件"))
        sheet_name = str(sr_section.get("sheet_name") or (self.sr_sheets[0] if self.sr_sheets else ""))
        self.sr_sheet_combo_values(self.sr_sheets, sheet_name if self.sr_sheets else "")
        columns = normalized_columns(self.sr_df)
        for combo in (self.sr_wave_combo, self.sr_col_combo):
            combo.configure(values=columns)
        self.sr_wave_col_var.set(str(sr_section.get("wavelength_column") or guess_wavelength_column(columns) or ""))
        self.sr_col_var.set(str(sr_section.get("response_column") or guess_response_column(columns) or ""))

    def _clear_result_display(self) -> None:
        self.last_result = None
        self.last_results_by_label = {}
        if hasattr(self, "metric_vars"):
            for variable in self.metric_vars.values():
                variable.set("-")
        for tree_name in ("band_tree", "data_tree"):
            tree = getattr(self, tree_name, None)
            if tree is not None:
                tree.delete(*tree.get_children())
        if hasattr(self, "canvas"):
            self.draw_plot()

    def _dataframe_to_input_payload(self, df: pd.DataFrame | None) -> dict[str, object] | None:
        if df is None:
            return None
        return {
            "columns": [str(column) for column in df.columns],
            "rows": [
                [self._json_safe_value(value) for value in row]
                for row in df.itertuples(index=False, name=None)
            ],
        }

    def _dataframe_from_input_payload(self, payload: object) -> pd.DataFrame | None:
        data = self._payload_dict(payload)
        if not data:
            return None
        columns = [str(column) for column in self._payload_list(data.get("columns"))]
        rows = self._payload_list(data.get("rows"))
        return pd.DataFrame(rows, columns=columns if columns else None)

    def _iv_curve_to_input_payload(self, curve: IVCurve) -> dict[str, object]:
        return {
            "label": curve.label,
            "source_name": curve.source_name,
            "sheet_name": curve.sheet_name,
            "voltage_column": curve.voltage_column,
            "current_column": curve.current_column,
            "header_row": curve.header_row,
            "first_data_row": curve.first_data_row,
            "voltage_v": [self._json_safe_value(value) for value in curve.voltage_v],
            "current_ma": [self._json_safe_value(value) for value in curve.current_ma],
            "isc_ma": curve.isc_ma,
        }

    def _iv_curve_from_input_payload(self, payload: dict[str, object]) -> IVCurve | None:
        if not payload:
            return None
        return IVCurve(
            source_name=str(payload.get("source_name") or ""),
            sheet_name=str(payload.get("sheet_name") or ""),
            voltage_column=str(payload.get("voltage_column") or ""),
            current_column=str(payload.get("current_column") or ""),
            header_row=self._optional_int(payload.get("header_row")),
            first_data_row=self._optional_int(payload.get("first_data_row")),
            voltage_v=self._float_array(payload.get("voltage_v")),
            current_ma=self._float_array(payload.get("current_ma")),
            isc_ma=self._float_value(payload.get("isc_ma"), 0.0),
            label=str(payload.get("label") or ""),
        )

    def _json_safe_value(self, value: object) -> object:
        if value is None:
            return None
        if isinstance(value, np.generic):
            value = value.item()
        if isinstance(value, pd.Timestamp):
            return None if pd.isna(value) else value.isoformat()
        if isinstance(value, float):
            return value if math.isfinite(value) else None
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(value, (str, int, bool)):
            return value
        if isinstance(value, float):
            return value
        return str(value)

    def _string_var_value(self, attr_name: str, default: str = "") -> str:
        variable = getattr(self, attr_name, None)
        if variable is None:
            return default
        try:
            return str(variable.get())
        except Exception:
            return default

    def _path_to_text(self, path: Path | None) -> str:
        return str(path) if path is not None else ""

    def _path_from_text(self, value: object) -> Path | None:
        text = str(value or "").strip()
        return Path(text) if text else None

    def _payload_dict(self, value: object) -> dict[str, object]:
        return value if isinstance(value, dict) else {}

    def _payload_list(self, value: object) -> list[object]:
        return value if isinstance(value, list) else []

    def _string_list(self, value: object) -> list[str]:
        return [str(item) for item in self._payload_list(value)]

    def _float_value(self, value: object, default: float) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def _optional_int(self, value: object) -> int | None:
        try:
            if value is None:
                return None
            return int(value)
        except (TypeError, ValueError):
            return None

    def _float_array(self, value: object) -> np.ndarray:
        values = []
        for item in self._payload_list(value):
            try:
                values.append(float(item))
            except (TypeError, ValueError):
                values.append(np.nan)
        return np.asarray(values, dtype=float)

    def _restored_display_name(
        self,
        payload: dict[str, object],
        path: Path | None,
        fallback_with_data: str,
        fallback_empty: str,
    ) -> str:
        display_name = str(payload.get("display_name") or "").strip()
        if display_name:
            return display_name
        if path is not None:
            return path.name
        return fallback_with_data if payload.get("data") else fallback_empty

    def _restored_iv_status_text(self) -> str:
        if self.current_inputs:
            preview = "；".join(
                f"{item.label}={fmt_number(item.isc_ma, 6)} mA"
                for item in self.current_inputs[:4]
            )
            suffix = "..." if len(self.current_inputs) > 4 else ""
            return f"输入文件已恢复 {len(self.current_inputs)} 组电流；{preview}{suffix}"
        if self.iv_curves:
            return f"输入文件已恢复 {len(self.iv_curves)} 组 IV 曲线"
        return "未读取 IV 表格"

    def _load_iv_path(self, path: Path) -> None:
        self._load_iv_paths([path])

    def _load_iv_paths(self, paths: list[Path]) -> list[str]:
        curves: list[IVCurve] = []
        failures: list[str] = []
        for path in paths:
            try:
                curves.extend(read_iv_curves(path))
            except Exception as exc:
                failures.append(f"{path.name}: {exc}")
        if not curves:
            detail = "\n".join(failures[:8])
            suffix = "\n..." if len(failures) > 8 else ""
            raise ValueError(f"未读取到可用的 IV 曲线。\n{detail}{suffix}" if detail else "未读取到可用的 IV 曲线。")

        added_inputs = []
        for curve in curves:
            label = self._next_current_label()
            curve.label = label
            self.iv_curves.append(curve)
            current = CurrentInput(
                label=label,
                isc_ma=curve.isc_ma,
                spectrum_label=self._default_current_spectrum_label(),
                temperature_c=self._default_current_temperature(),
                source=f"{curve.source_name} / {curve.current_column}",
                iv_curve=curve,
            )
            self.current_inputs.append(current)
            added_inputs.append(current)
        self.iv_path = paths[0]
        self.last_iv_curve = curves[-1]
        self.isc_var.set(fmt_number(curves[-1].isc_ma, 8))

        preview = "，".join(f"{item.label}={fmt_number(item.isc_ma, 6)} mA" for item in added_inputs[:6])
        suffix = "..." if len(added_inputs) > 6 else ""
        warning = f"；{len(failures)} 个文件未读取" if failures else ""
        self.iv_file_var.set(f"本次添加 {len(added_inputs)} 组；{preview}{suffix}{warning}")
        self._refresh_current_inputs_view(select_label=added_inputs[-1].label if added_inputs else None)
        self._refresh_plot_modes(select_mode="IV 曲线")
        self.draw_plot()
        self._refresh_correction_if_ready(show_errors=False)
        self.status_var.set(f"已添加 {len(added_inputs)} 组 IV，电流已换算为 mA。{warning}")
        return failures

    def add_manual_current(self) -> None:
        try:
            values = parse_current_values_ma(self.isc_var.get())
            if not values:
                raise ValueError("请输入电流值，单位 mA。")
            added = []
            for value in values:
                current = CurrentInput(
                    label=self._next_current_label(),
                    isc_ma=abs(float(value)),
                    spectrum_label=self._default_current_spectrum_label(),
                    temperature_c=self._default_current_temperature(),
                    source="手动输入",
                )
                self.current_inputs.append(current)
                added.append(current)
            self._refresh_current_inputs_view(select_label=added[-1].label if added else None)
            self._refresh_correction_if_ready(show_errors=False)
            self.status_var.set(f"已添加 {len(added)} 组手动电流。")
        except Exception as exc:
            messagebox.showerror("添加电流失败", str(exc))

    def _next_current_label(self) -> str:
        return f"I{len(self.current_inputs) + 1}"

    def _default_current_spectrum_label(self) -> str:
        if self.current_inputs and self.current_inputs[-1].spectrum_label:
            return self.current_inputs[-1].spectrum_label
        active = self._active_test_spectrum()
        if active is not None:
            return active.label
        return self.test_spectra[0].label if self.test_spectra else ""

    def _default_current_temperature(self) -> float:
        if self.current_inputs:
            return self.current_inputs[-1].temperature_c
        return float(self.temperature_var.get() or DEFAULT_REFERENCE_TEMPERATURE_C)

    def _current_input_by_label(self, label: str) -> CurrentInput | None:
        for current in self.current_inputs:
            if current.label == label:
                return current
        return None

    def _spectrum_name_for_label(self, label: str) -> str:
        dataset = self._test_spectrum_by_label(label)
        return dataset.name if dataset is not None else ""

    def _spectrum_display_for_label(self, label: str) -> str:
        dataset = self._test_spectrum_by_label(label)
        return self._test_spectrum_display(dataset) if dataset is not None else ""

    def _refresh_current_spectrum_choices(self) -> None:
        if not hasattr(self, "current_spectrum_combo"):
            return
        values = [self._test_spectrum_display(dataset) for dataset in self.test_spectra]
        self.current_spectrum_combo.configure(values=values)
        if not values:
            self.current_spectrum_var.set("")
            return
        if not self.current_inputs:
            active = self._active_test_spectrum()
            self.current_spectrum_var.set(self._test_spectrum_display(active) if active is not None else values[-1])
            return
        labels = {dataset.label for dataset in self.test_spectra}
        current_label = self._label_from_display(self.current_spectrum_var.get())
        if current_label not in labels:
            active = self._active_test_spectrum()
            self.current_spectrum_var.set(self._test_spectrum_display(active) if active is not None else values[-1])
        if self.test_spectra:
            default_label = self.test_spectra[-1].label
            for current in self.current_inputs:
                if not current.spectrum_label:
                    current.spectrum_label = default_label

    def _refresh_current_group_choices(self, select_label: str | None = None) -> None:
        if not hasattr(self, "current_group_combo"):
            return
        labels = [current.label for current in self.current_inputs]
        self.current_group_combo.configure(values=labels)
        if select_label in labels:
            self.current_group_var.set(select_label)
        elif self.current_group_var.get() not in labels:
            self.current_group_var.set(labels[-1] if labels else "")
        self._select_current_input()

    def _refresh_current_inputs_view(
        self,
        evaluations: list[CurrentCorrectionEvaluation] | None = None,
        select_label: str | None = None,
    ) -> None:
        if not hasattr(self, "cv_tree"):
            return
        evaluation_by_label = {evaluation.label: evaluation for evaluation in evaluations or []}
        self.cv_tree.delete(*self.cv_tree.get_children())
        for row_index, current in enumerate(self.current_inputs):
            evaluation = evaluation_by_label.get(current.label)
            if evaluation is None:
                values = (
                    current.label,
                    fmt_number(current.isc_ma, 6),
                    self._spectrum_name_for_label(current.spectrum_label),
                    fmt_number(current.temperature_c, 5),
                    "-",
                    "-",
                    "待计算",
                    current.source,
                    "-",
                    "-",
                    "-",
                )
                tags = ("even" if row_index % 2 else "odd",)
            else:
                status = "合格" if evaluation.valid else "剔除：" + "；".join(dict.fromkeys(evaluation.reasons))
                values = (
                    evaluation.label,
                    fmt_number(evaluation.correction.measured_isc, 6),
                    evaluation.spectrum_name,
                    fmt_number(evaluation.correction.test_temperature_c, 5),
                    fmt_number(evaluation.correction.corrected_cv, 6),
                    f"{evaluation.deviation_percent:.2f}",
                    status,
                    evaluation.source,
                    fmt_number(evaluation.correction.mg, 6),
                    fmt_number(evaluation.correction.mt, 6),
                    fmt_number(evaluation.correction.mmf, 6),
                )
                tags = ("even" if row_index % 2 else "odd", "valid" if evaluation.valid else "invalid")
            self.cv_tree.insert("", "end", iid=current.label, values=values, tags=tags)
        self._refresh_current_group_choices(select_label=select_label)
        if select_label and select_label in self.cv_tree.get_children():
            self.cv_tree.selection_set(select_label)

    def _select_current_input_from_tree(self) -> None:
        selected = self.cv_tree.selection()
        if selected:
            self.current_group_var.set(str(selected[0]))
            self._select_current_input()

    def _select_current_input(self) -> None:
        current = self._current_input_by_label(self.current_group_var.get())
        if current is None:
            self._refresh_selected_correction_display()
            if getattr(self, "plot_mode_var", None) is not None and self.plot_mode_var.get() == "IV 曲线":
                self.draw_plot()
            return
        spectrum_display = self._spectrum_display_for_label(current.spectrum_label)
        if spectrum_display:
            self.current_spectrum_var.set(spectrum_display)
        self.current_temp_edit_var.set(fmt_number(current.temperature_c, 5))
        if hasattr(self, "cv_tree") and current.label in self.cv_tree.get_children():
            if tuple(self.cv_tree.selection()) != (current.label,):
                self.cv_tree.selection_set(current.label)
            self.cv_tree.see(current.label)
        self._refresh_selected_correction_display()
        if getattr(self, "plot_mode_var", None) is not None and self.plot_mode_var.get() == "IV 曲线":
            self.draw_plot()

    def apply_current_settings(self) -> None:
        try:
            current = self._current_input_by_label(self.current_group_var.get())
            if current is None:
                raise ValueError("请先选择电流组。")
            spectrum_label = self._label_from_display(self.current_spectrum_var.get())
            if not self._test_spectrum_by_label(spectrum_label):
                raise ValueError("请先为该电流组选择测试光谱。")
            current.spectrum_label = spectrum_label
            current.temperature_c = float(self.current_temp_edit_var.get())
            self._refresh_current_inputs_view(select_label=current.label)
            self._refresh_correction_if_ready(show_errors=False)
            self.status_var.set(f"已更新 {current.label} 的光谱和温度。")
        except Exception as exc:
            messagebox.showerror("应用设置失败", str(exc))

    def delete_current_input(self) -> None:
        label = self.current_group_var.get()
        if not label:
            return
        self.current_inputs = [current for current in self.current_inputs if current.label != label]
        self.iv_curves = [curve for curve in self.iv_curves if curve.label != label]
        for index, current in enumerate(self.current_inputs, start=1):
            old_label = current.label
            current.label = f"I{index}"
            if current.iv_curve is not None:
                current.iv_curve.label = current.label
            for curve in self.iv_curves:
                if curve.label == old_label:
                    curve.label = current.label
        self._refresh_current_inputs_view()
        self._refresh_plot_modes()
        if self.plot_mode_var.get() == "IV 曲线":
            self.draw_plot()
        self._clear_correction()
        self.status_var.set(f"已删除 {label}。")

    def _load_sr_path(self, path: Path) -> None:
        self.sr_path = path
        suffix = path.suffix.lower()
        if suffix in EXCEL_TABLE_SUFFIXES:
            self.sr_sheets = list_excel_sheets(path)
            sheet = self.sr_sheets[0]
            self.sr_sheet_combo_values(self.sr_sheets, sheet)
            self.sr_df = read_table(path, sheet)
        else:
            self.sr_sheets = []
            self.sr_sheet_combo_values([], "")
            self.sr_df = read_table(path)
        self.sr_df.columns = normalized_columns(self.sr_df)
        self.sr_file_var.set(path.name)
        self._populate_sr_columns()
        self._clear_correction()
        self.draw_plot()
        self.status_var.set("已载入光谱响应 SR。")

    def _read_test_source_path(self, path: Path) -> tuple[pd.DataFrame, list[str], str]:
        suffix = path.suffix.lower()
        sheets: list[str] = []
        if suffix in EXCEL_TABLE_SUFFIXES:
            sheets = list_excel_sheets(path)
            sheet = sheets[0]
            df = read_table(path, sheet)
        else:
            sheet = ""
            df = read_table(path)
        df.columns = normalized_columns(df)
        return df, sheets, str(sheet) if sheets else ""

    def _load_test_source_path(self, path: Path) -> None:
        self.test_is_reference = False
        df, sheets, sheet = self._read_test_source_path(path)
        self.pending_test_path = path
        self.pending_test_df = df
        self.pending_test_sheets = sheets
        self.editing_pending_test_source = True
        self.test_path = path
        self.test_df = df
        self.sheet_combo_values(self.pending_test_sheets, sheet)
        self._populate_pending_test_columns()
        self.test_name_var.set("")
        self.test_file_var.set(f"待添加：{path.name}")
        self.status_var.set("已选择测试表格；选择数据列后点击“添加当前列”。")

    def _load_test_path(self, path: Path) -> None:
        self._load_test_source_path(path)
        self.add_selected_test_spectrum()

    def _load_test_batch_paths(self, paths: list[Path]) -> None:
        added: list[TestSpectrumDataset] = []
        failures: list[tuple[Path, str]] = []
        self.test_is_reference = False
        for path in paths:
            try:
                df, _sheets, sheet = self._read_test_source_path(path)
                wave_col, irrad_col = self._guess_test_columns(df)
                if not wave_col or not irrad_col:
                    raise ValueError("未识别到可用的波长列和辐照度列。")
                dataset = self._create_test_spectrum_dataset(
                    path=path,
                    sheet_name=sheet,
                    df=df,
                    wavelength_column=wave_col,
                    irradiance_column=irrad_col,
                    unit_name=self.unit_var.get() or "W/(m^2 nm)",
                )
                self.test_spectra.append(dataset)
                self._mark_spectrum_visible(dataset.label)
                added.append(dataset)
            except Exception as exc:
                failures.append((path, str(exc)))

        if added:
            last_label = added[-1].label
            self.editing_pending_test_source = False
            self._refresh_test_spectrum_controls(selected_label=last_label)
            self._refresh_current_spectrum_choices()
            self._refresh_spectrum_records(selected_label=last_label)
            self._clear_correction()
            self.test_name_var.set("")
            self.status_var.set(f"批量导入完成：新增 {len(added)} 组测试光谱。")

        if failures:
            summary = "\n".join(f"{path.name}：{reason}" for path, reason in failures[:8])
            extra = "" if len(failures) <= 8 else f"\n另有 {len(failures) - 8} 个文件未列出。"
            messagebox.showwarning(
                "批量导入部分失败",
                f"已导入 {len(added)} 个文件，{len(failures)} 个文件未导入：\n{summary}{extra}",
            )
            if not added:
                self.status_var.set("批量导入失败：未新增测试光谱。")
        elif added:
            messagebox.showinfo("批量导入完成", f"已新增 {len(added)} 组测试光谱。")

    def _populate_pending_test_columns(self) -> None:
        if self.pending_test_df is None:
            return
        columns = normalized_columns(self.pending_test_df)
        for combo in (self.wave_combo, self.irrad_combo):
            combo.configure(values=columns)
        wave_col, irrad_col = self._guess_test_columns(self.pending_test_df)
        self.wave_col_var.set(wave_col)
        self.irrad_col_var.set(irrad_col)
        self.test_df = self.pending_test_df

    def _guess_test_columns(self, df: pd.DataFrame) -> tuple[str, str]:
        columns = normalized_columns(df)
        wave_col = guess_wavelength_column(columns) or (columns[0] if columns else "")
        irrad_col = guess_irradiance_column(columns, self.ref_kind_var.get()) or (
            columns[1] if len(columns) > 1 else wave_col
        )
        return wave_col, irrad_col

    def _next_spectrum_label(self) -> str:
        existing_numbers = []
        for dataset in self.test_spectra:
            match = re.fullmatch(r"S(\d+)", dataset.label)
            if match:
                existing_numbers.append(int(match.group(1)))
        return f"S{max(existing_numbers, default=0) + 1}"

    def _default_spectrum_name(self, path: Path | None, irradiance_col: str) -> str:
        file_name = path.name if path is not None else "测试光谱"
        return f"{file_name}+{irradiance_col}" if irradiance_col else file_name

    def _create_test_spectrum_dataset(
        self,
        *,
        path: Path | None,
        sheet_name: str,
        df: pd.DataFrame,
        wavelength_column: str,
        irradiance_column: str,
        unit_name: str,
        name: str | None = None,
    ) -> TestSpectrumDataset:
        return TestSpectrumDataset(
            label=self._next_spectrum_label(),
            name=name or self._default_spectrum_name(path, irradiance_column),
            path=path,
            sheet_name=sheet_name,
            df=df.copy(),
            wavelength_column=wavelength_column,
            irradiance_column=irradiance_column,
            unit_name=unit_name or "W/(m^2 nm)",
        )

    def add_selected_test_spectrum(self) -> None:
        if self.pending_test_df is None:
            messagebox.showwarning("未选择表格", "请先点击“选择测试表格”。")
            return
        if not self.wave_col_var.get() or not self.irrad_col_var.get():
            messagebox.showwarning("未选择数据列", "请选择波长列和数据列。")
            return
        name = self.test_name_var.get().strip() or self._default_spectrum_name(self.pending_test_path, self.irrad_col_var.get())
        dataset = self._create_test_spectrum_dataset(
            path=self.pending_test_path,
            sheet_name=self.sheet_var.get(),
            df=self.pending_test_df.copy(),
            wavelength_column=self.wave_col_var.get(),
            irradiance_column=self.irrad_col_var.get(),
            unit_name=self.unit_var.get() or "W/(m^2 nm)",
            name=name,
        )
        self.test_spectra.append(dataset)
        self._mark_spectrum_visible(dataset.label)
        self._refresh_test_spectrum_controls(selected_label=dataset.label, sync_fields=False)
        self._refresh_current_spectrum_choices()
        self._refresh_spectrum_records(selected_label="")
        self._clear_correction()
        self.test_name_var.set("")
        self.status_var.set(f"已添加测试光谱：{dataset.name}")

    def sheet_combo_values(self, values: list[str], selected: str) -> None:
        self.sheet_var.set(selected)
        self.sheet_combo.configure(values=values)

    def sr_sheet_combo_values(self, values: list[str], selected: str) -> None:
        self.sr_sheet_var.set(selected)
        self.sr_sheet_combo.configure(values=values)

    def _change_test_sheet(self) -> None:
        if self.editing_pending_test_source:
            if self.pending_test_path is None or not self.sheet_var.get():
                return
            try:
                df = read_table(self.pending_test_path, self.sheet_var.get())
                df.columns = normalized_columns(df)
                self.pending_test_df = df
                self.test_df = df
                self._populate_pending_test_columns()
                self.test_file_var.set(f"待添加：{self.pending_test_path.name}")
            except Exception as exc:
                messagebox.showerror("读取工作表失败", str(exc))
            return

        dataset = self._active_test_spectrum()
        if dataset is None or dataset.path is None or not self.sheet_var.get():
            return
        try:
            df = read_table(dataset.path, self.sheet_var.get())
            df.columns = normalized_columns(df)
            dataset.df = df
            dataset.sheet_name = self.sheet_var.get()
            self._populate_test_columns()
            self._clear_correction()
        except Exception as exc:
            messagebox.showerror("读取工作表失败", str(exc))

    def _change_sr_sheet(self) -> None:
        if self.sr_path is None or not self.sr_sheet_var.get():
            return
        try:
            self.sr_df = read_table(self.sr_path, self.sr_sheet_var.get())
            self.sr_df.columns = normalized_columns(self.sr_df)
            self._populate_sr_columns()
            self._clear_correction()
            self.draw_plot()
        except Exception as exc:
            messagebox.showerror("读取 SR 工作表失败", str(exc))

    def _populate_test_columns(self) -> None:
        dataset = self._active_test_spectrum()
        if dataset is None:
            return
        self.test_df = dataset.df
        columns = normalized_columns(dataset.df)
        for combo in (self.wave_combo, self.irrad_combo):
            combo.configure(values=columns)
        wave_col = guess_wavelength_column(columns)
        irrad_col = guess_irradiance_column(columns, self.ref_kind_var.get())
        if wave_col:
            dataset.wavelength_column = wave_col
            self.wave_col_var.set(wave_col)
        if irrad_col:
            dataset.irradiance_column = irrad_col
            self.irrad_col_var.set(irrad_col)
        dataset.unit_name = self.unit_var.get() or dataset.unit_name

    def _populate_sr_columns(self) -> None:
        if self.sr_df is None:
            return
        columns = normalized_columns(self.sr_df)
        for combo in (self.sr_wave_combo, self.sr_col_combo):
            combo.configure(values=columns)
        wave_col = guess_wavelength_column(columns)
        sr_col = guess_response_column(columns)
        if wave_col:
            self.sr_wave_col_var.set(wave_col)
        if sr_col:
            self.sr_col_var.set(sr_col)

    def use_reference_as_test(self) -> None:
        if self.reference_df is None or self.reference_path is None:
            messagebox.showwarning("暂无参考光谱", "请先载入参考光谱。")
            return
        self.test_is_reference = True
        self.editing_pending_test_source = False
        df = self.reference_df.copy()
        columns = normalized_columns(df)
        wave_col = guess_wavelength_column(columns) or (columns[0] if columns else "")
        irrad_col = guess_irradiance_column(columns, self.ref_kind_var.get()) or (
            columns[1] if len(columns) > 1 else wave_col
        )
        dataset = TestSpectrumDataset(
            label="S1",
            name=f"{self.reference_path.name}（自测）",
            path=self.reference_path,
            sheet_name=self.reference_sheets[0] if self.reference_sheets else "",
            df=df,
            wavelength_column=wave_col,
            irradiance_column=irrad_col,
            unit_name="W/(m^2 nm)",
        )
        self.test_spectra = [dataset]
        self.visible_spectrum_labels = {dataset.label}
        self.highlighted_spectrum_labels.clear()
        self._refresh_test_spectrum_controls(selected_label=dataset.label)
        self._refresh_current_spectrum_choices()
        self._clear_correction()
        self.status_var.set("已使用参考光谱作为测试数据。")

    def _reference_result_name(self, reference_column: str) -> str:
        selected = self.ref_kind_var.get().strip()
        if not selected or selected == reference_column:
            return reference_column
        return f"{selected} - {reference_column}"

    def _calculate_spectrum_results(self) -> tuple[dict[str, CalculationResult], CalculationResult]:
        if self.reference_df is None:
            raise ValueError("请先载入参考光谱。")
        if not self.test_spectra:
            raise ValueError("请先添加测试光谱。")

        start_nm = float(self.start_var.get())
        end_nm = float(self.end_var.get())
        step_nm = float(self.step_var.get())
        ref_wave_col = guess_wavelength_column(normalized_columns(self.reference_df))
        ref_irrad_col = guess_irradiance_column(normalized_columns(self.reference_df), self.ref_kind_var.get())
        if ref_wave_col is None or ref_irrad_col is None:
            raise ValueError("参考表格中无法识别波长列或参考辐照度列。")

        reference = spectrum_from_dataframe(self.reference_df, ref_wave_col, ref_irrad_col, 1.0)
        reference_name = self._reference_result_name(ref_irrad_col)
        results: dict[str, CalculationResult] = {}
        for dataset in self.test_spectra:
            if not dataset.wavelength_column or not dataset.irradiance_column:
                raise ValueError(f"{dataset.name} 尚未选择波长列或辐照度列。")
            unit_factor = UNIT_FACTORS.get(dataset.unit_name, 1.0)
            test = spectrum_from_dataframe(dataset.df, dataset.wavelength_column, dataset.irradiance_column, unit_factor)
            results[dataset.label] = calculate_spd(
                reference,
                test,
                reference_name=reference_name,
                test_name=dataset.name,
                start_nm=start_nm,
                end_nm=end_nm,
                step_nm=step_nm,
                normalize=False,
            )

        active = self._active_test_spectrum() or self.test_spectra[0]
        return results, results[active.label]

    def calculate(self, show_errors: bool = True) -> None:
        try:
            results, result = self._calculate_spectrum_results()
            self.last_results_by_label = results
            self.last_result = result
            self._render_result(result)
            self._refresh_correction_if_ready(show_errors=False)
            self.status_var.set(f"计算完成：{len(results)} 组测试光谱。")
        except Exception as exc:
            if show_errors:
                messagebox.showerror("计算失败", str(exc))
            self.status_var.set(f"计算失败：{exc}")

    def calculate_correction(self) -> None:
        try:
            evaluations = self._calculate_corrections()
            self.last_correction_evaluations = evaluations
            self.last_correction = evaluations[0].correction if evaluations else None
            self._render_corrections(evaluations)
            self.status_var.set("ISC 修正计算完成。")
        except Exception as exc:
            messagebox.showerror("ISC 修正失败", str(exc))
            self.status_var.set(f"ISC 修正失败：{exc}")

    def _correction_inputs_ready(self) -> bool:
        return bool(
            self.reference_df is not None
            and bool(self.test_spectra)
            and self.sr_df is not None
            and self.sr_wave_col_var.get()
            and self.sr_col_var.get()
            and (self.current_inputs or self.isc_var.get().strip())
        )

    def _refresh_correction_if_ready(self, *, show_errors: bool) -> None:
        if not self._correction_inputs_ready():
            return
        try:
            evaluations = self._calculate_corrections()
            self.last_correction_evaluations = evaluations
            self.last_correction = evaluations[0].correction if evaluations else None
            self._render_corrections(evaluations)
        except Exception as exc:
            if show_errors:
                messagebox.showerror("ISC 修正失败", str(exc))
            else:
                self.status_var.set(f"ISC 修正未更新：{exc}")

    def _calculate_correction(self) -> CurrentCorrectionResult:
        evaluations = self._calculate_corrections()
        if not evaluations:
            raise ValueError("没有可用的 ISC 数据。")
        return evaluations[0].correction

    def _calculate_corrections(self) -> list[CurrentCorrectionEvaluation]:
        if self.reference_df is None:
            raise ValueError("请先载入参考光谱。")
        if not self.test_spectra:
            raise ValueError("请先添加测试光谱。")
        if self.sr_df is None:
            raise ValueError("请先载入光谱响应 SR 文件。")
        if not self.sr_wave_col_var.get() or not self.sr_col_var.get():
            raise ValueError("请选择 SR 表格的波长列和 SR 列。")

        ref_wave_col = guess_wavelength_column(normalized_columns(self.reference_df))
        ref_irrad_col = guess_irradiance_column(normalized_columns(self.reference_df), self.ref_kind_var.get())
        if ref_wave_col is None or ref_irrad_col is None:
            raise ValueError("参考表格中无法识别波长列或参考辐照度列。")

        temperature_coefficient = float(self.temp_coeff_var.get() or DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C)

        reference = spectrum_from_dataframe(self.reference_df, ref_wave_col, ref_irrad_col, 1.0)
        sr = response_from_dataframe(self.sr_df, self.sr_wave_col_var.get(), self.sr_col_var.get())

        if self.current_inputs:
            current_inputs = list(self.current_inputs)
        else:
            manual_values = parse_current_values_ma(self.isc_var.get())
            if not manual_values:
                raise ValueError("请输入或添加至少一组短路电流，单位 mA。")
            active = self._active_test_spectrum() or self.test_spectra[0]
            current_inputs = [
                CurrentInput(
                    label=f"I{index}",
                    isc_ma=abs(float(value)),
                    spectrum_label=active.label,
                    temperature_c=float(self.temperature_var.get() or DEFAULT_REFERENCE_TEMPERATURE_C),
                    source="手动输入",
                )
                for index, value in enumerate(manual_values, start=1)
            ]

        corrections = []
        for current in current_inputs:
            dataset = self._test_spectrum_by_label(current.spectrum_label)
            if dataset is None:
                raise ValueError(f"{current.label} 尚未选择有效测试光谱。")
            if not dataset.wavelength_column or not dataset.irradiance_column:
                raise ValueError(f"{dataset.name} 尚未选择波长列或辐照度列。")
            unit_factor = UNIT_FACTORS.get(dataset.unit_name, 1.0)
            test = spectrum_from_dataframe(dataset.df, dataset.wavelength_column, dataset.irradiance_column, unit_factor)
            corrections.append(
                (
                    current.label,
                    calculate_current_correction(
                        reference,
                        test,
                        sr,
                        sr_name=self.sr_path.name if self.sr_path is not None else "SR",
                        sr_sheet_name=self.sr_sheet_var.get(),
                        sr_wavelength_column=self.sr_wave_col_var.get(),
                        sr_column=self.sr_col_var.get(),
                        measured_isc=current.isc_ma,
                        test_temperature_c=current.temperature_c,
                        temperature_coefficient_percent_per_c=temperature_coefficient,
                        start_nm=CURRENT_CORRECTION_START_NM,
                        end_nm=CURRENT_CORRECTION_END_NM,
                    ),
                    dataset.label,
                    dataset.name,
                    current.source,
                )
            )
        return evaluate_current_corrections(corrections)

    def _clear_correction(self) -> None:
        self.last_correction = None
        self.last_correction_evaluations = []
        if not hasattr(self, "correction_vars"):
            return
        for variable in self.correction_vars.values():
            variable.set("-")
        self.correction_detail_var.set("载入 SR 并读取 IV 或输入 ISC 后计算。")
        if hasattr(self, "cv_tree"):
            self._refresh_current_inputs_view()
        self.correction_tree.delete(*self.correction_tree.get_children())

    def _correction_evaluation_for_display(
        self,
        evaluations: list[CurrentCorrectionEvaluation] | None = None,
    ) -> CurrentCorrectionEvaluation | None:
        correction_evaluations = list(evaluations if evaluations is not None else self.last_correction_evaluations)
        if not correction_evaluations:
            return None

        selected_labels: list[str] = []
        if hasattr(self, "current_group_var"):
            selected_labels.append(self.current_group_var.get())
        if hasattr(self, "cv_tree"):
            selected_labels.extend(str(item) for item in self.cv_tree.selection())

        for label in selected_labels:
            if not label:
                continue
            for evaluation in correction_evaluations:
                if evaluation.label == label:
                    return evaluation
        return correction_evaluations[0]

    def _refresh_selected_correction_display(
        self,
        evaluations: list[CurrentCorrectionEvaluation] | None = None,
    ) -> None:
        if not hasattr(self, "correction_vars"):
            return

        correction_evaluations = list(evaluations if evaluations is not None else self.last_correction_evaluations)
        evaluation = self._correction_evaluation_for_display(correction_evaluations)
        if evaluation is None:
            for key in ("mg", "mt", "mmf"):
                self.correction_vars[key].set("-")
            if hasattr(self, "correction_tree"):
                self.correction_tree.delete(*self.correction_tree.get_children())
            return

        correction = evaluation.correction
        self.last_correction = correction
        self.correction_vars["mg"].set(f"{correction.mg:.5f}")
        self.correction_vars["mt"].set(f"{correction.mt:.5f}")
        self.correction_vars["mmf"].set(f"{correction.mmf:.5f}")

        valid_cvs = [item.correction.corrected_cv for item in correction_evaluations if item.valid]
        valid_mean = float(np.mean(valid_cvs)) if valid_cvs else np.nan
        self.correction_detail_var.set(
            f"当前 {evaluation.label}（{evaluation.spectrum_name}）："
            f"SR 波段 {correction.grid_nm[0]:.0f}-{correction.grid_nm[-1]:.0f} nm；"
            f"标准辐照度积分 {fmt_number(correction.reference_irradiance_integral, 6)} W/m²，"
            f"实测辐照度积分 {fmt_number(correction.test_irradiance_integral, 6)} W/m²；"
            f"A5 有效 {len(valid_cvs)}/{len(correction_evaluations)} 组，平均 CV {fmt_number(valid_mean, 6)} mA。"
        )
        self._render_selected_correction_rows(correction)

    def _render_selected_correction_rows(self, correction: CurrentCorrectionResult) -> None:
        if not hasattr(self, "correction_tree"):
            return
        self.correction_tree.delete(*self.correction_tree.get_children())
        preview_step = max(1, int(len(correction.grid_nm) / 250))
        for row_index, idx in enumerate(range(0, len(correction.grid_nm), preview_step)):
            self.correction_tree.insert(
                "",
                "end",
                values=(
                    fmt_number(correction.grid_nm[idx], 5),
                    fmt_number(correction.sr[idx], 5),
                    fmt_number(correction.reference[idx], 5),
                    fmt_number(correction.test[idx], 5),
                    fmt_number(correction.reference_sr_weighted[idx], 5),
                    fmt_number(correction.test_sr_weighted[idx], 5),
                ),
                tags=("even" if row_index % 2 else "odd",),
            )

    def _render_corrections(self, evaluations: list[CurrentCorrectionEvaluation]) -> None:
        if not evaluations:
            self._clear_correction()
            return
        valid_cvs = [evaluation.correction.corrected_cv for evaluation in evaluations if evaluation.valid]
        valid_mean = float(np.mean(valid_cvs)) if valid_cvs else np.nan
        self.correction_vars["cv"].set(fmt_number(valid_mean, 6))

        if self.current_inputs:
            self._refresh_current_inputs_view(evaluations)
        else:
            self.cv_tree.delete(*self.cv_tree.get_children())
            for row_index, evaluation in enumerate(evaluations):
                status = "合格" if evaluation.valid else "剔除：" + "；".join(dict.fromkeys(evaluation.reasons))
                self.cv_tree.insert(
                    "",
                    "end",
                    iid=evaluation.label,
                    values=(
                        evaluation.label,
                        fmt_number(evaluation.correction.measured_isc, 6),
                        evaluation.spectrum_name,
                        fmt_number(evaluation.correction.test_temperature_c, 5),
                        fmt_number(evaluation.correction.corrected_cv, 6),
                        f"{evaluation.deviation_percent:.2f}",
                        status,
                        evaluation.source,
                        fmt_number(evaluation.correction.mg, 6),
                        fmt_number(evaluation.correction.mt, 6),
                        fmt_number(evaluation.correction.mmf, 6),
                    ),
                    tags=("even" if row_index % 2 else "odd", "valid" if evaluation.valid else "invalid"),
                )

        self._refresh_selected_correction_display(evaluations)

    def _render_correction(self, correction: CurrentCorrectionResult) -> None:
        self._render_corrections(evaluate_current_corrections([("I1", correction)]))

    def _iv_curve_by_label(self, label: str) -> IVCurve | None:
        for curve in self.iv_curves:
            if curve.label == label:
                return curve
        return None

    def _render_result(self, result: CalculationResult) -> None:
        self.last_result = result
        self.metric_vars["spd"].set(f"{result.spd_percent:.2f}%")
        self.metric_vars["spc"].set(f"{result.spc_percent:.2f}%")
        self.metric_vars["class"].set(result.overall_class)
        self.metric_vars["test_total"].set(f"{fmt_number(result.test_raw_total, 5)} W/m²")

        self.band_tree.delete(*self.band_tree.get_children())
        for row_index, row in enumerate(result.band_rows):
            class_text = str(row["class"]).upper()
            if class_text.startswith("A"):
                row_tags = ("good",)
            elif class_text.startswith("B"):
                row_tags = ("warn",)
            elif class_text.startswith("C"):
                row_tags = ("bad",)
            else:
                row_tags = ("even" if row_index % 2 else "odd",)
            self.band_tree.insert(
                "",
                "end",
                values=(
                    row["range"],
                    f"{row['ref_percent']:.2f}",
                    f"{row['test_percent']:.2f}",
                    f"{row['ratio']:.3f}",
                    row["class"],
                ),
                tags=row_tags,
            )

        self.data_tree.delete(*self.data_tree.get_children())
        preview_step = max(1, int(len(result.grid_nm) / 250))
        for row_index, idx in enumerate(range(0, len(result.grid_nm), preview_step)):
            self.data_tree.insert(
                "",
                "end",
                values=(
                    fmt_number(result.grid_nm[idx], 5),
                    fmt_number(result.reference[idx], 5),
                    fmt_number(result.test_scaled[idx], 5),
                    fmt_number(result.absolute_error[idx], 5),
                ),
                tags=("even" if row_index % 2 else "odd",),
            )
        self.draw_plot()

    def _selected_sr_curve(self) -> SpectralResponse | None:
        if self.sr_df is None or not self.sr_wave_col_var.get() or not self.sr_col_var.get():
            return None
        try:
            return response_from_dataframe(self.sr_df, self.sr_wave_col_var.get(), self.sr_col_var.get())
        except Exception:
            return None

    def _show_plot_message(self, canvas: tk.Canvas, width: int, height: int, message: str) -> None:
        if hasattr(self, "plot_status_var"):
            self.plot_status_var.set(message)
        canvas.create_text(width / 2, height / 2, text=message, fill=UI_TEXT_MUTED, font=(UI_FONT_FAMILY, 11))

    def _clear_plot_message(self) -> None:
        if hasattr(self, "plot_status_var"):
            self.plot_status_var.set("")

    def draw_plot(self) -> None:
        if not hasattr(self, "canvas"):
            return
        canvas = self.canvas
        canvas.delete("all")
        self.legend_hitboxes = []
        self._prune_spectrum_display_state()
        width = max(canvas.winfo_width(), 400)
        height = max(canvas.winfo_height(), 240)
        pad_left, pad_right, pad_top, pad_bottom = 72, 22, 28, 48

        canvas.create_rectangle(0, 0, width, height, fill="#ffffff", outline="")
        mode = self.plot_mode_var.get()
        result = self._active_calculation_result()
        title = "光谱曲线"
        x_label = "Wavelength (nm)"
        y_label = "W/(m^2 nm)"
        subtitle = ""
        isc_marker: tuple[float, str] | None = None

        if mode == "IV 曲线":
            curve = self._iv_curve_for_mode(mode)
            title = f"{curve.label} IV 曲线" if curve is not None and curve.label else "IV 曲线"
            if curve is None:
                self.plot_title_var.set(title)
                if self.iv_curves:
                    message = "当前电流组没有 IV 曲线，请选择一个来自 IV 表格的电流组"
                else:
                    message = "点击“读取 IV”选择电压/电流表格"
                self._show_plot_message(canvas, width, height, message)
                return
            x = curve.voltage_v
            series = [plot_series_data(curve.current_ma, "#0f766e", "实测 IV")]
            x_label = "Voltage (V)"
            y_label = "Current (mA)"
            row_text = f"，数据从第 {curve.first_data_row} 行开始" if curve.first_data_row else ""
            label_text = f"{curve.label} | " if curve.label else ""
            subtitle = f"{label_text}{curve.source_name} | ISC {fmt_number(curve.isc_ma, 6)} mA{row_text}"
            if curve.current_ma.size:
                isc_marker = (float(curve.current_ma[0]), f"{fmt_number(curve.isc_ma, 6)} mA")
        elif mode == "SR 曲线":
            title = "SR 曲线"
            sr_curve = self._selected_sr_curve()
            if sr_curve is None:
                self.plot_title_var.set(title)
                self._show_plot_message(canvas, width, height, "载入 SR 文件后显示 SR 曲线")
                return
            x = sr_curve.wavelength_nm
            series = [plot_series_data(sr_curve.response_a_w, "#7c3aed", "SR")]
            y_label = "A/W"
            sr_name = self.sr_path.name if self.sr_path is not None else "SR"
            subtitle = f"{sr_name} | {self.sr_wave_col_var.get()} vs {self.sr_col_var.get()}"
        elif mode == "绝对偏差":
            title = "绝对偏差"
            if result is None:
                self.plot_title_var.set(title)
                self._show_plot_message(canvas, width, height, "选择测试表格后点击左上方“开始计算”")
                return
            x = result.grid_nm
            series = [plot_series_data(result.absolute_error, "#f59e0b", "绝对偏差")]
            subtitle = (
                f"{result.reference_name} vs {result.test_name} | "
                f"SPD {result.spd_percent:.2f}%"
            )
        else:
            if result is None and not self.last_results_by_label:
                self.plot_title_var.set(title)
                self._show_plot_message(canvas, width, height, "添加测试光谱后点击左上方“开始计算”")
                return
            active_result = result or next(iter(self.last_results_by_label.values()))
            x = active_result.grid_nm
            series = [self._plot_reference_series(active_result.reference, reference_legend_label(active_result.reference_name))]
            if self.last_results_by_label:
                total_test_count = 0
                visible_test_count = 0
                for dataset in self.test_spectra:
                    dataset_result = self.last_results_by_label.get(dataset.label)
                    if dataset_result is not None:
                        total_test_count += 1
                        if self._is_spectrum_visible(dataset.label):
                            visible_test_count += 1
                            series.append(self._plot_series_for_dataset(dataset, dataset_result.test_scaled))
                if total_test_count and visible_test_count == 0:
                    subtitle = f"全部 {total_test_count} 组测试光谱已隐藏；参考曲线保持显示。"
                else:
                    subtitle = f"已绘制 {visible_test_count} / {total_test_count} 组测试光谱；图例按光谱名称显示。"
            else:
                active_dataset = self._active_test_spectrum()
                if active_dataset is None or self._is_spectrum_visible(active_dataset.label):
                    if active_dataset is not None:
                        series.append(self._plot_series_for_dataset(active_dataset, active_result.test_scaled))
                    else:
                        series.append(plot_series_data(active_result.test_scaled, spectrum_color_for_index(0), active_result.test_name))
                else:
                    self.highlighted_spectrum_labels.discard(active_dataset.label)
                subtitle = (
                    f"{active_result.reference_name} vs {active_result.test_name} | "
                    f"积分参考 {active_result.reference_total:.2f} W/m^2，测试 {active_result.test_scaled_total:.2f} W/m^2"
                )
                if active_dataset is not None and not self._is_spectrum_visible(active_dataset.label):
                    subtitle += "；当前测试光谱已隐藏。"

        self.plot_title_var.set(title)
        x = np.asarray(x, dtype=float)
        all_y = np.concatenate([item.y_values for item in series])
        finite_x = x[np.isfinite(x)]
        finite_y = all_y[np.isfinite(all_y)]
        if finite_x.size < 2 or finite_y.size == 0:
            self._show_plot_message(canvas, width, height, "曲线数据点不足")
            return
        self._clear_plot_message()

        x_min, x_max = float(np.min(finite_x)), float(np.max(finite_x))
        if x_max <= x_min:
            x_min -= 0.5
            x_max += 0.5

        raw_y_min, raw_y_max = float(np.min(finite_y)), float(np.max(finite_y))
        if raw_y_min >= 0:
            y_min, y_max = 0.0, raw_y_max
        elif raw_y_max <= 0:
            y_min, y_max = raw_y_min, 0.0
        else:
            y_min, y_max = raw_y_min, raw_y_max
        if y_max <= y_min:
            span = abs(y_max) if y_max else 1.0
            y_min -= span * 0.5
            y_max += span * 0.5
        padding = (y_max - y_min) * 0.08
        if raw_y_min < 0:
            y_min -= padding
        if raw_y_max > 0:
            y_max += padding

        legend_layout = _plot_legend_layout(len(series), width, height, pad_left, pad_top, pad_bottom)
        pad_right = int(math.ceil(legend_layout.panel_width + legend_layout.panel_gap))
        plot_w = max(80, width - pad_left - pad_right)
        plot_h = height - pad_top - pad_bottom
        plot_right = pad_left + plot_w

        def px(x_value: float) -> float:
            return pad_left + (x_value - x_min) / (x_max - x_min) * plot_w

        def py(y_value: float) -> float:
            return pad_top + plot_h - ((y_value - y_min) / (y_max - y_min)) * plot_h

        for i in range(6):
            y_value = y_min + (y_max - y_min) * i / 5
            yy = py(y_value)
            canvas.create_line(pad_left, yy, plot_right, yy, fill=UI_GRID_MAJOR)
            canvas.create_text(pad_left - 8, yy, text=fmt_number(y_value, 3), fill=UI_TEXT_MUTED, anchor="e", font=(UI_FONT_FAMILY, 8))

        for x_value in np.linspace(x_min, x_max, 7):
            xx = px(float(x_value))
            canvas.create_line(xx, pad_top, xx, height - pad_bottom, fill=UI_GRID)
            tick_text = f"{x_value:.0f}" if x_label == "Wavelength (nm)" else fmt_number(float(x_value), 4)
            canvas.create_text(xx, height - pad_bottom + 18, text=tick_text, fill=UI_TEXT_MUTED, font=(UI_FONT_FAMILY, 8))

        canvas.create_line(pad_left, pad_top, pad_left, height - pad_bottom, fill=UI_AXIS, width=1)
        canvas.create_line(pad_left, height - pad_bottom, plot_right, height - pad_bottom, fill=UI_AXIS, width=1)
        if y_min < 0 < y_max:
            canvas.create_line(pad_left, py(0.0), plot_right, py(0.0), fill=UI_BORDER_STRONG, width=1)
        if isc_marker is not None and np.isfinite(isc_marker[0]):
            marker_y = py(isc_marker[0])
            canvas.create_line(pad_left - 6, marker_y, pad_left + 6, marker_y, fill="#dc2626", width=2)
            canvas.create_text(
                pad_left + 10,
                marker_y,
                text=isc_marker[1],
                fill="#dc2626",
                anchor="w",
                font=(UI_FONT_FAMILY, 9, "bold"),
            )
        canvas.create_text(pad_left + plot_w / 2, height - 12, text=x_label, fill=UI_TEXT, font=(UI_FONT_FAMILY, 9))
        canvas.create_text(18, height / 2, text=y_label, fill=UI_TEXT, angle=90, font=(UI_FONT_FAMILY, 9))

        active_highlights = set(self.highlighted_spectrum_labels)
        for item in series:
            faded = bool(
                active_highlights
                and item.spectrum_label
                and item.spectrum_label not in active_highlights
                and not item.is_reference
            )
            line_color = FADED_SERIES_COLOR if faded else item.color
            line_width = 3 if item.spectrum_label and item.spectrum_label in active_highlights else 2
            if faded:
                line_width = 1
            points = []
            for x_value, y_value in zip(x, item.y_values):
                if not np.isfinite(x_value) or not np.isfinite(y_value):
                    continue
                points.extend([px(float(x_value)), py(float(y_value))])
            if len(points) >= 4:
                line_options: dict[str, object] = {
                    "fill": line_color,
                    "width": line_width,
                    "smooth": mode != "IV 曲线",
                }
                if item.dash:
                    line_options["dash"] = item.dash
                canvas.create_line(*points, **line_options)

        panel_left = pad_left + plot_w + legend_layout.panel_gap / 2
        canvas.create_rectangle(panel_left, pad_top, width - 8, height - pad_bottom, fill="#fbfcfe", outline=UI_BORDER)
        legend_x = pad_left + plot_w + legend_layout.panel_gap
        legend_y = pad_top + 8
        legend_font = (UI_FONT_FAMILY, legend_layout.font_size)
        legend_font_bold = (UI_FONT_FAMILY, legend_layout.font_size, "bold")
        for idx, item in enumerate(series):
            column = idx // legend_layout.rows_per_column
            row = idx % legend_layout.rows_per_column
            item_x = legend_x + column * legend_layout.column_width
            y0 = legend_y + row * legend_layout.row_step
            line_end = item_x + legend_layout.line_length
            highlighted = bool(item.spectrum_label and item.spectrum_label in active_highlights)
            faded = bool(
                active_highlights
                and item.spectrum_label
                and item.spectrum_label not in active_highlights
                and not item.is_reference
            )
            legend_color = FADED_SERIES_COLOR if faded else item.color
            legend_text_color = FADED_LEGEND_TEXT_COLOR if faded else UI_TEXT
            legend_width = 4 if highlighted else 3
            legend_line_options: dict[str, object] = {
                "fill": legend_color,
                "width": legend_width,
            }
            if item.dash:
                legend_line_options["dash"] = item.dash
            canvas.create_line(item_x, y0, line_end, y0, **legend_line_options)
            text_x = line_end + 8
            text_width = max(16, legend_layout.column_width - legend_layout.line_length - 14)
            legend_text = _fit_canvas_text(canvas, item.label, legend_font, text_width)
            canvas.create_text(
                text_x,
                y0,
                text=legend_text,
                fill=legend_text_color,
                anchor="w",
                font=legend_font_bold if highlighted else legend_font,
            )
            hitbox_label = item.spectrum_label if item.spectrum_label and not item.is_reference else None
            self.legend_hitboxes.append(
                (
                    item_x - 6,
                    y0 - max(8, legend_layout.row_step / 2),
                    item_x + legend_layout.column_width - 6,
                    y0 + max(8, legend_layout.row_step / 2),
                    hitbox_label,
                )
            )

        subtitle_font = (UI_FONT_FAMILY, 8)
        canvas.create_text(pad_left, 12, text=_fit_canvas_text(canvas, subtitle, subtitle_font, plot_w), fill=UI_TEXT_MUTED, anchor="w", font=subtitle_font)

    def _on_plot_double_click(self, event: tk.Event) -> str:
        for x0, y0, x1, y1, label in reversed(self.legend_hitboxes):
            if x0 <= event.x <= x1 and y0 <= event.y <= y1:
                if label:
                    if label in self.highlighted_spectrum_labels:
                        self.highlighted_spectrum_labels.remove(label)
                    elif self._is_spectrum_visible(label):
                        self.highlighted_spectrum_labels.add(label)
                    self._prune_spectrum_display_state()
                    self.draw_plot()
                return "break"
        if self.highlighted_spectrum_labels:
            self.highlighted_spectrum_labels.clear()
            self.draw_plot()
        return "break"

    def export_result(self) -> None:
        if self.last_result is None:
            messagebox.showinfo("暂无结果", "请先完成一次计算。")
            return

        path_str = filedialog.asksaveasfilename(
            title="导出 SPD 计算结果",
            initialdir=str(APP_DIR),
            initialfile="SPD计算结果.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("CSV 文件", "*.csv")],
        )
        if not path_str:
            return
        path = Path(path_str)
        try:
            self._refresh_correction_if_ready(show_errors=False)
            evaluations = self.last_correction_evaluations
            export_evaluations = [evaluation for evaluation in evaluations if evaluation.valid] if evaluations else []
            correction_groups = self._group_correction_evaluations_by_spectrum(export_evaluations) if evaluations else None
            first_correction = export_evaluations[0].correction if export_evaluations else None
            if path.suffix.lower() == ".csv":
                self._export_csv_frame(self.last_result, export_evaluations).to_csv(path, index=False, encoding="utf-8-sig")
            else:
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    self._summary_frame(self.last_result, export_evaluations, correction_groups).to_excel(writer, sheet_name="Summary", index=False)
                    if correction_groups is not None:
                        self._correction_group_frame(correction_groups).to_excel(writer, sheet_name="CV_Groups", index=False)
                    self._band_rows_frame(correction_groups).to_excel(writer, sheet_name="Bands", index=False)
                    self._result_data_frame_for_export(correction_groups).to_excel(writer, sheet_name="Data", index=False)
                    if evaluations:
                        self._correction_summary_frame(export_evaluations).to_excel(writer, sheet_name="ISC_Correction", index=False)
                    if correction_groups:
                        self._correction_data_frame_for_groups(correction_groups).to_excel(writer, sheet_name="SR_Data", index=False)
                    elif first_correction is not None:
                        self._correction_data_frame(first_correction).to_excel(writer, sheet_name="SR_Data", index=False)
                    self._add_charts_sheet(writer, self.last_result, self.iv_curves, correction_groups)
                    self._format_summary_sheet(writer, correction_groups)
            input_path = self._normalized_input_path(path)
            try:
                self._save_input_state_path(input_path)
                self.status_var.set(f"已导出：{path.name}；已保存输入文件：{input_path.name}")
                messagebox.showinfo("导出完成", f"结果已保存到：\n{path}\n\n输入文件已同步保存到：\n{input_path}")
            except Exception as exc:
                self.status_var.set(f"已导出：{path.name}；输入文件保存失败：{exc}")
                messagebox.showwarning(
                    "输入文件保存失败",
                    f"结果已保存到：\n{path}\n\n但输入文件未能保存：\n{exc}",
                )
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc))

    def export_spectrum_report(self) -> None:
        try:
            results, result = self._calculate_spectrum_results()
            self.last_results_by_label = results
            self.last_result = result
            self._render_result(result)
        except Exception as exc:
            messagebox.showerror("导出失败", f"光谱数据无法计算：{exc}")
            self.status_var.set(f"光谱导出失败：{exc}")
            return

        path_str = filedialog.asksaveasfilename(
            title="导出光谱相关数据",
            initialdir=str(APP_DIR),
            initialfile="光谱相关数据.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not path_str:
            return
        path = Path(path_str)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        try:
            factors, factor_errors = self._calculate_spectrum_mismatch_factors()
            used_titles = {"Summary"}
            sheet_titles = {
                dataset.label: self._safe_sheet_title(dataset.name or dataset.label, used_titles)
                for dataset in self.test_spectra
                if dataset.label in self.last_results_by_label
            }
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                summary_header_row = 29
                self._spectrum_report_summary_frame(factors, factor_errors, sheet_titles).to_excel(
                    writer,
                    sheet_name="Summary",
                    index=False,
                    startrow=summary_header_row - 1,
                )
                workbook = writer.book
                for page_index, dataset in enumerate(self.test_spectra, start=1):
                    dataset_result = self.last_results_by_label.get(dataset.label)
                    sheet_title = sheet_titles.get(dataset.label)
                    if dataset_result is None or not sheet_title:
                        continue
                    sheet = workbook.create_sheet(sheet_title)
                    writer.sheets[sheet.title] = sheet
                    self._write_spectrum_report_page(
                        sheet,
                        dataset,
                        dataset_result,
                        factors.get(dataset.label),
                        factor_errors.get(dataset.label, ""),
                        page_index,
                    )
                self._add_spectrum_report_summary_chart(writer, result, summary_header_row)
                self._format_spectrum_report_summary_sheet(writer, summary_header_row)
            self.status_var.set(f"已导出光谱数据：{path.name}")
            messagebox.showinfo("导出完成", f"光谱相关数据已保存到：\n{path}")
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc))

    def _calculate_spectrum_mismatch_factors(
        self,
    ) -> tuple[dict[str, CurrentCorrectionResult], dict[str, str]]:
        factors: dict[str, CurrentCorrectionResult] = {}
        errors: dict[str, str] = {}

        if self.reference_df is None:
            return factors, {dataset.label: "未载入参考光谱，MG/MMF 未计算。" for dataset in self.test_spectra}
        if self.sr_df is None or not self.sr_wave_col_var.get() or not self.sr_col_var.get():
            return factors, {dataset.label: "未载入 SR 或未选择 SR 列，MG/MMF 未计算。" for dataset in self.test_spectra}

        ref_wave_col = guess_wavelength_column(normalized_columns(self.reference_df))
        ref_irrad_col = guess_irradiance_column(normalized_columns(self.reference_df), self.ref_kind_var.get())
        if ref_wave_col is None or ref_irrad_col is None:
            return factors, {dataset.label: "参考表格中无法识别波长列或参考辐照度列。" for dataset in self.test_spectra}

        try:
            reference = spectrum_from_dataframe(self.reference_df, ref_wave_col, ref_irrad_col, 1.0)
            sr = response_from_dataframe(self.sr_df, self.sr_wave_col_var.get(), self.sr_col_var.get())
            temperature_coefficient = float(self.temp_coeff_var.get() or DEFAULT_ISC_TEMP_COEFF_PERCENT_PER_C)
        except Exception as exc:
            return factors, {dataset.label: str(exc) for dataset in self.test_spectra}

        for dataset in self.test_spectra:
            try:
                unit_factor = UNIT_FACTORS.get(dataset.unit_name, 1.0)
                test = spectrum_from_dataframe(dataset.df, dataset.wavelength_column, dataset.irradiance_column, unit_factor)
                factors[dataset.label] = calculate_current_correction(
                    reference,
                    test,
                    sr,
                    sr_name=self.sr_path.name if self.sr_path is not None else "SR",
                    sr_sheet_name=self.sr_sheet_var.get(),
                    sr_wavelength_column=self.sr_wave_col_var.get(),
                    sr_column=self.sr_col_var.get(),
                    measured_isc=1.0,
                    test_temperature_c=DEFAULT_REFERENCE_TEMPERATURE_C,
                    temperature_coefficient_percent_per_c=temperature_coefficient,
                    start_nm=CURRENT_CORRECTION_START_NM,
                    end_nm=CURRENT_CORRECTION_END_NM,
                )
            except Exception as exc:
                errors[dataset.label] = str(exc)
        return factors, errors

    def _spectrum_report_summary_frame(
        self,
        factors: dict[str, CurrentCorrectionResult],
        factor_errors: dict[str, str],
        sheet_titles: dict[str, str],
    ) -> pd.DataFrame:
        columns = [
            "页面",
            "光谱标签",
            "光谱名称",
            "文件",
            "工作表",
            "波长列",
            "辐照度列",
            "单位",
            "计算波段_nm",
            "参考积分_W_m2",
            "测试原始积分_W_m2",
            "SPD_%",
            "SPC_%",
            "分段等级",
            "MG",
            "MMF",
            "MG/MMF说明",
        ]
        rows = []
        for dataset in self.test_spectra:
            result = self.last_results_by_label.get(dataset.label)
            if result is None:
                continue
            factor = factors.get(dataset.label)
            rows.append(
                {
                    "页面": sheet_titles.get(dataset.label, dataset.label),
                    "光谱标签": dataset.label,
                    "光谱名称": dataset.name,
                    "文件": dataset.path.name if dataset.path is not None else "",
                    "工作表": dataset.sheet_name,
                    "波长列": dataset.wavelength_column,
                    "辐照度列": dataset.irradiance_column,
                    "单位": dataset.unit_name,
                    "计算波段_nm": f"{result.start_nm:g}-{result.end_nm:g}",
                    "参考积分_W_m2": result.reference_total,
                    "测试原始积分_W_m2": result.test_raw_total,
                    "SPD_%": result.spd_percent,
                    "SPC_%": result.spc_percent,
                    "分段等级": result.overall_class,
                    "MG": factor.mg if factor is not None else "",
                    "MMF": factor.mmf if factor is not None else "",
                    "MG/MMF说明": factor_errors.get(dataset.label, "MG/MMF 使用 SR 波段计算，不包含电流。"),
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _spectrum_report_data_frame(self, dataset: TestSpectrumDataset, result: CalculationResult) -> pd.DataFrame:
        unit_factor = UNIT_FACTORS.get(dataset.unit_name, 1.0)
        raw_spectrum = spectrum_from_dataframe(
            dataset.df,
            dataset.wavelength_column,
            dataset.irradiance_column,
            unit_factor,
        )
        return pd.DataFrame(
            {
                "原始波长_nm": pd.Series(raw_spectrum.wavelength_nm),
                "原始辐照度_W_m2_nm": pd.Series(raw_spectrum.irradiance_w_m2_nm),
            }
        )

    def _write_spectrum_report_page(
        self,
        sheet,
        dataset: TestSpectrumDataset,
        result: CalculationResult,
        factor: CurrentCorrectionResult | None,
        factor_error: str,
        page_index: int,
    ) -> None:
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows

        sheet.sheet_view.showGridLines = False
        widths = [20, 42, 16, 16, 16, 16, 16, 16]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        title_fill = PatternFill("solid", fgColor="EAF2FF")
        metric_fill = PatternFill("solid", fgColor="F8FAFC")
        header_fill = PatternFill("solid", fgColor="E2E8F0")

        sheet["A1"] = dataset.name or dataset.label
        sheet["A1"].font = Font(bold=True, size=14, color="0F172A")
        sheet["A1"].fill = title_fill
        sheet["A1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        sheet.row_dimensions[1].height = 28

        source_name = dataset.path.name if dataset.path is not None else "未记录文件"
        sr_band = f"{factor.grid_nm[0]:g}-{factor.grid_nm[-1]:g}" if factor is not None else ""
        factor_note = factor_error or "MG/MMF 使用 SR 波段计算，不包含电流。"
        metric_rows: list[tuple[str, object]] = [
            ("文件", source_name),
            ("工作表", dataset.sheet_name or "-"),
            ("波长列", dataset.wavelength_column),
            ("辐照度列", dataset.irradiance_column),
            ("输入单位", dataset.unit_name),
            ("参考光谱", result.reference_name),
            ("计算波段_nm", f"{result.start_nm:g}-{result.end_nm:g}"),
            ("参考积分_W_m2", result.reference_total),
            ("测试原始积分_W_m2", result.test_raw_total),
            ("SPD_%", result.spd_percent),
            ("SPC_%", result.spc_percent),
            ("分段等级", result.overall_class),
            ("MG", factor.mg if factor is not None else ""),
            ("MMF", factor.mmf if factor is not None else ""),
            ("SR波段_nm", sr_band),
            ("SR参考积分_W_m2", factor.reference_irradiance_integral if factor is not None else ""),
            ("SR测试积分_W_m2", factor.test_irradiance_integral if factor is not None else ""),
            ("说明", factor_note),
        ]
        sheet["A3"] = "光谱指标"
        sheet["A3"].font = Font(bold=True, color="334155")
        sheet["A3"].fill = title_fill
        sheet.row_dimensions[3].height = 22

        for row_offset, (label, value) in enumerate(metric_rows, start=4):
            label_cell = sheet.cell(row=row_offset, column=1, value=label)
            value_cell = sheet.cell(row=row_offset, column=2, value=self._excel_cell_value(value))
            label_cell.font = Font(bold=True, color="334155")
            label_cell.fill = metric_fill
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
            if isinstance(value_cell.value, float):
                value_cell.number_format = "0.000000"
            sheet.row_dimensions[row_offset].height = 22 if label != "说明" else 42

        chart_title_row = 24
        sheet.cell(row=chart_title_row, column=1, value="光谱与标准光谱对比图")
        sheet.cell(row=chart_title_row, column=1).font = Font(bold=True, color="334155")
        sheet.cell(row=chart_title_row, column=1).fill = title_fill
        image_width, image_height = _plot_image_dimensions_for_series(2, base_width=1080, base_height=340)
        spectrum_image = _render_plot_image(
            result.grid_nm,
            [
                self._plot_reference_series(result.reference, reference_legend_label(result.reference_name)),
                self._plot_series_for_dataset(dataset, result.test_raw),
            ],
            "Wavelength (nm)",
            "W/(m^2 nm)",
            f"SPD {result.spd_percent:.2f}% | SPC {result.spc_percent:.2f}% | MG {fmt_number(factor.mg, 6) if factor is not None else '-'} | MMF {fmt_number(factor.mmf, 6) if factor is not None else '-'}",
            width=image_width,
            height=image_height,
        )
        if spectrum_image is not None:
            image = ExcelImage(spectrum_image)
            image.width = image_width
            image.height = image_height
            sheet.add_image(image, "A25")
        else:
            sheet["A25"] = "光谱数据点不足，无法生成光谱对比图。"
            sheet["A25"].font = Font(color="64748B")
        for row in range(25, 46):
            sheet.row_dimensions[row].height = 18

        data_start_row = 48
        sheet.cell(row=data_start_row - 1, column=1, value="原始光谱数据")
        sheet.cell(row=data_start_row - 1, column=1).font = Font(bold=True, color="334155")
        sheet.cell(row=data_start_row - 1, column=1).fill = title_fill

        frame = self._spectrum_report_data_frame(dataset, result)
        for row_index, row in enumerate(dataframe_to_rows(frame, index=False, header=True), start=data_start_row):
            for column_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=self._excel_cell_value(value))
                if row_index == data_start_row:
                    cell.font = Font(bold=True, color="0F172A")
                    cell.fill = header_fill
                elif isinstance(cell.value, float):
                    cell.number_format = "0.000000"

        last_column = get_column_letter(frame.shape[1])
        sheet.auto_filter.ref = f"A{data_start_row}:{last_column}{sheet.max_row}"
        sheet.freeze_panes = None

    def _add_spectrum_report_summary_chart(
        self,
        writer: pd.ExcelWriter,
        result: CalculationResult,
        table_header_row: int,
    ) -> None:
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Font, PatternFill

        sheet = writer.sheets.get("Summary")
        if sheet is None:
            return

        title_fill = PatternFill("solid", fgColor="EAF2FF")
        sheet["A1"] = "所有光谱对比图"
        sheet["A1"].font = Font(bold=True, size=14, color="0F172A")
        sheet["A1"].fill = title_fill
        sheet["A1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        sheet["A2"] = f"参考光谱：{reference_legend_label(result.reference_name)}；图例按光谱名称显示。"
        sheet["A2"].font = Font(color="475569")
        sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        sheet.merge_cells("A1:H1")
        sheet.merge_cells("A2:H2")
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 24

        spectrum_series = [self._plot_reference_series(result.reference, reference_legend_label(result.reference_name))]
        for dataset in self.test_spectra:
            dataset_result = self.last_results_by_label.get(dataset.label)
            if dataset_result is not None:
                spectrum_series.append(self._plot_series_for_dataset(dataset, dataset_result.test_scaled))

        image_width, image_height = _plot_image_dimensions_for_series(len(spectrum_series))
        spectrum_image = _render_plot_image(
            result.grid_nm,
            spectrum_series,
            "Wavelength (nm)",
            "W/(m^2 nm)",
            f"已绘制 {max(len(spectrum_series) - 1, 0)} 组测试光谱；图例按光谱名称显示。",
            width=image_width,
            height=image_height,
        )
        if spectrum_image is not None:
            image = ExcelImage(spectrum_image)
            image.width = image_width
            image.height = image_height
            sheet.add_image(image, "A4")
        else:
            sheet["A4"] = "光谱数据点不足，无法生成光谱对比图。"
            sheet["A4"].font = Font(color="64748B")
        for row in range(4, table_header_row - 1):
            sheet.row_dimensions[row].height = 18

        table_title_row = table_header_row - 1
        sheet.cell(row=table_title_row, column=1, value="光谱数据汇总")
        sheet.cell(row=table_title_row, column=1).font = Font(bold=True, color="334155")
        sheet.cell(row=table_title_row, column=1).fill = title_fill

    def _format_spectrum_report_summary_sheet(self, writer: pd.ExcelWriter, header_row: int = 1) -> None:
        from openpyxl.styles import Font, PatternFill

        sheet = writer.sheets.get("Summary")
        if sheet is None:
            return
        header_fill = PatternFill("solid", fgColor="EAF2FF")
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = header_fill
        for column_cells in sheet.iter_cols(min_row=header_row, max_row=sheet.max_row):
            column_letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 36)
        sheet.freeze_panes = None
        sheet.sheet_view.showGridLines = False

    @staticmethod
    def _safe_sheet_title(base: str, used_titles: set[str]) -> str:
        cleaned = "".join(
            char if unicodedata.category(char)[0] in {"L", "N"} else "_"
            for char in str(base or "")
        )
        cleaned = re.sub(r"_+", "_", cleaned).strip("_") or "Sheet"
        cleaned = cleaned[:31].rstrip("_") or "Sheet"
        candidate = cleaned
        counter = 2
        used_lower = {title.lower() for title in used_titles}
        while candidate.lower() in used_lower:
            suffix = f"_{counter}"
            candidate = f"{cleaned[:31 - len(suffix)].rstrip('_')}{suffix}"
            counter += 1
        used_titles.add(candidate)
        return candidate

    @staticmethod
    def _excel_cell_value(value: object) -> object:
        if value is None:
            return None
        if isinstance(value, (float, np.floating)):
            return float(value) if np.isfinite(value) else None
        if isinstance(value, (int, np.integer)):
            return int(value)
        return value

    def _group_correction_evaluations_by_spectrum(
        self,
        evaluations: list[CurrentCorrectionEvaluation],
    ) -> list[tuple[str, str, list[CurrentCorrectionEvaluation]]]:
        grouped: dict[str, tuple[str, list[CurrentCorrectionEvaluation]]] = {}
        for evaluation in evaluations:
            spectrum_label = evaluation.spectrum_label or evaluation.spectrum_name or "未指定光谱"
            spectrum_name = evaluation.spectrum_name or self._spectrum_name_for_label(spectrum_label) or spectrum_label
            if spectrum_label not in grouped:
                grouped[spectrum_label] = (spectrum_name, [])
            grouped[spectrum_label][1].append(evaluation)

        ordered_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]] = []
        seen: set[str] = set()
        for dataset in self.test_spectra:
            if dataset.label in grouped:
                spectrum_name, group_evaluations = grouped[dataset.label]
                ordered_groups.append((dataset.label, spectrum_name, group_evaluations))
                seen.add(dataset.label)
        for spectrum_label, (spectrum_name, group_evaluations) in grouped.items():
            if spectrum_label not in seen:
                ordered_groups.append((spectrum_label, spectrum_name, group_evaluations))
        return ordered_groups

    def _correction_group_frame(
        self,
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]],
    ) -> pd.DataFrame:
        columns = [
            "使用光谱标签",
            "使用光谱",
            "有效CV组数",
            "有效电流组",
            "电流来源",
            "组内有效CV平均值_mA",
            "组内短路电流平均值_mA",
            "SPD_%",
            "SPC_%",
            "分段等级",
            "MG",
            "MMF",
            "SR文件",
            "SR工作表",
            "SR波段_nm",
            "标准温度_C",
            "温度系数_%_C",
            "测试温度_C",
        ]
        rows = []
        for spectrum_label, spectrum_name, group_evaluations in correction_groups:
            if not group_evaluations:
                continue
            corrections = [evaluation.correction for evaluation in group_evaluations]
            first = corrections[0]
            result = self.last_results_by_label.get(spectrum_label)
            rows.append(
                {
                    "使用光谱标签": spectrum_label,
                    "使用光谱": spectrum_name,
                    "有效CV组数": len(group_evaluations),
                    "有效电流组": "、".join(evaluation.label for evaluation in group_evaluations),
                    "电流来源": "；".join(dict.fromkeys(evaluation.source for evaluation in group_evaluations if evaluation.source)),
                    "组内有效CV平均值_mA": float(np.mean([correction.corrected_cv for correction in corrections])),
                    "组内短路电流平均值_mA": float(np.mean([correction.measured_isc for correction in corrections])),
                    "SPD_%": result.spd_percent if result is not None else "",
                    "SPC_%": result.spc_percent if result is not None else "",
                    "分段等级": result.overall_class if result is not None else "",
                    "MG": first.mg,
                    "MMF": first.mmf,
                    "SR文件": first.sr_name,
                    "SR工作表": first.sr_sheet_name,
                    "SR波段_nm": f"{first.grid_nm[0]:g}-{first.grid_nm[-1]:g}",
                    "标准温度_C": first.reference_temperature_c,
                    "温度系数_%_C": first.temperature_coefficient_percent_per_c,
                    "测试温度_C": "、".join(
                        dict.fromkeys(fmt_number(correction.test_temperature_c, 5) for correction in corrections)
                    ),
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _band_rows_frame(
        self,
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]] | None = None,
    ) -> pd.DataFrame:
        if correction_groups is None:
            return pd.DataFrame(self.last_result.band_rows if self.last_result is not None else [])
        frames = []
        for spectrum_label, spectrum_name, group_evaluations in correction_groups:
            result = self.last_results_by_label.get(spectrum_label)
            if result is None:
                continue
            frame = pd.DataFrame(result.band_rows)
            if frame.empty:
                continue
            frame.insert(0, "使用光谱标签", spectrum_label)
            frame.insert(1, "使用光谱", spectrum_name)
            frames.append(frame)
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    def _result_data_frame_for_export(
        self,
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]] | None = None,
    ) -> pd.DataFrame:
        if correction_groups is None:
            return self._result_data_frame(self.last_result)
        frames = []
        for spectrum_label, spectrum_name, group_evaluations in correction_groups:
            result = self.last_results_by_label.get(spectrum_label)
            if result is None:
                continue
            frame = self._result_data_frame(result)
            frame.insert(0, "使用光谱标签", spectrum_label)
            frames.append(frame)
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    def _summary_frame(
        self,
        result: CalculationResult,
        evaluations: list[CurrentCorrectionEvaluation] | None = None,
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]] | None = None,
    ) -> pd.DataFrame:
        columns = [
            "光谱标签",
            "用于CV修正",
            "测试光谱",
            "参考光谱",
            "计算波段_nm",
            "网格步长_nm",
            "SPD_%",
            "SPC_%",
            "分段等级",
            "参考积分_W_m2",
            "测试原始积分_W_m2",
            "有效CV组数",
            "有效电流组",
            "有效CV平均值_mA",
            "SR文件",
            "导出说明",
        ]
        group_by_label: dict[str, list[CurrentCorrectionEvaluation]] = {}
        if correction_groups:
            group_by_label = {spectrum_label: group_evaluations for spectrum_label, _name, group_evaluations in correction_groups}

        labels_in_order = [dataset.label for dataset in self.test_spectra if dataset.label in self.last_results_by_label]
        for label in self.last_results_by_label:
            if label not in labels_in_order:
                labels_in_order.append(label)
        if not labels_in_order and result is not None:
            labels_in_order = [""]

        rows = []
        for spectrum_label in labels_in_order:
            dataset_result = self.last_results_by_label.get(spectrum_label, result)
            dataset = self._test_spectrum_by_label(spectrum_label)
            group_evaluations = group_by_label.get(spectrum_label, [])
            valid_cvs = [evaluation.correction.corrected_cv for evaluation in group_evaluations]
            first_correction = group_evaluations[0].correction if group_evaluations else None
            rows.append(
                {
                    "光谱标签": spectrum_label,
                    "用于CV修正": "是" if group_evaluations else "否",
                    "测试光谱": dataset.name if dataset is not None else dataset_result.test_name,
                    "参考光谱": dataset_result.reference_name,
                    "计算波段_nm": f"{dataset_result.start_nm:g}-{dataset_result.end_nm:g}",
                    "网格步长_nm": dataset_result.step_nm,
                    "SPD_%": dataset_result.spd_percent,
                    "SPC_%": dataset_result.spc_percent,
                    "分段等级": dataset_result.overall_class,
                    "参考积分_W_m2": dataset_result.reference_total,
                    "测试原始积分_W_m2": dataset_result.test_raw_total,
                    "有效CV组数": len(group_evaluations) if group_evaluations else "",
                    "有效电流组": "、".join(evaluation.label for evaluation in group_evaluations),
                    "有效CV平均值_mA": float(np.mean(valid_cvs)) if valid_cvs else "",
                    "SR文件": first_correction.sr_name if first_correction is not None else "",
                    "导出说明": "仅保留A5合格CV；分组修正参数见CV_Groups" if group_evaluations else "",
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _format_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]] | None,
    ) -> None:
        from openpyxl.styles import Font, PatternFill

        sheet = writer.sheets.get("Summary")
        if sheet is None:
            return
        used_labels = {spectrum_label for spectrum_label, _name, _group in correction_groups or []}
        header_fill = PatternFill("solid", fgColor="EAF2FF")
        used_fill = PatternFill("solid", fgColor="FFF2CC")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = header_fill
        label_column = None
        for cell in sheet[1]:
            if cell.value == "光谱标签":
                label_column = cell.column
                break
        if label_column is not None:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=label_column)
                if cell.value in used_labels:
                    cell.fill = used_fill
                    cell.font = Font(bold=True, color="7A4F00")
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 34)
        sheet.freeze_panes = "A2"

    def _result_data_frame(self, result: CalculationResult) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Wavelength_nm": result.grid_nm,
                "Reference_W_m2_nm": result.reference,
                "Test_raw_W_m2_nm": result.test_raw,
                "Test_scaled_W_m2_nm": result.test_scaled,
                "Abs_error_W_m2_nm": result.absolute_error,
            }
        )

    def _correction_summary_frame(self, evaluations: list[CurrentCorrectionEvaluation]) -> pd.DataFrame:
        columns = [
            "组别",
            "使用光谱标签",
            "使用光谱",
            "电流来源",
            "Ii_mA",
            "CVi_mA",
            "偏离有效均值_%",
            "测试温度_C",
            "MG",
            "MT",
            "MMF",
        ]
        rows = []
        for evaluation in evaluations:
            correction = evaluation.correction
            rows.append(
                {
                    "组别": evaluation.label,
                    "使用光谱标签": evaluation.spectrum_label,
                    "使用光谱": evaluation.spectrum_name,
                    "电流来源": evaluation.source,
                    "Ii_mA": correction.measured_isc,
                    "CVi_mA": correction.corrected_cv,
                    "偏离有效均值_%": evaluation.deviation_percent,
                    "测试温度_C": correction.test_temperature_c,
                    "MG": correction.mg,
                    "MT": correction.mt,
                    "MMF": correction.mmf,
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _correction_data_frame(self, correction: CurrentCorrectionResult) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Wavelength_nm": correction.grid_nm,
                "SR_A_W": correction.sr,
                "Reference_W_m2_nm": correction.reference,
                "Test_W_m2_nm": correction.test,
                "Reference_times_SR": correction.reference_sr_weighted,
                "Test_times_SR": correction.test_sr_weighted,
            }
        )

    def _correction_data_frame_for_groups(
        self,
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]],
    ) -> pd.DataFrame:
        frames = []
        for spectrum_label, spectrum_name, group_evaluations in correction_groups:
            if not group_evaluations:
                continue
            correction = group_evaluations[0].correction
            frame = self._correction_data_frame(correction)
            frame.insert(0, "使用光谱标签", spectrum_label)
            frames.append(frame)
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    def _iv_data_frame(self, curves: list[IVCurve]) -> pd.DataFrame:
        columns = ["组别", "电流来源", "工作表", "短路电流ISC_mA"]
        rows = []
        for curve in curves:
            rows.append(
                {
                    "组别": curve.label,
                    "电流来源": curve.source_name,
                    "工作表": curve.sheet_name,
                    "短路电流ISC_mA": curve.isc_ma,
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _iv_curve_for_evaluation(self, evaluation: CurrentCorrectionEvaluation) -> IVCurve | None:
        current = self._current_input_by_label(evaluation.label)
        if current is not None and current.iv_curve is not None:
            return current.iv_curve
        return self._iv_curve_by_label(evaluation.label)

    def _iv_data_frame_for_evaluations(self, evaluations: list[CurrentCorrectionEvaluation]) -> pd.DataFrame:
        columns = ["组别", "使用光谱标签", "使用光谱", "电流来源", "工作表", "短路电流ISC_mA"]
        rows = []
        for evaluation in evaluations:
            curve = self._iv_curve_for_evaluation(evaluation)
            rows.append(
                {
                    "组别": evaluation.label,
                    "使用光谱标签": evaluation.spectrum_label,
                    "使用光谱": evaluation.spectrum_name,
                    "电流来源": curve.source_name if curve is not None else evaluation.source,
                    "工作表": curve.sheet_name if curve is not None else "",
                    "短路电流ISC_mA": curve.isc_ma if curve is not None else evaluation.correction.measured_isc,
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _add_charts_sheet(
        self,
        writer: pd.ExcelWriter,
        result: CalculationResult,
        iv_curves: list[IVCurve],
        correction_groups: list[tuple[str, str, list[CurrentCorrectionEvaluation]]] | None = None,
    ) -> None:
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.pagebreak import Break

        workbook = writer.book
        used_sheet_titles = {worksheet.title for worksheet in workbook.worksheets}

        def create_chart_sheet(title: str, index: int | None = 3):
            sheet_title = self._safe_sheet_title(title or "Charts", used_sheet_titles)
            chart_ws = workbook.create_sheet(sheet_title, index)
            writer.sheets[chart_ws.title] = chart_ws
            chart_ws.sheet_view.showGridLines = False
            chart_ws.freeze_panes = "A2"
            for column in "ABCDEFGHIJKLMN":
                chart_ws.column_dimensions[column].width = 14
            chart_ws.page_setup.orientation = "landscape"
            chart_ws.page_setup.fitToWidth = 1
            chart_ws.page_setup.fitToHeight = 0
            return chart_ws

        title_fill = PatternFill("solid", fgColor="EAF2FF")

        def write_source_note(chart_ws, row: int, start_column: int, end_column: int, iv_curve: IVCurve) -> None:
            note = f"来源：{iv_curve.source_name or '未记录'}"
            if iv_curve.sheet_name:
                note += f"；工作表：{iv_curve.sheet_name}"
            if iv_curve.first_data_row:
                note += f"；起始行：{iv_curve.first_data_row}"
            chart_ws.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)
            cell = chart_ws.cell(row=row, column=start_column)
            cell.value = note
            cell.font = Font(size=8, color="475569")
            cell.alignment = Alignment(wrap_text=False)

        if correction_groups is not None:
            if not correction_groups:
                chart_ws = create_chart_sheet("Charts", 3)
                chart_ws["A1"] = "图表"
                chart_ws["A1"].font = Font(bold=True, size=14, color="0F172A")
                chart_ws["A1"].fill = title_fill
                chart_ws.merge_cells("A1:N1")
                chart_ws["A3"] = "没有 A5 判定合格的 CV 数据，导出图表页不包含电流曲线。"
                chart_ws["A3"].font = Font(color="64748B")
                return

            for group_index, (spectrum_label, spectrum_name, group_evaluations) in enumerate(correction_groups, start=1):
                chart_ws = create_chart_sheet(spectrum_name, 3 + group_index - 1)
                page_start = 1
                valid_mean = float(np.mean([evaluation.correction.corrected_cv for evaluation in group_evaluations]))
                title_cell = f"A{page_start}"
                subtitle_cell = f"A{page_start + 1}"
                chart_ws[title_cell] = (
                    f"{spectrum_name} | 有效 CV {len(group_evaluations)} 组 | "
                    f"平均 {fmt_number(valid_mean, 6)} mA"
                )
                chart_ws[title_cell].font = Font(bold=True, size=14, color="0F172A")
                chart_ws[title_cell].fill = title_fill
                chart_ws[subtitle_cell] = f"光谱对比：{result.reference_name} vs {spectrum_name}"
                chart_ws[subtitle_cell].font = Font(bold=True, color="334155")
                chart_ws[subtitle_cell].alignment = Alignment(wrap_text=True)
                chart_ws.merge_cells(start_row=page_start, start_column=1, end_row=page_start, end_column=14)
                chart_ws.merge_cells(start_row=page_start + 1, start_column=1, end_row=page_start + 1, end_column=14)

                group_result = self.last_results_by_label.get(spectrum_label)
                if group_result is not None:
                    spectrum_x = group_result.grid_nm
                    spectrum_series = [
                        self._plot_reference_series(group_result.reference, reference_legend_label(group_result.reference_name)),
                        self._plot_series_for_label(spectrum_label, spectrum_name, group_result.test_scaled),
                    ]
                else:
                    correction = group_evaluations[0].correction
                    spectrum_x = correction.grid_nm
                    spectrum_series = [
                        self._plot_reference_series(correction.reference, reference_legend_label(result.reference_name)),
                        self._plot_series_for_label(spectrum_label, spectrum_name, correction.test),
                    ]
                spectrum_image = _render_plot_image(
                    spectrum_x,
                    spectrum_series,
                    "Wavelength (nm)",
                    "W/(m^2 nm)",
                    f"该页仅显示使用 {spectrum_name} 修正且 A5 合格的电流组。",
                )
                if spectrum_image is not None:
                    image = ExcelImage(spectrum_image)
                    image.width = 1050
                    image.height = 320
                    chart_ws.add_image(image, f"A{page_start + 3}")
                else:
                    chart_ws[f"A{page_start + 3}"] = "光谱数据点不足，无法生成光谱对比图。"
                    chart_ws[f"A{page_start + 3}"].font = Font(color="64748B")

                iv_title_row = page_start + 22
                chart_ws[f"A{iv_title_row}"] = "有效电流图（IV 曲线）"
                chart_ws[f"A{iv_title_row}"].font = Font(bold=True, color="334155")
                chart_ws[f"A{iv_title_row}"].fill = title_fill
                chart_ws.merge_cells(start_row=iv_title_row, start_column=1, end_row=iv_title_row, end_column=14)

                curve_items = [
                    (evaluation, self._iv_curve_for_evaluation(evaluation))
                    for evaluation in group_evaluations
                ]
                curve_items = [(evaluation, curve) for evaluation, curve in curve_items if curve is not None]
                current_start = page_start + 24
                if not curve_items:
                    chart_ws[f"A{current_start}"] = "该组有效 CV 来自手动电流或未关联 IV 曲线，图表页不包含电流曲线。"
                    chart_ws[f"A{current_start}"].font = Font(color="64748B")
                    page_end = current_start + 3
                else:
                    for curve_index, (evaluation, iv_curve) in enumerate(curve_items):
                        note_row = current_start + (curve_index // 2) * 15
                        anchor_row = note_row + 1
                        anchor_col = "A" if curve_index % 2 == 0 else "H"
                        note_start_col = 1 if anchor_col == "A" else 8
                        write_source_note(chart_ws, note_row, note_start_col, note_start_col + 6, iv_curve)
                        current_image = _render_plot_image(
                            iv_curve.voltage_v,
                            [(iv_curve.current_ma, "#0f766e", "实测 IV")],
                            "Voltage (V)",
                            "Current (mA)",
                            (
                                f"{evaluation.label} | ISC {fmt_number(iv_curve.isc_ma, 6)} mA | "
                                f"CV {fmt_number(evaluation.correction.corrected_cv, 6)} mA"
                            ),
                            isc_marker=(
                                (float(iv_curve.current_ma[0]), f"{fmt_number(iv_curve.isc_ma, 6)} mA")
                                if iv_curve.current_ma.size
                                else None
                            ),
                        )
                        if current_image is None:
                            chart_ws[f"{anchor_col}{anchor_row}"] = f"{evaluation.label} IV 数据点不足，无法生成电流图。"
                            chart_ws[f"{anchor_col}{anchor_row}"].font = Font(color="64748B")
                            continue
                        image = ExcelImage(current_image)
                        image.width = 520
                        image.height = 240
                        chart_ws.add_image(image, f"{anchor_col}{anchor_row}")
                    page_end = current_start + math.ceil(len(curve_items) / 2) * 15 + 1

                for row in range(page_start, page_end + 1):
                    chart_ws.row_dimensions[row].height = 18
                chart_ws.print_area = f"A1:N{page_end}"
            return

        chart_ws = create_chart_sheet("Charts", 3)
        max_chart_row = max(56, 30 + max(len(iv_curves), 1) * 24)
        for row in range(1, max_chart_row):
            chart_ws.row_dimensions[row].height = 18
        chart_ws["A1"] = "图表"
        chart_ws["A1"].font = Font(bold=True, size=14, color="0F172A")
        chart_ws["A1"].fill = title_fill
        chart_ws["A2"] = f"光谱对比：{result.reference_name} vs {result.test_name}"
        chart_ws["A2"].font = Font(bold=True, color="334155")
        chart_ws["A2"].alignment = Alignment(wrap_text=True)
        chart_ws.merge_cells("A1:H1")
        chart_ws.merge_cells("A2:H2")

        data_ws = writer.sheets["Data"]
        data_max_row = data_ws.max_row
        if data_max_row >= 3:
            spectrum_series = [self._plot_reference_series(result.reference, reference_legend_label(result.reference_name))]
            for dataset in self.test_spectra:
                dataset_result = self.last_results_by_label.get(dataset.label)
                if dataset_result is not None:
                    spectrum_series.append(self._plot_series_for_dataset(dataset, dataset_result.test_scaled))
            image_width, image_height = _plot_image_dimensions_for_series(len(spectrum_series))
            spectrum_image = _render_plot_image(
                result.grid_nm,
                spectrum_series,
                "Wavelength (nm)",
                "W/(m^2 nm)",
                f"已绘制 {max(len(spectrum_series) - 1, 0)} 组测试光谱；图例按光谱名称显示。",
                width=image_width,
                height=image_height,
            )
            if spectrum_image is not None:
                image = ExcelImage(spectrum_image)
                image.width = image_width
                image.height = image_height
                chart_ws.add_image(image, "A4")
            else:
                chart_ws["A4"] = "光谱数据点不足，无法生成光谱对比图。"
        else:
            chart_ws["A4"] = "光谱数据点不足，无法生成光谱对比图。"

        chart_ws["A28"] = "电流图（IV 曲线）"
        chart_ws["A28"].font = Font(bold=True, color="334155")
        chart_ws["A28"].fill = title_fill
        chart_ws.merge_cells("A28:H28")

        if not iv_curves:
            chart_ws["A30"] = "未读取 IV 表格，导出文件暂不包含电流曲线。"
            chart_ws["A30"].font = Font(color="64748B")
            return

        start_row = 30
        for index, iv_curve in enumerate(iv_curves, start=1):
            note_row = start_row + (index - 1) * 23
            anchor_row = note_row + 1
            write_source_note(chart_ws, note_row, 1, 8, iv_curve)
            current_image = _render_plot_image(
                iv_curve.voltage_v,
                [(iv_curve.current_ma, "#0f766e", "实测 IV")],
                "Voltage (V)",
                "Current (mA)",
                f"{iv_curve.label} | ISC {fmt_number(iv_curve.isc_ma, 6)} mA",
                isc_marker=(float(iv_curve.current_ma[0]), f"{fmt_number(iv_curve.isc_ma, 6)} mA") if iv_curve.current_ma.size else None,
            )
            if current_image is None:
                chart_ws[f"A{anchor_row}"] = f"{iv_curve.label} IV 数据点不足，无法生成电流图。"
                chart_ws[f"A{anchor_row}"].font = Font(color="64748B")
                continue
            image = ExcelImage(current_image)
            image.width = 1100
            image.height = 380
            chart_ws.add_image(image, f"A{anchor_row}")

    def _export_csv_frame(
        self,
        result: CalculationResult,
        evaluations: list[CurrentCorrectionEvaluation],
    ) -> pd.DataFrame:
        frame = self._result_data_frame(result)
        if not evaluations or frame.empty:
            return frame
        first = evaluations[0].correction
        valid_cvs = [evaluation.correction.corrected_cv for evaluation in evaluations if evaluation.valid]
        values = {
            "Correction_test_temperature_C": first.test_temperature_c,
            "Correction_MG": first.mg,
            "Correction_MT": first.mt,
            "Correction_MMF": first.mmf,
            "Correction_valid_count": len(valid_cvs),
            "Correction_valid_CV_mean_mA": float(np.mean(valid_cvs)) if valid_cvs else "",
        }
        for evaluation in evaluations:
            label = evaluation.label
            values[f"{label}_spectrum"] = evaluation.spectrum_name
            values[f"{label}_mA"] = evaluation.correction.measured_isc
            values[f"CV{label[1:]}_mA" if label.startswith("I") else f"{label}_CV_mA"] = evaluation.correction.corrected_cv
            values[f"{label}_A5_status"] = "valid" if evaluation.valid else "invalid"
            values[f"{label}_A5_reason"] = "；".join(dict.fromkeys(evaluation.reasons))
        for column, value in values.items():
            frame[column] = ""
            frame.loc[frame.index[0], column] = value
        return frame


def run_self_test() -> None:
    path = find_default_reference_file()
    if path is None:
        raise SystemExit("未找到 AM1.5 标准光谱文件。")
    df = read_table(path, list_excel_sheets(path)[0])
    df.columns = normalized_columns(df)
    wave_col = guess_wavelength_column(df.columns)
    g_col = guess_irradiance_column(df.columns, "AM1.5G")
    d_col = guess_irradiance_column(df.columns, "AM1.5D")
    if wave_col is None or g_col is None or d_col is None:
        raise SystemExit("无法识别标准光谱列。")
    reference_choices = reference_irradiance_choices(df)
    if "AM1.5G" not in reference_choices or "AM1.5D" not in reference_choices:
        raise SystemExit("参考光谱下拉选项未识别 AM1.5G / AM1.5D。")
    am0_frame = pd.DataFrame({"Wavelength/nm": [300, 301], "AM0": [1.0, 1.1]})
    if reference_irradiance_choices(am0_frame) != ["AM0"]:
        raise SystemExit("参考光谱下拉选项未识别 AM0 表头。")
    if len(TEST_SERIES_COLORS) < 20 or len(set(TEST_SERIES_COLORS)) != len(TEST_SERIES_COLORS):
        raise SystemExit("测试光谱颜色池应至少包含 20 个不重复颜色。")
    first_cycle_styles = [spectrum_style_for_index(index) for index in range(len(TEST_SERIES_COLORS))]
    if any(style.dash for style in first_cycle_styles):
        raise SystemExit("测试光谱颜色第一轮应优先使用实线。")
    repeated_style = spectrum_style_for_index(len(TEST_SERIES_COLORS))
    if repeated_style.color != first_cycle_styles[0].color or not repeated_style.dash:
        raise SystemExit("测试光谱颜色循环后应使用不同线型区分。")

    app = SpectralSPDApp.__new__(SpectralSPDApp)
    app.test_spectra = [
        TestSpectrumDataset("S1", "one", None, "", am0_frame, "Wavelength/nm", "AM0", "W/(m^2 nm)"),
        TestSpectrumDataset("S3", "three", None, "", am0_frame, "Wavelength/nm", "AM0", "W/(m^2 nm)"),
    ]
    app.current_inputs = [CurrentInput("I1", 1.0, "S3", DEFAULT_REFERENCE_TEMPERATURE_C, "自检")]
    app.active_test_label = "S3"
    app.last_results_by_label = {"S1": object(), "S3": object()}
    app.last_correction_evaluations = []
    app.visible_spectrum_labels = {"S1", "S3"}
    app.highlighted_spectrum_labels = {"S3"}
    app._renumber_test_spectra()
    if [dataset.label for dataset in app.test_spectra] != ["S1", "S2"]:
        raise SystemExit("光谱删除后的编号重排自检失败。")
    if app.current_inputs[0].spectrum_label != "S2" or app.active_test_label != "S2":
        raise SystemExit("光谱编号重排后关联光谱未同步更新。")
    if set(app.last_results_by_label) != {"S1", "S2"}:
        raise SystemExit("光谱编号重排后计算结果索引未同步更新。")
    if app.visible_spectrum_labels != {"S1", "S2"} or app.highlighted_spectrum_labels != {"S2"}:
        raise SystemExit("光谱编号重排后显示/高亮状态未同步更新。")

    am15g = spectrum_from_dataframe(df, wave_col, g_col)
    am15d = spectrum_from_dataframe(df, wave_col, d_col)
    same = calculate_spd(am15g, am15g, reference_name="AM1.5G", test_name="AM1.5G")
    direct = calculate_spd(am15g, am15d, reference_name="AM1.5G", test_name="AM1.5D")
    print(f"AM1.5G self-test SPD={same.spd_percent:.6f}%, SPC={same.spc_percent:.3f}%, class={same.overall_class}")
    print(f"AM1.5D vs AM1.5G SPD={direct.spd_percent:.3f}%, SPC={direct.spc_percent:.3f}%, class={direct.overall_class}")
    print(f"Columns: wavelength={wave_col}, G={g_col}, D={d_col}")


def main() -> None:
    parser = argparse.ArgumentParser(description="光谱 SPD 自动计算器")
    parser.add_argument("--self-test", action="store_true", help="运行核心计算自检，不打开界面。")
    args = parser.parse_args()

    if args.self_test:
        run_self_test()
        return
    if tk is None:
        raise SystemExit("当前环境无法打开 Tkinter 图形界面。")
    root = tk.Tk()
    SpectralSPDApp(root)
    if os.environ.get("SPD_GUI_SMOKE") == "1":
        root.after(500, root.destroy)
    root.mainloop()


if __name__ == "__main__":
    main()
